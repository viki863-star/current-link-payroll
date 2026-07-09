import os, base64, re, math
from datetime import date, datetime
from flask import render_template, request, redirect, url_for, flash, current_app, send_file, session, jsonify
from markupsafe import Markup
from . import customer_bp

def _get_db():
    from ..database import open_db
    return open_db()

def _safe_rollback(db):
    try:
        db.rollback()
    except Exception:
        pass

def _safe_execute(db, sql, params=()):
    try:
        db.execute(sql, params)
        db.commit()
    except Exception:
        _safe_rollback(db)

def _open_db():
    from ..database import DatabaseAdapter, _connect_postgres, _connect_sqlite
    backend = current_app.config.get("DATABASE_BACKEND", "sqlite")
    if backend == "postgres":
        connection = _connect_postgres(current_app.config["DATABASE_URL"])
    else:
        connection = _connect_sqlite(current_app.config["DATABASE_PATH"])
    return DatabaseAdapter(connection, backend)

def _ensure_tables():
    backend = current_app.config.get("DATABASE_BACKEND", "postgres")
    db = _open_db()
    _safe_rollback(db)
    autoinc = "BIGSERIAL PRIMARY KEY" if backend == "postgres" else "INTEGER PRIMARY KEY AUTOINCREMENT"
    now = "NOW()" if backend == "postgres" else "datetime('now')"
    ignore = "ON CONFLICT DO NOTHING" if backend == "postgres" else "OR IGNORE"
    int_type = "INTEGER" if backend == "postgres" else "INTEGER"
    real_type = "DOUBLE PRECISION" if backend == "postgres" else "REAL"
    # Create each table individually so a failure in one doesn't affect others
    table_ddl = [
        f"""CREATE TABLE IF NOT EXISTS customers (
            id {autoinc}, customer_name TEXT NOT NULL, customer_code TEXT,
            contact_person TEXT, phone TEXT, email TEXT, address TEXT,
            trn TEXT, trade_license TEXT, credit_limit {real_type} DEFAULT 0,
            payment_terms TEXT, status TEXT NOT NULL DEFAULT 'active',
            notes TEXT, created_at TEXT NOT NULL DEFAULT ({now})
        )""",
        f"""CREATE TABLE IF NOT EXISTS customer_invoices (
            id {autoinc}, customer_id {int_type} NOT NULL, invoice_no TEXT,
            invoice_date TEXT NOT NULL, amount {real_type} NOT NULL,
            vat_percent {real_type} DEFAULT 5, vat_amount {real_type} DEFAULT 0,
            total_amount {real_type} NOT NULL,
            notes TEXT, created_at TEXT NOT NULL DEFAULT ({now})
        )""",
        f"""CREATE TABLE IF NOT EXISTS customer_payments (
            id {autoinc}, customer_id {int_type} NOT NULL, invoice_id {int_type},
            payment_date TEXT NOT NULL, amount {real_type} NOT NULL,
            payment_method TEXT DEFAULT 'Cash', reference_no TEXT,
            notes TEXT, created_at TEXT NOT NULL DEFAULT ({now})
        )""",
        f"""CREATE TABLE IF NOT EXISTS customer_contracts (
            id {autoinc}, customer_id {int_type} NOT NULL, contract_no TEXT,
            contract_date TEXT NOT NULL, start_date TEXT, end_date TEXT,
            contract_type TEXT DEFAULT 'rental', amount {real_type},
            status TEXT NOT NULL DEFAULT 'active', notes TEXT,
            created_at TEXT NOT NULL DEFAULT ({now})
        )""",
        f"""CREATE TABLE IF NOT EXISTS customer_quotations (
            id {autoinc}, customer_id {int_type} NOT NULL, quotation_no TEXT,
            quotation_date TEXT NOT NULL, amount {real_type},
            status TEXT NOT NULL DEFAULT 'pending', notes TEXT,
            created_at TEXT NOT NULL DEFAULT ({now})
        )""",
        f"""CREATE TABLE IF NOT EXISTS customer_lpos (
            id {autoinc}, customer_id {int_type} NOT NULL, lpo_no TEXT,
            lpo_date TEXT NOT NULL, amount {real_type},
            status TEXT NOT NULL DEFAULT 'pending', notes TEXT,
            created_at TEXT NOT NULL DEFAULT ({now})
        )""",
        f"""CREATE TABLE IF NOT EXISTS lpo_items (
            id {autoinc}, lpo_id {int_type} NOT NULL, description TEXT,
            quantity {real_type} NOT NULL DEFAULT 1, rate {real_type} NOT NULL DEFAULT 0,
            amount {real_type} NOT NULL DEFAULT 0, unit_type TEXT NOT NULL DEFAULT 'hour',
            vehicle_no TEXT, sort_order {int_type} DEFAULT 0
        )""",
        f"""CREATE TABLE IF NOT EXISTS customer_documents (
            id {autoinc}, customer_id {int_type} NOT NULL,
            doc_type TEXT, doc_name TEXT, file_data TEXT, file_type TEXT,
            expiry_date TEXT, created_at TEXT NOT NULL DEFAULT ({now})
        )""",
        f"""CREATE TABLE IF NOT EXISTS customer_invoice_items (
            id {autoinc}, invoice_id {int_type} NOT NULL, description TEXT,
            quantity {real_type} DEFAULT 1, rate {real_type} DEFAULT 0,
            amount {real_type} DEFAULT 0, sort_order {int_type} DEFAULT 0
        )""",
        f"""CREATE TABLE IF NOT EXISTS service_items (
            id {autoinc}, description TEXT NOT NULL UNIQUE,
            default_rate {real_type} DEFAULT 0, category TEXT DEFAULT '',
            created_at TEXT NOT NULL DEFAULT ({now})
        )""",
        f"""CREATE TABLE IF NOT EXISTS customer_service_orders (
            id {autoinc}, customer_id {int_type} NOT NULL, so_no TEXT,
            so_date TEXT NOT NULL, amount {real_type},
            status TEXT NOT NULL DEFAULT 'pending', notes TEXT,
            created_at TEXT NOT NULL DEFAULT ({now})
        )""",
        f"""CREATE TABLE IF NOT EXISTS tabreed_tripsheets (
            id {autoinc}, customer_id {int_type} NOT NULL,
            entry_date TEXT NOT NULL, time_in TEXT, time_out TEXT,
            meter_start {real_type} DEFAULT 0, meter_stop {real_type} DEFAULT 0,
            total_reading {real_type} DEFAULT 0, tanker_gln TEXT,
            trips {real_type} DEFAULT 1, tanker_reg TEXT, notes TEXT,
            created_at TEXT NOT NULL DEFAULT ({now})
        )""",
        f"""CREATE TABLE IF NOT EXISTS customer_so_items (
            id {autoinc}, so_id {int_type} NOT NULL, description TEXT,
            quantity {real_type} NOT NULL DEFAULT 1, rate {real_type} NOT NULL DEFAULT 0,
            amount {real_type} NOT NULL DEFAULT 0, unit_type TEXT NOT NULL DEFAULT 'hour',
            vehicle_no TEXT, sort_order {int_type} DEFAULT 0
        )""",
        f"""CREATE TABLE IF NOT EXISTS customer_quotation_items (
            id {autoinc}, quotation_id {int_type} NOT NULL, description TEXT,
            quantity {real_type} DEFAULT 1, rate {real_type} DEFAULT 0,
            amount {real_type} DEFAULT 0, unit TEXT DEFAULT 'hr',
            sort_order {int_type} DEFAULT 0
        )""",
        f"""CREATE TABLE IF NOT EXISTS quotation_sequence (last_number {int_type} DEFAULT 0)""",
        f"""CREATE TABLE IF NOT EXISTS invoice_sequence (last_number {int_type} DEFAULT 0)""",
        f"""CREATE TABLE IF NOT EXISTS customer_credit_notes (
            id {autoinc}, customer_id {int_type} NOT NULL,
            credit_note_no TEXT, credit_note_date TEXT NOT NULL,
            invoice_id {int_type}, amount {real_type} NOT NULL DEFAULT 0,
            vat_percent {real_type} DEFAULT 0, vat_amount {real_type} DEFAULT 0,
            total_amount {real_type} NOT NULL DEFAULT 0,
            reason TEXT, notes TEXT, created_at TEXT NOT NULL DEFAULT ({now})
        )""",
    ]
    for ddl in table_ddl:
        _safe_execute(db, ddl)
    # seed service_items from existing invoice items
    _safe_execute(db, f"""
        INSERT INTO service_items (description)
        SELECT DISTINCT TRIM(description) FROM customer_invoice_items
        WHERE description IS NOT NULL AND TRIM(description) != ''
        {ignore}
    """)
    # ALTER TABLE additions (best-effort on both backends)
    alter_ops = [
        ("customer_invoices", "lpo_no", "TEXT"),
        ("customer_invoices", "lpo_date", "TEXT"),
        ("customer_invoices", "so_no", "TEXT"),
        ("customer_invoices", "project_no", "TEXT"),
        ("customer_invoices", "invoice_template", "TEXT DEFAULT 'standard'"),
        ("customer_invoices", "discount", real_type + " DEFAULT 0"),
        ("customer_invoices", "ref_no", "TEXT"),
        ("customer_invoices", "service_order_no", "TEXT"),
        ("customer_invoices", "so_no", "TEXT"),
        ("customer_invoice_items", "vehicle_no", "TEXT"),
        ("customer_invoice_items", "capacity_gallon", "TEXT"),
        ("customer_invoice_items", "unit", "TEXT"),
        ("customer_invoice_items", "vat_percent_item", real_type),
        ("customer_invoice_items", "vat_amount_item", real_type),
        ("customer_invoice_items", "total_incl_vat", real_type),
        ("customer_invoice_items", "lpo_id", int_type),
        ("customer_invoice_items", "unit", "TEXT"),
        ("customer_lpos", "file_data", "TEXT"),
        ("customer_lpos", "file_type", "TEXT"),
        ("customer_lpos", "service_order_no", "TEXT"),
        ("lpo_items", "vehicle_no", "TEXT"),
        ("customer_so_items", "vehicle_no", "TEXT"),
        ("customer_quotations", "vat_percent", real_type + " DEFAULT 0"),
        ("customer_quotations", "vat_amount", real_type + " DEFAULT 0"),
        ("customer_quotations", "total_amount", real_type + " DEFAULT 0"),
        ("customer_quotations", "terms", "TEXT DEFAULT ''"),
        ("customer_quotations", "sub_total", real_type + " DEFAULT 0"),
        ("customer_quotations", "location", "TEXT"),
        ("customer_quotations", "contact_details", "TEXT"),
        ("company_profile", "logo_data", "TEXT"),
        ("company_profile", "logo_type", "TEXT"),
        ("company_profile", "theme_color", "TEXT DEFAULT '#0F2B52'"),
        ("company_profile", "bank_name", "TEXT"),
        ("company_profile", "bank_account_name", "TEXT"),
        ("company_profile", "bank_account_number", "TEXT"),
        ("company_profile", "iban", "TEXT"),
    ]
    for table, col, dtype in alter_ops:
        sql = f"ALTER TABLE {table} ADD COLUMN IF NOT EXISTS {col} {dtype}" if backend == "postgres" else f"ALTER TABLE {table} ADD COLUMN {col} {dtype}"
        _safe_execute(db, sql)
    _safe_execute(db, "ALTER TABLE customer_invoices DROP COLUMN IF EXISTS status" if backend == "postgres" else "ALTER TABLE customer_invoices DROP COLUMN IF EXISTS status")
    _safe_execute(db, "ALTER TABLE customer_invoices DROP COLUMN IF EXISTS paid" if backend == "postgres" else "ALTER TABLE customer_invoices DROP COLUMN IF EXISTS paid")
    _safe_execute(db, "INSERT INTO quotation_sequence (last_number) SELECT 0 WHERE NOT EXISTS (SELECT 1 FROM quotation_sequence)")
    _safe_execute(db, "INSERT INTO invoice_sequence (last_number) SELECT 0 WHERE NOT EXISTS (SELECT 1 FROM invoice_sequence)")
    _safe_execute(db, "ALTER TABLE customer_quotation_items ADD COLUMN IF NOT EXISTS unit TEXT DEFAULT 'hr'" if backend == "postgres" else "ALTER TABLE customer_quotation_items ADD COLUMN unit TEXT DEFAULT 'hr'")
    # Ensure customer_id columns exist on tables that may have been created without them
    for tbl in ["customer_invoices","customer_payments","customer_contracts","customer_quotations",
                "customer_lpos","customer_documents","customer_invoice_items",
                "customer_service_orders","customer_so_items","customer_quotation_items",
                "customer_credit_notes","tabreed_tripsheets"]:
        _safe_execute(db, f"ALTER TABLE {tbl} ADD COLUMN IF NOT EXISTS customer_id {int_type}" if backend == "postgres" else f"ALTER TABLE {tbl} ADD COLUMN customer_id {int_type}")
    # Ensure created_at columns exist on tables that may have been created without them
    for tbl in ["customer_documents","customer_invoice_items","customer_so_items",
                "customer_quotation_items","lpo_items","service_items"]:
        _safe_execute(db, f"ALTER TABLE {tbl} ADD COLUMN IF NOT EXISTS created_at TEXT" if backend == "postgres" else f"ALTER TABLE {tbl} ADD COLUMN created_at TEXT")
    db.close()

# ─── HELPERS ───

def _next_code(db):
    last = db.execute("SELECT customer_code FROM customers ORDER BY id DESC LIMIT 1").fetchone()
    if last and last["customer_code"]:
        m = re.search(r'(\d+)', last["customer_code"])
        n = int(m.group(1)) + 1 if m else 1
    else:
        n = 1
    return f"CUS-{n:04d}"

def _next_invoice_no(db):
    db.execute("UPDATE invoice_sequence SET last_number = last_number + 1")
    n = db.execute("SELECT last_number FROM invoice_sequence").fetchone()[0]
    return f"NS{n + 883}"

def _get_customer_or_404(cid):
    db = _get_db()
    c = db.execute("SELECT * FROM customers WHERE id=?", (cid,)).fetchone()
    db.close()
    if not c:
        flash("Customer not found.", "error")
        return None
    return c

# ─── DASHBOARD ───

@customer_bp.route("/")
def customer_dashboard():
    from ..routes import _touch_admin_workspace
    _touch_admin_workspace("customers")
    _ensure_tables()
    db = _get_db()
    total = db.execute("SELECT COUNT(*) FROM customers").fetchone()[0]
    active = db.execute("SELECT COUNT(*) FROM customers WHERE status='active'").fetchone()[0]
    total_receivable = db.execute("SELECT COALESCE(SUM(total_amount),0) FROM customer_invoices").fetchone()[0]
    total_cn = db.execute("SELECT COALESCE(SUM(total_amount),0) FROM customer_credit_notes").fetchone()[0]
    inv_count = db.execute("SELECT COUNT(*) FROM customer_invoices").fetchone()[0]
    paid_total = db.execute("SELECT COALESCE(SUM(amount),0) FROM customer_payments").fetchone()[0]
    recent = db.execute("""SELECT i.*, c.customer_name FROM customer_invoices i
        JOIN customers c ON i.customer_id=c.id ORDER BY i.created_at DESC LIMIT 8""").fetchall()
    recent_pmts = db.execute("""SELECT p.*, c.customer_name FROM customer_payments p
        JOIN customers c ON p.customer_id=c.id ORDER BY p.created_at DESC LIMIT 6""").fetchall()
    monthly = db.execute("""
        SELECT substr(invoice_date,1,7) AS mon,
               COUNT(*) AS cnt, COALESCE(SUM(total_amount),0) AS tot
        FROM customer_invoices GROUP BY mon ORDER BY mon DESC LIMIT 12
    """).fetchall()
    top_customers = db.execute("""
        SELECT c.id, c.customer_name, COUNT(i.id) AS inv_cnt,
               COALESCE(SUM(i.total_amount),0) AS total
        FROM customers c LEFT JOIN customer_invoices i ON i.customer_id=c.id
        GROUP BY c.id, c.customer_name ORDER BY total DESC LIMIT 5
    """).fetchall()
    db.close()
    return render_template("customer/dashboard.html",
        total=total, active=active,
        total_receivable=total_receivable, total_cn=total_cn,
        inv_count=inv_count, paid_total=paid_total,
        outstanding=total_receivable - paid_total - total_cn,
        recent_invoices=recent, recent_payments=recent_pmts,
        monthly_trend=monthly, top_customers=top_customers)

# ─── CUSTOMER CRUD ───

@customer_bp.route("/add", methods=["GET", "POST"])
def customer_add():
    _ensure_tables()
    db = _get_db()
    if request.method == "POST":
        name = request.form.get("customer_name", "").strip()
        if not name:
            flash("Customer name is required.", "error")
            code = _next_code(db)
            db.close()
            return render_template("customer/form.html", cus={}, code=code)
        code = request.form.get("customer_code", "").strip() or _next_code(db)
        c = db.execute("""INSERT INTO customers (customer_name,customer_code,contact_person,phone,email,address,trn,trade_license,credit_limit,payment_terms,status,notes)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?)""",
            (name, code, request.form.get("contact_person"), request.form.get("phone"),
             request.form.get("email"), request.form.get("address"), request.form.get("trn"),
             request.form.get("trade_license"), float(request.form.get("credit_limit", 0) or 0),
             request.form.get("payment_terms"), request.form.get("status", "active"), request.form.get("notes")))
        new_id = c.lastrowid
        db.commit()
        db.close()
        flash("Customer added.", "success")
        return redirect(url_for("customer.customer_profile", cid=new_id))
    code = _next_code(db)
    db.close()
    return render_template("customer/form.html", cus={}, code=code)

@customer_bp.route("/<int:cid>/edit", methods=["GET", "POST"])
def customer_edit(cid):
    _ensure_tables()
    c = _get_customer_or_404(cid)
    if not c: return redirect(url_for("customer.customer_dashboard"))
    if request.method == "POST":
        db = _get_db()
        db.execute("""UPDATE customers SET customer_name=?,contact_person=?,phone=?,email=?,address=?,trn=?,trade_license=?,credit_limit=?,payment_terms=?,status=?,notes=? WHERE id=?""",
            (request.form.get("customer_name"), request.form.get("contact_person"), request.form.get("phone"),
             request.form.get("email"), request.form.get("address"), request.form.get("trn"),
             request.form.get("trade_license"), float(request.form.get("credit_limit", 0) or 0),
             request.form.get("payment_terms"), request.form.get("status", "active"), request.form.get("notes"), cid))
        db.commit()
        db.close()
        flash("Customer updated.", "success")
        return redirect(url_for("customer.customer_profile", cid=cid))
    return render_template("customer/form.html", cus=c)

@customer_bp.route("/<int:cid>")
def customer_profile(cid):
    _ensure_tables()
    c = _get_customer_or_404(cid)
    if not c: return redirect(url_for("customer.customer_dashboard"))
    db = _get_db()
    tab = request.args.get("tab", "overview")
    invoices = db.execute("SELECT * FROM customer_invoices WHERE customer_id=? ORDER BY invoice_date DESC", (cid,)).fetchall()
    payments = db.execute("SELECT p.*, i.invoice_no FROM customer_payments p LEFT JOIN customer_invoices i ON p.invoice_id=i.id WHERE p.customer_id=? ORDER BY p.payment_date DESC", (cid,)).fetchall()
    contracts = db.execute("SELECT * FROM customer_contracts WHERE customer_id=? ORDER BY contract_date DESC", (cid,)).fetchall()
    quotations = db.execute("""
        SELECT q.*, (SELECT COUNT(*) FROM customer_quotation_items WHERE quotation_id=q.id) AS items_count
        FROM customer_quotations q WHERE q.customer_id=? ORDER BY q.quotation_date DESC
    """, (cid,)).fetchall()
    lpos = db.execute("SELECT * FROM customer_lpos WHERE customer_id=? ORDER BY lpo_date DESC", (cid,)).fetchall()
    service_orders = db.execute("SELECT * FROM customer_service_orders WHERE customer_id=? ORDER BY so_date DESC", (cid,)).fetchall()
    docs = db.execute("SELECT * FROM customer_documents WHERE customer_id=? ORDER BY created_at DESC", (cid,)).fetchall()
    credit_notes = db.execute("SELECT * FROM customer_credit_notes WHERE customer_id=? ORDER BY credit_note_date DESC", (cid,)).fetchall()
    tripsheets = db.execute("""
        SELECT * FROM tabreed_tripsheets
        WHERE customer_id=?
        ORDER BY entry_date DESC, id DESC
        LIMIT 100
    """, (cid,)).fetchall()
    total_inv = db.execute("SELECT COALESCE(SUM(total_amount),0) FROM customer_invoices WHERE customer_id=?", (cid,)).fetchone()[0]
    total_paid = db.execute("SELECT COALESCE(SUM(amount),0) FROM customer_payments WHERE customer_id=?", (cid,)).fetchone()[0]
    total_cn = db.execute("SELECT COALESCE(SUM(total_amount),0) FROM customer_credit_notes WHERE customer_id=?", (cid,)).fetchone()[0]
    balance = round(total_inv - total_paid - total_cn, 2)
    db.close()
    return render_template("customer/profile.html", c=c, active_tab=tab, invoices=invoices,
        payments=payments, contracts=contracts, quotations=quotations, lpos=lpos, service_orders=service_orders, docs=docs,
        credit_notes=credit_notes, tripsheets=tripsheets,
        total_inv=total_inv, total_paid=total_paid, total_cn=total_cn, balance=balance)

# ─── INVOICES ───

@customer_bp.route("/<int:cid>/invoice/add", methods=["GET", "POST"])
def customer_invoice_add(cid):
    _ensure_tables()
    c = _get_customer_or_404(cid)
    if not c: return redirect(url_for("customer.customer_dashboard"))
    is_nmdc = "nmdc" in (c["customer_name"] or "").lower()
    db = _get_db()
    next_no = _next_invoice_no(db)
    lpos = db.execute("SELECT id,lpo_no,lpo_date,amount FROM customer_lpos WHERE customer_id=? AND status!='closed' ORDER BY lpo_date DESC", (cid,)).fetchall()
    sos = db.execute("SELECT id,so_no,so_date,amount FROM customer_service_orders WHERE customer_id=? AND status!='closed' ORDER BY so_date DESC", (cid,)).fetchall()
    svc_items = db.execute("SELECT description FROM service_items ORDER BY description LIMIT 500").fetchall()
    if request.method == "POST":
        try:
            _safe_rollback(db)
            inv_date = request.form.get("invoice_date", date.today().isoformat())
            inv_no = request.form.get("invoice_no", "").strip() or next_no
            existing = db.execute("SELECT id FROM customer_invoices WHERE invoice_no=?", (inv_no,)).fetchone()
            if existing:
                flash(f"Invoice number '{inv_no}' already exists. Use a different number.", "error")
                db.close()
                tmpl = "customer/invoice_form_nmdc.html" if is_nmdc else "customer/invoice_form.html"
                return render_template(tmpl, c=c, inv={}, lpos=lpos, sos=sos, svc_items=svc_items, today=date.today().isoformat(), next_no=next_no)
            vat_pct = float(request.form.get("vat_percent", 5))
            so_id = request.form.get("so_id", "").strip()
            so_no = None
            if so_id:
                so_row = db.execute("SELECT so_no FROM customer_service_orders WHERE id=? AND customer_id=?", (so_id, cid)).fetchone()
                if so_row:
                    so_no = so_row["so_no"]
            notes = request.form.get("notes", "").strip()
            items = []
            sub_total = 0
            lpo_no = None
            lpo_date = None
            project_no = None

            if is_nmdc:
                main_desc = request.form.get("main_description", "").strip()
                nmdc_monthly_rate = float(request.form.get("monthly_rate", 0) or 0)
                eq_plants = request.form.getlist("eq_plant[]")
                eq_regs = request.form.getlist("eq_reg[]")
                eq_hours_list = request.form.getlist("eq_hours[]")
                eq_pf_list = request.form.getlist("eq_period_from[]")
                eq_pt_list = request.form.getlist("eq_period_to[]")
                eq_periods = []
                if main_desc:
                    items.append({"desc": main_desc, "qty": 1, "rate": nmdc_monthly_rate, "amt": 0, "unit": "month", "vehicle": "", "hours": 0})
                for i in range(len(eq_plants)):
                    plant = eq_plants[i].strip()
                    reg = eq_regs[i].strip() if i < len(eq_regs) else ""
                    hours = float(eq_hours_list[i]) if i < len(eq_hours_list) and eq_hours_list[i].strip() else 0
                    eq_pf = eq_pf_list[i].strip() if i < len(eq_pf_list) else ""
                    eq_pt = eq_pt_list[i].strip() if i < len(eq_pt_list) else ""
                    if hours > 0:
                        qyt = round(hours / 260, 3)
                        amt = round(qyt * nmdc_monthly_rate, 2)
                        items.append({"desc": reg, "qty": qyt, "rate": nmdc_monthly_rate, "amt": amt, "unit": "month", "vehicle": plant, "hours": hours})
                        sub_total += amt
                        eq_periods.append({"from": eq_pf, "to": eq_pt})
                if not items:
                    flash("At least one equipment with hours is required.", "error")
                    db.close()
                    return render_template("customer/invoice_form_nmdc.html", c=c, inv={}, lpos=lpos, sos=sos, svc_items=svc_items, today=date.today().isoformat(), next_no=next_no)
            else:
                lpo_id = request.form.get("lpo_id", "").strip()
                lpo_no = None
                lpo_date = request.form.get("lpo_date", "").strip() or None
                if lpo_id:
                    lpo_row = db.execute("SELECT lpo_no,lpo_date FROM customer_lpos WHERE id=? AND customer_id=?", (lpo_id, cid)).fetchone()
                    if lpo_row:
                        lpo_no = lpo_row["lpo_no"]
                        if not lpo_date:
                            lpo_date = lpo_row["lpo_date"]
                project_no = request.form.get("project_no", "").strip() or None
                descs = request.form.getlist("item_desc[]")
                qtys = request.form.getlist("item_qty[]")
                units = request.form.getlist("item_unit[]")
                rates = request.form.getlist("item_rate[]")
                vehicles = request.form.getlist("item_vehicle[]")
                for i in range(len(descs)):
                    desc = descs[i].strip()
                    qty = float(qtys[i]) if i < len(qtys) and qtys[i].strip() else 1
                    unit = units[i] if i < len(units) else "hour"
                    rate = float(rates[i]) if i < len(rates) and rates[i].strip() else 0
                    vehicle = vehicles[i].strip().upper() if i < len(vehicles) else ""
                    if desc or rate > 0:
                        amt = round(qty * rate, 2)
                        sub_total += amt
                        items.append({"desc": desc, "qty": qty, "rate": rate, "amt": amt, "unit": unit, "vehicle": vehicle, "hours": 0})
                if not items:
                    flash("At least one line item is required.", "error")
                    db.close()
                    return render_template("customer/invoice_form.html", c=c, inv={}, lpos=lpos, sos=sos, svc_items=svc_items, today=date.today().isoformat(), next_no=next_no)
            vat_amt = round(sub_total * vat_pct / 100, 2)
            total = round(sub_total + vat_amt, 2)
            if is_nmdc:
                import json
                nmdc_meta = {
                    "period_from": request.form.get("period_from", ""),
                    "period_to": request.form.get("period_to", ""),
                    "monthly_rate": float(request.form.get("monthly_rate", 0) or 0),
                    "month_label": request.form.get("month_label", ""),
                    "eq_periods": eq_periods,
                }
                notes = json.dumps(nmdc_meta) if not notes else json.dumps(nmdc_meta) + "\n" + notes
            c_inv = db.execute("""INSERT INTO customer_invoices (customer_id,invoice_no,invoice_date,amount,vat_percent,vat_amount,total_amount,lpo_no,lpo_date,so_no,project_no,notes)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?)""",
                (cid, inv_no, inv_date, sub_total, vat_pct, vat_amt, total, lpo_no, lpo_date, so_no, project_no, notes))
            inv_id = c_inv.lastrowid
            for idx, it in enumerate(items):
                db.execute("INSERT INTO customer_invoice_items (invoice_id,description,quantity,rate,amount,sort_order) VALUES (?,?,?,?,?,?)",
                    (inv_id, it["desc"], it["qty"], it["rate"], it["amt"], idx))
                if it["desc"]:
                    try:
                        db.execute("INSERT OR IGNORE INTO service_items (description, default_rate) VALUES (?,?)", (it["desc"], it["rate"]))
                    except Exception:
                        pass
            db.commit()
            db.close()
            flash(f"Invoice {inv_no} created.", "success")
            return redirect(url_for("customer.customer_profile", cid=cid, tab="invoices"))
        except Exception as e:
            db.rollback()
            db.close()
            import traceback
            current_app.logger.error("Invoice add failed: %s\n%s", e, traceback.format_exc())
            flash(f"Error creating invoice: {e}", "error")
            tmpl = "customer/invoice_form_nmdc.html" if is_nmdc else "customer/invoice_form.html"
            return render_template(tmpl, c=c, inv={}, lpos=lpos, sos=sos, svc_items=svc_items, today=date.today().isoformat(), next_no=next_no)
    db.close()
    tmpl = "customer/invoice_form_nmdc.html" if is_nmdc else "customer/invoice_form.html"
    return render_template(tmpl, c=c, inv={}, lpos=lpos, sos=sos, svc_items=svc_items, today=date.today().isoformat(), next_no=next_no)

@customer_bp.route("/service-items/search")
def service_items_search():
    _ensure_tables()
    q = request.args.get("q", "").strip()
    db = _get_db()
    if q:
        items = db.execute(
            "SELECT id, description, default_rate FROM service_items WHERE description LIKE ? ORDER BY description LIMIT 20",
            (f"%{q}%",)
        ).fetchall()
    else:
        items = db.execute(
            "SELECT id, description, default_rate FROM service_items ORDER BY description LIMIT 50"
        ).fetchall()
    db.close()
    return jsonify([{"id": r["id"], "description": r["description"], "rate": r["default_rate"]} for r in items])

# ─── NOUROL INVOICE ───

@customer_bp.route("/<int:cid>/invoice/<int:iid>/edit", methods=["GET", "POST"])
def customer_invoice_edit(cid, iid):
    _ensure_tables()
    c = _get_customer_or_404(cid)
    if not c: return redirect(url_for("customer.customer_dashboard"))
    db = _get_db()
    inv = db.execute("SELECT * FROM customer_invoices WHERE id=? AND customer_id=?", (iid, cid)).fetchone()
    items = db.execute("SELECT * FROM customer_invoice_items WHERE invoice_id=? ORDER BY sort_order", (iid,)).fetchall()
    lpos = db.execute("SELECT id,lpo_no,lpo_date,amount FROM customer_lpos WHERE customer_id=? AND status!='closed' ORDER BY lpo_date DESC", (cid,)).fetchall()
    sos = db.execute("SELECT id,so_no,so_date,amount FROM customer_service_orders WHERE customer_id=? AND status!='closed' ORDER BY so_date DESC", (cid,)).fetchall()
    svc_items = db.execute("SELECT description FROM service_items ORDER BY description LIMIT 500").fetchall()
    if not inv:
        db.close()
        flash("Invoice not found.", "error")
        return redirect(url_for("customer.customer_profile", cid=cid, tab="invoices"))
    selected_lpo_id = None
    if inv["lpo_no"]:
        row = db.execute("SELECT id FROM customer_lpos WHERE lpo_no=? AND customer_id=?", (inv["lpo_no"], cid)).fetchone()
        if row:
            selected_lpo_id = row["id"]
    selected_so_id = None
    try:
        if inv["so_no"]:
            row = db.execute("SELECT id FROM customer_service_orders WHERE so_no=? AND customer_id=?", (inv["so_no"], cid)).fetchone()
            if row:
                selected_so_id = row["id"]
    except (KeyError, AttributeError):
        pass
    is_nmdc_edit = "nmdc" in (c["customer_name"] or "").lower()
    if request.method == "POST":
        try:
            _safe_rollback(db)
            inv_date = request.form.get("invoice_date", date.today().isoformat())
            inv_no = request.form.get("invoice_no", "").strip() or inv["invoice_no"]
            dup = db.execute("SELECT id FROM customer_invoices WHERE invoice_no=? AND id!=?", (inv_no, iid)).fetchone()
            if dup:
                flash(f"Invoice number '{inv_no}' already in use.", "error")
                db.close()
                tmpl_e = "customer/invoice_form_nmdc.html" if is_nmdc_edit else "customer/invoice_form.html"
                return render_template(tmpl_e, c=c, inv=inv, items=items, lpos=lpos, sos=sos, svc_items=svc_items, today=date.today().isoformat(), edit=True, selected_lpo_id=selected_lpo_id, selected_so_id=selected_so_id)
            vat_pct = float(request.form.get("vat_percent", 5))
            so_id = request.form.get("so_id", "").strip()
            so_no = None
            if so_id:
                so_row = db.execute("SELECT so_no FROM customer_service_orders WHERE id=? AND customer_id=?", (so_id, cid)).fetchone()
                if so_row:
                    so_no = so_row["so_no"]
            notes = request.form.get("notes", "").strip()
            new_items = []
            sub_total = 0
            lpo_no = None
            lpo_date = None
            project_no = None
            if is_nmdc_edit:
                main_desc = request.form.get("main_description", "").strip()
                nmdc_monthly_rate = float(request.form.get("monthly_rate", 0) or 0)
                eq_plants = request.form.getlist("eq_plant[]")
                eq_regs = request.form.getlist("eq_reg[]")
                eq_hours_list = request.form.getlist("eq_hours[]")
                eq_pf_list = request.form.getlist("eq_period_from[]")
                eq_pt_list = request.form.getlist("eq_period_to[]")
                eq_periods = []
                if main_desc:
                    new_items.append({"desc": main_desc, "qty": 1, "rate": nmdc_monthly_rate, "amt": 0, "unit": "month", "vehicle": "", "hours": 0})
                for i in range(len(eq_plants)):
                    plant = eq_plants[i].strip()
                    reg = eq_regs[i].strip() if i < len(eq_regs) else ""
                    hours = float(eq_hours_list[i]) if i < len(eq_hours_list) and eq_hours_list[i].strip() else 0
                    eq_pf = eq_pf_list[i].strip() if i < len(eq_pf_list) else ""
                    eq_pt = eq_pt_list[i].strip() if i < len(eq_pt_list) else ""
                    if hours > 0:
                        qyt = round(hours / 260, 3)
                        amt = round(qyt * nmdc_monthly_rate, 2)
                        new_items.append({"desc": reg, "qty": qyt, "rate": nmdc_monthly_rate, "amt": amt, "unit": "month", "vehicle": plant, "hours": hours})
                        sub_total += amt
                        eq_periods.append({"from": eq_pf, "to": eq_pt})
                if not new_items:
                    flash("At least one equipment with hours is required.", "error")
                    db.close()
                    return render_template("customer/invoice_form_nmdc.html", c=c, inv=inv, items=items, lpos=lpos, sos=sos, svc_items=svc_items, today=date.today().isoformat(), edit=True, selected_lpo_id=selected_lpo_id, selected_so_id=selected_so_id)
                import json
                nmdc_meta_e = {
                    "period_from": request.form.get("period_from", ""),
                    "period_to": request.form.get("period_to", ""),
                    "monthly_rate": nmdc_monthly_rate,
                    "month_label": request.form.get("month_label", ""),
                    "eq_periods": eq_periods,
                }
                notes = json.dumps(nmdc_meta_e) if not notes else json.dumps(nmdc_meta_e) + "\n" + notes
            else:
                lpo_id = request.form.get("lpo_id", "").strip()
                lpo_no = None
                lpo_date = request.form.get("lpo_date", "").strip() or None
                if lpo_id:
                    lpo_row = db.execute("SELECT lpo_no,lpo_date FROM customer_lpos WHERE id=? AND customer_id=?", (lpo_id, cid)).fetchone()
                    if lpo_row:
                        lpo_no = lpo_row["lpo_no"]
                        if not lpo_date:
                            lpo_date = lpo_row["lpo_date"]
                project_no = request.form.get("project_no", "").strip() or None
                descs = request.form.getlist("item_desc[]")
                qtys = request.form.getlist("item_qty[]")
                units = request.form.getlist("item_unit[]")
                rates = request.form.getlist("item_rate[]")
                vehicles = request.form.getlist("item_vehicle[]")
                for i in range(len(descs)):
                    desc = descs[i].strip()
                    qty = float(qtys[i]) if i < len(qtys) and qtys[i].strip() else 1
                    unit = units[i] if i < len(units) else "hour"
                    rate = float(rates[i]) if i < len(rates) and rates[i].strip() else 0
                    vehicle = vehicles[i].strip().upper() if i < len(vehicles) else ""
                    if desc or rate > 0:
                        amt = round(qty * rate, 2)
                        sub_total += amt
                        new_items.append({"desc": desc, "qty": qty, "rate": rate, "amt": amt, "unit": unit, "vehicle": vehicle, "hours": 0})
            if not new_items:
                flash("At least one line item is required.", "error")
                db.close()
                tmpl_e = "customer/invoice_form_nmdc.html" if is_nmdc_edit else "customer/invoice_form.html"
                return render_template(tmpl_e, c=c, inv=inv, items=items, lpos=lpos, sos=sos, svc_items=svc_items, today=date.today().isoformat(), edit=True, selected_lpo_id=selected_lpo_id, selected_so_id=selected_so_id)
            vat_amt = round(sub_total * vat_pct / 100, 2)
            total = round(sub_total + vat_amt, 2)
            db.execute("""UPDATE customer_invoices SET invoice_no=?,invoice_date=?,amount=?,vat_percent=?,vat_amount=?,total_amount=?,lpo_no=?,lpo_date=?,so_no=?,project_no=?,notes=? WHERE id=?""",
                (inv_no, inv_date, sub_total, vat_pct, vat_amt, total, lpo_no, lpo_date, so_no, project_no, notes, iid))
            db.execute("DELETE FROM customer_invoice_items WHERE invoice_id=?", (iid,))
            for idx, it in enumerate(new_items):
                db.execute("INSERT INTO customer_invoice_items (invoice_id,description,quantity,rate,amount,sort_order) VALUES (?,?,?,?,?,?)",
                    (iid, it["desc"], it["qty"], it["rate"], it["amt"], idx))
                if it["desc"]:
                    try:
                        db.execute("INSERT OR IGNORE INTO service_items (description, default_rate) VALUES (?,?)", (it["desc"], it["rate"]))
                    except Exception:
                        pass
            db.commit()
            db.close()
            flash(f"Invoice {inv_no} updated.", "success")
            return redirect(url_for("customer.customer_profile", cid=cid, tab="invoices"))
        except Exception as e:
            db.rollback()
            db.close()
            import traceback
            current_app.logger.error("Invoice edit failed: %s\n%s", e, traceback.format_exc())
            flash(f"Error updating invoice: {e}", "error")
            tmpl_e = "customer/invoice_form_nmdc.html" if is_nmdc_edit else "customer/invoice_form.html"
            return render_template(tmpl_e, c=c, inv=inv, items=items, lpos=lpos, sos=sos, svc_items=svc_items, today=date.today().isoformat(), edit=True, selected_lpo_id=selected_lpo_id, selected_so_id=selected_so_id)
    db.close()
    nmdc_meta = {}
    display_notes = inv.get("notes", "") or ""
    if is_nmdc_edit:
        try:
            import json
            nmdc_meta = json.loads(inv.get("notes", "{}"))
        except Exception:
            nmdc_meta = {}
        if display_notes:
            lines = display_notes.split("\n", 1)
            if lines and lines[0].strip().startswith("{"):
                try:
                    json.loads(lines[0])
                    display_notes = lines[1].strip() if len(lines) > 1 else ""
                except Exception:
                    pass
    tmpl_e = "customer/invoice_form_nmdc.html" if is_nmdc_edit else "customer/invoice_form.html"
    return render_template(tmpl_e, c=c, inv=inv, items=items, lpos=lpos, sos=sos, svc_items=svc_items, today=date.today().isoformat(), edit=True, selected_lpo_id=selected_lpo_id, selected_so_id=selected_so_id, nmdc_meta=nmdc_meta, display_notes=display_notes)

@customer_bp.route("/<int:cid>/invoice/<int:iid>")
def customer_invoice_view(cid, iid):
    _ensure_tables()
    c = _get_customer_or_404(cid)
    if not c: return redirect(url_for("customer.customer_dashboard"))
    db = _get_db()
    inv = db.execute("SELECT * FROM customer_invoices WHERE id=? AND customer_id=?", (iid, cid)).fetchone()
    if not inv:
        db.close()
        flash("Invoice not found.", "error")
        return redirect(url_for("customer.customer_profile", cid=cid, tab="invoices"))
    items = db.execute("SELECT * FROM customer_invoice_items WHERE invoice_id=? ORDER BY sort_order", (iid,)).fetchall()
    company = db.execute("SELECT * FROM company_profile LIMIT 1").fetchone()
    so_date_val = None
    if inv.get("so_no"):
        so_row = db.execute("SELECT so_date FROM customer_service_orders WHERE so_no=? AND customer_id=?", (inv["so_no"], cid)).fetchone()
        if so_row:
            so_date_val = so_row["so_date"]
    db.close()
    try:
        tmpl_t = inv["invoice_template"] or "standard"
    except (IndexError, KeyError):
        tmpl_t = "standard"
    is_nmdc = "nmdc" in (c["customer_name"] or "").lower()
    nmdc_meta = {}
    if is_nmdc:
        tmpl = "customer/invoice_view_nmdc.html"
        try:
            import json
            nmdc_meta = json.loads(inv.get("notes", "{}"))
        except Exception:
            nmdc_meta = {}
    else:
        tmpl = "customer/invoice_view.html"
    sum_taxable = sum(float(it.get("amount") or 0) for it in items)
    sum_vat = sum(float(it.get("vat_amount_item") or 0) for it in items)
    sum_total = sum(float(it.get("total_incl_vat") or (float(it.get("amount") or 0) + float(it.get("vat_amount_item") or 0))) for it in items)
    display_notes = inv.get("notes", "") or ""
    if is_nmdc and display_notes:
        import json
        lines = display_notes.split("\n", 1)
        if lines and lines[0].strip().startswith("{"):
            try:
                json.loads(lines[0])
                display_notes = lines[1].strip() if len(lines) > 1 else ""
            except Exception:
                pass
    return render_template(tmpl, c=c, inv=inv, items=items, company=company, nmdc_meta=nmdc_meta, sum_taxable=sum_taxable, sum_vat=sum_vat, sum_total=sum_total, display_notes=display_notes, so_date=so_date_val)

@customer_bp.route("/<int:cid>/invoice/<int:iid>/pdf")
def customer_invoice_pdf(cid, iid):
    import tempfile
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
    from io import BytesIO

    _logo_tmp_files = []
    _ensure_tables()
    db = _get_db()
    c = db.execute("SELECT * FROM customers WHERE id=?", (cid,)).fetchone()
    inv = db.execute("SELECT * FROM customer_invoices WHERE id=? AND customer_id=?", (iid, cid)).fetchone()
    items = db.execute("SELECT * FROM customer_invoice_items WHERE invoice_id=? ORDER BY sort_order", (iid,)).fetchall()
    company = db.execute("SELECT * FROM company_profile LIMIT 1").fetchone()
    pdf_so_date = None
    if inv.get("so_no"):
        so_r = db.execute("SELECT so_date FROM customer_service_orders WHERE so_no=? AND customer_id=?", (inv["so_no"], cid)).fetchone()
        if so_r:
            pdf_so_date = so_r["so_date"]
    db.close()
    if not c or not inv:
        flash("Invoice not found.", "error")
        return redirect(url_for("customer.customer_dashboard"))

    is_nmdc = "nmdc" in (c["customer_name"] or "").lower()

    buf = BytesIO()
    LM, RM, TM, BM = 14*mm, 14*mm, 10*mm, 8*mm
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=LM, rightMargin=RM, topMargin=TM, bottomMargin=BM)
    W = A4[0] - LM - RM

    tc = company["theme_color"] or "#1a3a5c" if company else "#1a3a5c"
    try: TH = colors.HexColor(tc)
    except: TH = colors.HexColor("#1a3a5c")
    WH = colors.white; BG = colors.HexColor("#f8fafc")
    C3 = colors.HexColor("#e2e8f0"); C4 = colors.HexColor("#0f172a")
    C5 = colors.HexColor("#64738b"); C6 = colors.HexColor("#dc2626")
    DH = colors.HexColor("#1e293b")

    cn = (company["company_name"] if company else "CURRENT LINK TRANSPORT AND GENERAL CONTRACTING") or "CURRENT LINK TRANSPORT AND GENERAL CONTRACTING"
    c_addr = (company["address"] or "") if company else ""
    c_ph = (company["phone_number"] or "") if company else ""
    c_em = (company["email"] or "") if company else ""
    c_trn = company["trn_no"] or "—" if company else "—"

    def S(name, **kw):
        kw.setdefault("fontSize", 8)
        kw.setdefault("leading", 12)
        return ParagraphStyle(name, **kw)

    def L(t, **kw):
        kw.setdefault("textColor", C5)
        return Paragraph(str(t), S("_L", **kw))

    def V(t, **kw):
        kw.setdefault("fontName", "Helvetica-Bold")
        kw.setdefault("textColor", C4)
        kw.setdefault("fontSize", 8.5)
        return Paragraph(str(t), S("_V", **kw))

    def C(t, **kw):
        kw.setdefault("alignment", TA_CENTER)
        return Paragraph(str(t), S("_C", **kw))

    def R(t, **kw):
        kw.setdefault("alignment", TA_RIGHT)
        return Paragraph(str(t), S("_R", **kw))

    def RB(t, **kw):
        kw.setdefault("fontName", "Helvetica-Bold")
        kw.setdefault("alignment", TA_RIGHT)
        return Paragraph(f"<b>{t}</b>", S("_RB", **kw))

    safe = lambda v, d="—": str(v) if v else d
    els = []
    inv_no = inv["invoice_no"] or "—"
    inv_dt = inv["invoice_date"] or "—"

    NMDC_CONV_HOURS = 260

    # ═══════════════════════════════════
    # 1. HEADER (matching web view)
    # ═══════════════════════════════════
    logo = None; LW = 0
    if company and company["logo_data"]:
        try:
            lb = base64.b64decode(company["logo_data"])
            f = tempfile.NamedTemporaryFile(delete=False, suffix=".png")
            f.write(lb); f.close()
            _logo_tmp_files.append(f.name)
        except: pass

    # Company info lines
    ci_lines = []
    c_contact = []
    if c_ph: c_contact.append(f"Phone: {c_ph}")
    if c_em: c_contact.append(f"Email: {c_em}")
    if c_contact: ci_lines.append('<font size=6 color="#64748b">' + ' &middot; '.join(c_contact) + '</font>')
    ci_lines.append(f"<font size=6 color='#64748b'><b>TRN: {c_trn}</b></font>")
    ci_html = f"<font size=11><b>{cn}</b></font><br/>" + "<br/>".join(ci_lines)
    co_p = Paragraph(ci_html, S("CO", fontSize=11, fontName="Helvetica-Bold", textColor=TH, leading=13))

    # Measure the company text height, then size logo to match
    logo_w = 0
    if _logo_tmp_files:
        try:
            from PIL import Image as PILImage
            with PILImage.open(_logo_tmp_files[-1]) as img:
                ow, oh = img.size
            from reportlab.pdfgen import canvas as rlcanvas
            from io import BytesIO
            tmp_buf = BytesIO()
            tmp_c = rlcanvas.Canvas(tmp_buf)
            ci_width = W*0.65 - 4*mm
            co_p.wrapOn(tmp_c, ci_width, 1000)
            text_h = co_p.height
            tmp_c.save()
            target_h = text_h
            ratio = target_h / oh
            logo_w = int(ow * ratio)
            logo_h = int(target_h)
            logo = Image(_logo_tmp_files[-1], width=logo_w, height=logo_h)
            LW = logo_w
        except:
            pass

    if logo:
        lh = Table([[logo, Spacer(1, 4*mm), co_p]], colWidths=[LW if LW else 0, 4*mm, W*0.65 - (LW if LW else 0) - 4*mm])
        lh.setStyle(TableStyle([("VALIGN",(0,0),(-1,-1),"TOP"),("LEFTPADDING",(0,0),(-1,-1),0),("RIGHTPADDING",(0,0),(-1,-1),0)]))
    else:
        lh = co_p

    rh = Paragraph(
        f"<b>TAX INVOICE</b><br/>"
        f"<font size=6 color='#64748b'># {inv_no}<br/>{inv_dt}</font>",
        S("TI", fontSize=13, fontName="Helvetica-Bold", textColor=TH, leading=16, alignment=TA_RIGHT))

    ht = Table([[lh, rh]], colWidths=[W*0.65, W*0.35])
    ht.setStyle(TableStyle([("VALIGN",(0,0),(-1,-1),"TOP"),("LEFTPADDING",(0,0),(-1,-1),0),("RIGHTPADDING",(0,0),(-1,-1),0)]))
    els.append(ht)

    # Bottom border line
    bl = Table([[""]], colWidths=[W], rowHeights=[1.5])
    bl.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,-1),TH),("LEFTPADDING",(0,0),(-1,-1),0),("RIGHTPADDING",(0,0),(-1,-1),0)]))
    els.append(bl)
    els.append(Spacer(1, 3*mm))

    # ═══════════════════════════════════
    # 2. BILL TO / INVOICE INFO (matching web)
    # ═══════════════════════════════════
    def card(title, pairs):
        cw = W*0.50
        r = [[
            Paragraph(f"<b>{title}</b>", S("_ch", fontSize=6, fontName="Helvetica-Bold", textColor=C5, leading=7.5)),
            Paragraph("", S("_cs", fontSize=1.5, leading=1.5)),
        ]]
        for a, b in pairs:
            r.append([
                Paragraph(a, S("_cl", fontSize=6.5, textColor=C5, leading=8.5)),
                Paragraph(f"{b}", S("_cv", fontSize=7, fontName="Helvetica-Bold", textColor=C4, leading=9)),
            ])
        t = Table(r, colWidths=[cw*0.28, cw*0.72])
        t.setStyle(TableStyle([
            ("VALIGN",(0,0),(-1,-1),"TOP"),
            ("TOPPADDING",(0,0),(-1,-1),1.5), ("BOTTOMPADDING",(0,0),(-1,-1),1.5),
            ("LEFTPADDING",(0,0),(-1,-1),6), ("RIGHTPADDING",(0,0),(-1,-1),6),
            ("BOX",(0,0),(-1,-1),0.5,colors.HexColor("#e2e8f0")),
        ]))
        return t

    bd = [("Customer", safe(c["customer_name"])), ("TRN", safe(c["trn"]))]
    if c["phone"]: bd.append(("Phone", c["phone"]))
    if c["email"]: bd.append(("Email", c["email"]))
    if c["address"]: bd.append(("Address", c["address"]))
    id_ = [("Invoice #", inv_no), ("Date", inv_dt)]
    if inv.get("so_no"): id_.append(("SO No.", inv["so_no"]))
    if pdf_so_date: id_.append(("SO Date", pdf_so_date))
    if inv.get("lpo_no"): id_.append(("LPO No.", inv["lpo_no"]))
    if inv.get("lpo_date"): id_.append(("LPO Date", inv["lpo_date"]))
    try:
        if inv.get("project_no"): id_.append(("Project No.", inv["project_no"]))
    except (IndexError, KeyError):
        pass
    try:
        if inv.get("ref_no"): id_.append(("Ref No.", inv["ref_no"]))
    except (IndexError, KeyError):
        pass

    iw = Table([[card("BILL TO", bd), Spacer(1, 2*mm), card("INVOICE INFO", id_)]], colWidths=[W*0.50, 2*mm, W*0.50])
    iw.setStyle(TableStyle([("VALIGN",(0,0),(-1,-1),"TOP"),("LEFTPADDING",(0,0),(-1,-1),0),("RIGHTPADDING",(0,0),(-1,-1),0)]))
    els.append(iw)
    els.append(Spacer(1, 2*mm))

    # ═══════════════════════════════════
    # 3. ITEMS TABLE — auto-fit on one page
    # ═══════════════════════════════════
    # Estimate fixed content height (in points)
    fixed_pt = 22*mm + 1.5*mm + 22*mm + 2*mm + 2*mm + 12*mm + 2*mm + 5*mm + (3*mm if inv["notes"] else 0) + 6*mm + 5*mm + 5*mm
    avail_pt = A4[1] - TM - BM - fixed_pt
    num_rows = len(items) + (1 if is_nmdc else 0)
    fs = 7.0
    if num_rows > 0:
        target = avail_pt / (num_rows + 1)
        fs = max(4.0, min(7.0, target / 2.6))
    ldr = fs * 1.15
    pad_t = max(1.0, fs * 0.25)
    pad_b = max(1.0, fs * 0.25)

    def _pc(t, **kw):
        kw.setdefault("fontSize", fs)
        kw.setdefault("leading", ldr)
        return Paragraph(str(t), S("_pc", **kw))

    sub = inv["amount"] or 0; vat = inv["vat_amount"] or 0; tot = inv["total_amount"] or 0; vp = inv["vat_percent"] or 0

    nmdc_eq_periods = []
    if is_nmdc:
        nmdc_main_desc = items[0]["description"] if items else ""
        try:
            import json
            nmdc_meta = json.loads(inv.get("notes", "{}"))
        except Exception:
            nmdc_meta = {}
        nmdc_pf = nmdc_meta.get("period_from", "") or ""
        nmdc_pt = nmdc_meta.get("period_to", "") or ""
        nmdc_mr = nmdc_meta.get("monthly_rate", 0) or 0
        nmdc_ml = nmdc_meta.get("month_label", "") or ""
        nmdc_eq_periods = nmdc_meta.get("eq_periods", []) or []

    cw = [9*mm, 44*mm, 16*mm, 12*mm, 15*mm, 18*mm, 13*mm, 16*mm, 25*mm]
    hdr = [
        Paragraph("<b>#</b>", S("_h0", fontSize=fs, fontName="Helvetica-Bold", textColor=WH, alignment=TA_CENTER, leading=ldr)),
        Paragraph("<b>Description</b>", S("_h1", fontSize=fs, fontName="Helvetica-Bold", textColor=WH, leading=ldr)),
        Paragraph("<b>Qty</b>", S("_h2", fontSize=fs, fontName="Helvetica-Bold", textColor=WH, alignment=TA_CENTER, leading=ldr)),
        Paragraph("<b>Unit</b>", S("_hu", fontSize=fs, fontName="Helvetica-Bold", textColor=WH, alignment=TA_CENTER, leading=ldr)),
        Paragraph("<b>Unit Price</b>", S("_h3", fontSize=fs, fontName="Helvetica-Bold", textColor=WH, alignment=TA_RIGHT, leading=ldr)),
        Paragraph("<b>Taxable Amount</b>", S("_h4", fontSize=fs, fontName="Helvetica-Bold", textColor=WH, alignment=TA_RIGHT, leading=ldr)),
        Paragraph("<b>VAT %</b>", S("_h5", fontSize=fs, fontName="Helvetica-Bold", textColor=WH, alignment=TA_CENTER, leading=ldr)),
        Paragraph("<b>VAT Amount</b>", S("_h6", fontSize=fs, fontName="Helvetica-Bold", textColor=WH, alignment=TA_RIGHT, leading=ldr)),
        Paragraph("<b>Total Amount<br/>(Including VAT)</b>", S("_h7", fontSize=fs, fontName="Helvetica-Bold", textColor=WH, alignment=TA_RIGHT, leading=ldr)),
    ]
    rws = [hdr]
    if is_nmdc:
        # Row 1: Main description
        rws.append([
            _pc("1", alignment=TA_CENTER, fontName="Helvetica-Bold"),
            _pc(nmdc_main_desc or "—", fontSize=fs, leading=ldr*0.95),
            _pc("—", alignment=TA_CENTER),
            _pc("—", alignment=TA_CENTER),
            _pc("—", alignment=TA_RIGHT),
            _pc("—", alignment=TA_RIGHT),
            _pc("—", alignment=TA_CENTER),
            _pc("—", alignment=TA_RIGHT),
            _pc("—", alignment=TA_RIGHT),
        ])
    table_items = items[1:] if is_nmdc else items
    for idx, it in enumerate(table_items):
        vp_item = float(it.get("vat_percent_item") or inv["vat_percent"] or 5)
        amt = float(it.get("amount") or 0)
        va_item = float(it.get("vat_amount_item") or (amt * vp_item / 100))
        ti_item = float(it.get("total_incl_vat") or (amt + va_item))
        eq_p = nmdc_eq_periods[idx] if is_nmdc and idx < len(nmdc_eq_periods) else {}
        eq_period_text = ""
        if is_nmdc and (eq_p.get("from") or eq_p.get("to")):
            eq_period_text += f" | Period: {eq_p.get('from','')} to {eq_p.get('to','')}"
        cap = it.get("capacity_gallon")
        eq_hours = f"<br/><font size=1 color='#94a3b8'>Hours: {float(cap):,.2f}</font>" if is_nmdc and cap and float(cap) > 0 else ""
        desc_html = (it.get("description") or "—")
        if is_nmdc:
            parts = []
            if it.get("vehicle_no"): parts.append(f"<b>Plant No:</b> {it['vehicle_no']}")
            if desc_html != "—": parts.append(f"<b>Reg#</b> {desc_html}")
            plant_reg = " | ".join(parts)
            desc_html = plant_reg + eq_period_text + eq_hours
        rws.append([
            _pc(str(idx + (2 if is_nmdc else 1)), alignment=TA_CENTER, fontName="Helvetica-Bold"),
            _pc(desc_html, fontSize=fs, leading=ldr*0.9),
            _pc(f"{float(it.get('quantity') or 0):,.3f}", alignment=TA_CENTER),
            _pc((it.get('unit') or 'mo'), alignment=TA_CENTER),
            _pc(f"{float(it.get('rate') or 0):,.2f}", alignment=TA_RIGHT),
            _pc(f"{amt:,.2f}", alignment=TA_RIGHT),
            _pc(f"{vp_item:.2f}%", alignment=TA_CENTER),
            _pc(f"{va_item:,.2f}", alignment=TA_RIGHT, textColor=C6),
            Paragraph(f"<b>{ti_item:,.2f}</b>", S("_b", fontSize=fs, fontName="Helvetica-Bold", alignment=TA_RIGHT, leading=ldr)),
        ])

    itt = Table(rws, colWidths=cw, repeatRows=1)
    itt.setStyle(TableStyle([
        ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
        ("BACKGROUND",(0,0),(-1,0),DH), ("TEXTCOLOR",(0,0),(-1,0),WH),
        ("BOX",(0,0),(-1,-1),0.5,C3),
        ("INNERGRID",(0,0),(-1,-1),0.3,C3),
        ("TOPPADDING",(0,0),(-1,-1),pad_t), ("BOTTOMPADDING",(0,0),(-1,-1),pad_b),
        ("LEFTPADDING",(0,0),(-1,-1),6), ("RIGHTPADDING",(0,0),(-1,-1),6),
        ("ROWBACKGROUNDS",(0,1),(-1,-1),[WH, colors.HexColor("#f8fafc")]),
    ]))
    els.append(itt)

    # ═══════════════════════════════════
    # 4. TOTALS (matching web)
    # ═══════════════════════════════════
    tw = 80*mm
    trows = [
        [Paragraph("Sub Total", S("_st", fontSize=8, textColor=C5, leading=11)),
         Paragraph(f"<b>AED {sub:,.2f}</b>", S("_stv", fontSize=8, fontName="Helvetica-Bold", textColor=C4, leading=11, alignment=TA_RIGHT))],
        [Paragraph(f"VAT @ {vp:.0f}%", S("_vt", fontSize=8, textColor=C5, leading=11)),
         Paragraph(f"<b>AED {vat:,.2f}</b>", S("_vtv", fontSize=8, fontName="Helvetica-Bold", textColor=C6, leading=11, alignment=TA_RIGHT))],
        [Paragraph("<b>Total Due</b>", S("_td", fontSize=10, fontName="Helvetica-Bold", textColor=C4, leading=13)),
         Paragraph(f"<b>AED {tot:,.2f}</b>", S("_tdv", fontSize=11, fontName="Helvetica-Bold", textColor=TH, leading=14, alignment=TA_RIGHT))],
    ]
    tt = Table(trows, colWidths=[tw*0.45, tw*0.55])
    tt.setStyle(TableStyle([
        ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
        ("TOPPADDING",(0,0),(-1,-1),2), ("BOTTOMPADDING",(0,0),(-1,-1),2),
        ("LEFTPADDING",(0,0),(-1,-1),8), ("RIGHTPADDING",(0,0),(-1,-1),8),
        ("BOX",(0,0),(-1,-1),0.5,colors.HexColor("#e2e8f0")),
        ("LINEABOVE",(0,2),(-1,2),1.5,TH),
    ]))

    ft = Table([["", tt]], colWidths=[W - tw, tw])
    ft.setStyle(TableStyle([("VALIGN",(0,0),(-1,-1),"TOP"),("LEFTPADDING",(0,0),(-1,-1),0),("RIGHTPADDING",(0,0),(-1,-1),0)]))
    els.append(Spacer(1, 1*mm))
    els.append(ft)
    # ═══════════════════════════════════
    # 5. AMOUNT IN WORDS
    # ═══════════════════════════════════
    def n2w(n):
        if n == 0: return "Zero"
        o = ["","One","Two","Three","Four","Five","Six","Seven","Eight","Nine","Ten","Eleven","Twelve",
             "Thirteen","Fourteen","Fifteen","Sixteen","Seventeen","Eighteen","Nineteen"]
        t = ["","","Twenty","Thirty","Forty","Fifty","Sixty","Seventy","Eighty","Ninety"]
        sc = ["","Thousand","Million","Billion"]
        def h(num):
            r = ""
            if num >= 100: r += o[num//100] + " Hundred"; num %= 100
            if num and r: r += " "
            if num >= 20: r += t[num//10]; num %= 10
            if num and r: r += " "
            if num > 0: r += o[num]
            return r.strip()
        ip = int(n)
        dp = min(int(round((n - ip) * 100)), 99)
        if ip == 0: w = "Zero"
        else:
            w = ""; i = 0
            while ip > 0:
                ck = ip % 1000
                if ck:
                    cw = h(ck)
                    if sc[i]: cw += " " + sc[i]
                    w = cw + (" " + w if w else "")
                ip //= 1000; i += 1
        if dp: w += f" and {dp:02d}/100"
        return "AED " + w + " Only"

    els.append(Spacer(1, 3*mm))
    ab = Table([[Paragraph(f"<b>Amount in Words:</b> {n2w(tot)}", S("AW", fontSize=9, textColor=C4, leading=13))]], colWidths=[W])
    ab.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,-1),BG),("LEFTPADDING",(0,0),(-1,-1),8),("RIGHTPADDING",(0,0),(-1,-1),8),("TOPPADDING",(0,0),(-1,-1),5),("BOTTOMPADDING",(0,0),(-1,-1),5)]))
    els.append(ab)

    display_notes = inv.get("notes", "") or ""
    if is_nmdc and display_notes:
        import json
        lines = display_notes.split("\n", 1)
        if lines and lines[0].strip().startswith("{"):
            try:
                json.loads(lines[0])
                display_notes = lines[1].strip() if len(lines) > 1 else ""
            except Exception:
                pass
    if display_notes:
        els.append(Spacer(1, 1.5*mm))
        nb = Table([[Paragraph(f"<b>Notes:</b> {display_notes}", S("NW", fontSize=7, textColor=C4, leading=9))]], colWidths=[W])
        nb.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,-1),colors.HexColor("#f8fafc")),("BOX",(0,0),(-1,-1),0.5,colors.HexColor("#e2e8f0")),("LEFTPADDING",(0,0),(-1,-1),5),("RIGHTPADDING",(0,0),(-1,-1),5),("TOPPADDING",(0,0),(-1,-1),2),("BOTTOMPADDING",(0,0),(-1,-1),2)]))
        els.append(nb)

    # ═══════════════════════════════════
    # 6. BANK DETAILS (single column label-value pairs)
    # ═══════════════════════════════════
    if company and (company["bank_name"] or company["bank_account_name"] or company["bank_account_number"] or company["iban"]):
        bk_items = []
        if company["bank_name"]: bk_items.append(("Bank", company["bank_name"]))
        if company["bank_account_name"]: bk_items.append(("Account", company["bank_account_name"]))
        if company["bank_account_number"]: bk_items.append(("A/C No.", company["bank_account_number"]))
        if company["iban"]: bk_items.append(("IBAN", company["iban"]))
        if company["swift_code"]: bk_items.append(("Swift", company["swift_code"]))
        if bk_items:
            els.append(Spacer(1, 1.5*mm))
            els.append(Paragraph("<b>BANK DETAILS</b>", S("BD", fontSize=7, fontName="Helvetica-Bold", textColor=C5, leading=8, spaceAfter=1)))
            bk_rows = [[
                Paragraph(f"<font color='#64748b'>{lbl}:</font>", S("_bkl", fontSize=7, textColor=C5, leading=9)),
                Paragraph(f"<b>{val}</b>", S("_bkv", fontSize=7, fontName="Helvetica-Bold", textColor=C4, leading=9)),
            ] for lbl, val in bk_items]
            bkt = Table(bk_rows, colWidths=[20*mm, W - 20*mm])
            bkt.setStyle(TableStyle([
                ("VALIGN",(0,0),(-1,-1),"TOP"),
                ("TOPPADDING",(0,0),(-1,-1),1), ("BOTTOMPADDING",(0,0),(-1,-1),1),
                ("LEFTPADDING",(0,0),(-1,-1),0), ("RIGHTPADDING",(0,0),(-1,-1),0),
            ]))
            els.append(bkt)

    # ═══════════════════════════════════
    # 7. SIGNATURES — stamp & sign side by side
    # ═══════════════════════════════════
    els.append(Spacer(1, 2*mm))
    sg = ParagraphStyle("SG", fontSize=8, alignment=TA_CENTER, leading=11)
    stamp_path = os.path.join(current_app.root_path, 'static', 'Stamp.png')
    sign_path = os.path.join(current_app.root_path, 'static', 'Sign (1).png')
    auth_img = []
    if os.path.exists(stamp_path):
        auth_img.append(Image(stamp_path, width=28, height=28))
    if os.path.exists(sign_path):
        auth_img.append(Image(sign_path, width=28, height=28))
    if auth_img:
        auth_imgs = Table([auth_img], colWidths=[28]*len(auth_img))
        auth_imgs.setStyle(TableStyle([
            ("ALIGN",(0,0),(-1,-1),"CENTER"),
            ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
            ("LEFTPADDING",(0,0),(-1,-1),2), ("RIGHTPADDING",(0,0),(-1,-1),2),
        ]))
    else:
        auth_imgs = Paragraph("", sg)
    auth_cell = Table([
        [Paragraph("_________________________", sg)],
        [auth_imgs],
        [Paragraph("<b>Authorized Signatory</b>", sg)],
    ], colWidths=[W*0.38])
    auth_cell.setStyle(TableStyle([
        ("TOPPADDING",(0,0),(-1,-1),0), ("BOTTOMPADDING",(0,0),(-1,-1),0),
        ("LEFTPADDING",(0,0),(-1,-1),0), ("RIGHTPADDING",(0,0),(-1,-1),0),
    ]))
    auth_cell.setStyle(TableStyle([
        ("ALIGN",(0,0),(-1,-1),"CENTER"),
        ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
        ("TOPPADDING",(0,0),(-1,-1),0), ("BOTTOMPADDING",(0,0),(-1,-1),1),
    ]))
    sgt = Table([[
        auth_cell,
        C("", fontSize=4),
        Paragraph("", sg),
    ]], colWidths=[W*0.38, W*0.24, W*0.38])
    sgt.setStyle(TableStyle([
        ("VALIGN",(0,0),(-1,-1),"TOP"),
        ("LINEABOVE",(0,0),(0,0),0.5,C5), ("LINEABOVE",(2,0),(2,0),0.5,C5),
        ("LEFTPADDING",(0,0),(-1,-1),0), ("RIGHTPADDING",(0,0),(-1,-1),0),
    ]))
    els.append(sgt)
    # ═══════════════════════════════════
    # 8. COMPANY ADDRESS / FOOTER
    # ═══════════════════════════════════
    if c_addr:
        els.append(Spacer(1, 1.5*mm))
        els.append(Paragraph(f"<font size=6 color='#64748b'>{c_addr}</font>", S("_ad", fontSize=6, textColor=C5, alignment=TA_CENTER, leading=7)))
    els.append(Spacer(1, 2*mm))

    pp = []
    if company:
        parts = []
        if company["bank_name"]: parts.append(f"Bank: <b>{company['bank_name']}</b>")
        if company["bank_account_number"]: parts.append(f"A/C: <b>{company['bank_account_number']}</b>")
        if company["iban"]: parts.append(f"IBAN: <b>{company['iban']}</b>")
        if parts:
            pp.append(Paragraph("Payable at: " + " | ".join(parts), S("FP", fontSize=6.5, textColor=C4, alignment=TA_CENTER, leading=8)))

    pp.append(Paragraph(
        "This is a computer-generated Tax Invoice. Valid without signature.",
        S("FN", fontSize=6.5, textColor=C5, alignment=TA_CENTER, leading=8)))

    fh = Table([[""]], colWidths=[W], rowHeights=[0.3])
    fh.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,-1),TH),("LEFTPADDING",(0,0),(-1,-1),0),("RIGHTPADDING",(0,0),(-1,-1),0)]))
    els.append(fh)
    els.append(Spacer(1, 1*mm))
    for p in pp:
        els.append(p)
        els.append(Spacer(1, 0.5*mm))

    doc.build(els)
    for f in _logo_tmp_files:
        try: os.remove(f)
        except: pass
    pdf_data = buf.getvalue(); buf.close()
    return send_file(BytesIO(pdf_data), mimetype="application/pdf", as_attachment=True, download_name=f"Invoice_{inv_no}.pdf")

@customer_bp.route("/<int:cid>/invoice/<int:iid>/delete", methods=["POST"])
def customer_invoice_delete(cid, iid):
    db = _get_db()
    db.execute("DELETE FROM customer_invoice_items WHERE invoice_id=?", (iid,))
    db.execute("DELETE FROM customer_invoices WHERE id=? AND customer_id=?", (iid, cid))
    db.commit()
    db.close()
    flash("Invoice deleted.", "success")
    return redirect(url_for("customer.customer_profile", cid=cid, tab="invoices"))

# ─── CREDIT NOTES ───

@customer_bp.route("/<int:cid>/credit-note/add", methods=["GET", "POST"])
def customer_credit_note_add(cid):
    _ensure_tables()
    c = _get_customer_or_404(cid)
    if not c: return redirect(url_for("customer.customer_dashboard"))
    db = _get_db()
    invoices = db.execute("SELECT id,invoice_no,total_amount,invoice_date FROM customer_invoices WHERE customer_id=? ORDER BY invoice_date DESC", (cid,)).fetchall()

    if request.method == "POST":
        cn_date = request.form.get("credit_note_date", date.today().isoformat())
        cn_no = request.form.get("credit_note_no", "").strip()
        inv_id = request.form.get("invoice_id") or None
        amount = float(request.form.get("amount", 0) or 0)
        vat_pct = float(request.form.get("vat_percent", 0) or 0)
        reason = request.form.get("reason", "").strip()
        notes = request.form.get("notes", "").strip()

        if not cn_no:
            flash("Credit note number is required.", "error")
            db.close()
            return render_template("customer/credit_note_form.html", c=c, invoices=invoices, today=date.today().isoformat())
        if amount <= 0:
            flash("Amount must be greater than zero.", "error")
            db.close()
            return render_template("customer/credit_note_form.html", c=c, invoices=invoices, today=date.today().isoformat())

        vat_amt = round(amount * vat_pct / 100, 2)
        total = round(amount + vat_amt, 2)

        db.execute(
            "INSERT INTO customer_credit_notes (customer_id, credit_note_no, credit_note_date, invoice_id, amount, vat_percent, vat_amount, total_amount, reason, notes) VALUES (?,?,?,?,?,?,?,?,?,?)",
            (cid, cn_no, cn_date, inv_id, amount, vat_pct, vat_amt, total, reason or None, notes or None),
        )
        db.commit()
        db.close()
        flash(f"Credit Note {cn_no} created.", "success")
        return redirect(url_for("customer.customer_profile", cid=cid, tab="credit_notes"))

    db.close()
    return render_template("customer/credit_note_form.html", c=c, invoices=invoices, today=date.today().isoformat())

@customer_bp.route("/<int:cid>/credit-note/<int:cnid>/delete", methods=["POST"])
def customer_credit_note_delete(cid, cnid):
    _ensure_tables()
    db = _get_db()
    row = db.execute("SELECT credit_note_no FROM customer_credit_notes WHERE id=? AND customer_id=?", (cnid, cid)).fetchone()
    if row:
        db.execute("DELETE FROM customer_credit_notes WHERE id=?", (cnid,))
        db.commit()
        flash(f"Credit Note {row['credit_note_no']} deleted.", "success")
    else:
        flash("Credit note not found.", "error")
    db.close()
    return redirect(url_for("customer.customer_profile", cid=cid, tab="credit_notes"))

# ─── PAYMENTS ───

@customer_bp.route("/<int:cid>/payment/add", methods=["GET", "POST"])
def customer_payment_add(cid):
    c = _get_customer_or_404(cid)
    if not c: return redirect(url_for("customer.customer_dashboard"))
    db = _get_db()
    invoices_raw = db.execute(
        "SELECT i.id,i.invoice_no,i.invoice_date,i.total_amount,COALESCE((SELECT SUM(p.amount) FROM customer_payments p WHERE p.invoice_id=i.id),0) AS paid,COALESCE((SELECT SUM(cn.total_amount) FROM customer_credit_notes cn WHERE cn.invoice_id=i.id),0) AS credit_notes FROM customer_invoices i WHERE i.customer_id=? ORDER BY i.invoice_date DESC",
        (cid,),
    ).fetchall()
    invoices = []
    for inv in invoices_raw:
        d = dict(inv)
        owed = d["total_amount"] - d["paid"] - d["credit_notes"]
        d["balance"] = round(owed, 2)
        if owed > 0.005:
            invoices.append(d)
    total_balance = round(sum(d["balance"] for d in invoices), 2)
    if request.method == "POST":
        pmt_date = request.form.get("payment_date", date.today().isoformat())
        method = request.form.get("payment_method", "Cheque")
        ref = request.form.get("reference_no", "").strip()
        notes = request.form.get("notes", "").strip()
        inv_ids = request.form.getlist("inv_ids")
        created_ids = []
        total_alloc = 0
        for inv_id in inv_ids:
            alloc_key = f"alloc_{inv_id}"
            alloc_amt = float(request.form.get(alloc_key, 0) or 0)
            if alloc_amt > 0:
                cur = db.execute(
                    "INSERT INTO customer_payments (customer_id,invoice_id,payment_date,amount,payment_method,reference_no,notes) VALUES (?,?,?,?,?,?,?)",
                    (cid, int(inv_id), pmt_date, alloc_amt, method, ref or None, notes or None),
                )
                created_ids.append(str(cur.lastrowid))
                total_alloc += alloc_amt
        if created_ids:
            db.commit()
            ids_param = ",".join(created_ids)
            flash(f"Payment of AED {total_alloc:.2f} recorded.", "success")
            undo_url = url_for('customer.customer_payment_undo', cid=cid, ids=ids_param)
            flash(Markup(f'<a href="{undo_url}" style="color:#fff;text-decoration:underline;font-weight:700">Undo this payment</a>'), "info")
        else:
            flash("No amount allocated to any invoice.", "error")
        db.close()
        return redirect(url_for("customer.customer_profile", cid=cid, tab="payments"))
    db.close()
    return render_template("customer/payment_form.html", c=c, invoices=invoices, today=date.today().isoformat(), balance=total_balance)

@customer_bp.route("/<int:cid>/payment/undo")
def customer_payment_undo(cid):
    ids_str = request.args.get("ids", "")
    if ids_str:
        db = _get_db()
        ids = [int(x) for x in ids_str.split(",") if x.strip().isdigit()]
        for pid in ids:
            db.execute("DELETE FROM customer_payments WHERE id=? AND customer_id=?", (pid, cid))
        db.commit()
        db.close()
        flash(f"Payment undone — {len(ids)} record(s) deleted.", "warning")
    return redirect(url_for("customer.customer_profile", cid=cid, tab="payments"))

@customer_bp.route("/<int:cid>/payment/<int:pid>/delete", methods=["POST"])
def customer_payment_delete(cid, pid):
    db = _get_db()
    db.execute("DELETE FROM customer_payments WHERE id=? AND customer_id=?", (pid, cid))
    db.commit()
    db.close()
    flash("Payment deleted.", "success")
    return redirect(url_for("customer.customer_profile", cid=cid, tab="payments"))

# ─── CONTRACTS ───

@customer_bp.route("/<int:cid>/contract/add", methods=["GET", "POST"])
def customer_contract_add(cid):
    c = _get_customer_or_404(cid)
    if not c: return redirect(url_for("customer.customer_dashboard"))
    db = _get_db()
    if request.method == "POST":
        db.execute("INSERT INTO customer_contracts (customer_id,contract_no,contract_date,start_date,end_date,contract_type,amount,status,notes) VALUES (?,?,?,?,?,?,?,?,?)",
            (cid, request.form.get("contract_no"), request.form.get("contract_date", date.today().isoformat()),
             request.form.get("start_date"), request.form.get("end_date"), request.form.get("contract_type", "rental"),
             float(request.form.get("amount", 0) or 0), request.form.get("status", "active"), request.form.get("notes")))
        db.commit()
        db.close()
        flash("Contract added.", "success")
        return redirect(url_for("customer.customer_profile", cid=cid, tab="contracts"))
    db.close()
    return render_template("customer/contract_form.html", c=c, contract={}, today=date.today().isoformat())

@customer_bp.route("/<int:cid>/contract/<int:ctid>/close", methods=["POST"])
def customer_contract_close(cid, ctid):
    db = _get_db()
    db.execute("UPDATE customer_contracts SET status='closed' WHERE id=? AND customer_id=?", (ctid, cid))
    db.commit()
    db.close()
    flash("Contract closed.", "success")
    return redirect(url_for("customer.customer_profile", cid=cid, tab="contracts"))

# ─── QUOTATIONS ───

def _next_quotation_no(db):
    db.execute("UPDATE quotation_sequence SET last_number = last_number + 1")
    n = db.execute("SELECT last_number FROM quotation_sequence").fetchone()[0]
    return f"QTN{n + 1000}"

@customer_bp.route("/<int:cid>/quotation/add", methods=["GET", "POST"])
def customer_quotation_add(cid):
    _ensure_tables()
    c = _get_customer_or_404(cid)
    if not c: return redirect(url_for("customer.customer_dashboard"))
    db = _get_db()
    next_no = _next_quotation_no(db)
    company = db.execute("SELECT * FROM company_profile LIMIT 1").fetchone()
    if request.method == "POST":
        q_date = request.form.get("quotation_date", date.today().isoformat())
        q_no = request.form.get("quotation_no", "").strip() or next_no
        location = request.form.get("location", "").strip()
        contact_details = request.form.get("contact_details", "").strip()
        existing = db.execute("SELECT id FROM customer_quotations WHERE quotation_no=?", (q_no,)).fetchone()
        if existing:
            flash(f"Quotation number '{q_no}' already exists.", "error")
            db.close()
            return render_template("customer/quotation_form.html", c=c, q={}, company=company, today=date.today().isoformat(), next_no=next_no)
        vat_pct = float(request.form.get("vat_percent", 5))
        terms = request.form.get("terms", "").strip()
        notes = request.form.get("notes", "").strip()
        descs = request.form.getlist("item_desc[]")
        qtys = request.form.getlist("item_qty[]")
        rates = request.form.getlist("item_rate[]")
        units = request.form.getlist("item_unit[]")
        items = []
        sub_total = 0
        for i in range(len(descs)):
            desc = descs[i].strip()
            qty = float(qtys[i]) if i < len(qtys) and qtys[i].strip() else 1
            rate = float(rates[i]) if i < len(rates) and rates[i].strip() else 0
            unit = units[i].strip() if i < len(units) and units[i].strip() else "hr"
            if desc or rate > 0:
                amt = round(qty * rate, 2)
                sub_total += amt
                items.append({"desc": desc, "qty": qty, "rate": rate, "amt": amt, "unit": unit})
        if not items:
            flash("At least one line item is required.", "error")
            db.close()
            return render_template("customer/quotation_form.html", c=c, q={}, company=company, today=date.today().isoformat(), next_no=next_no)
        vat_amt = round(sub_total * vat_pct / 100, 2)
        total = round(sub_total + vat_amt, 2)
        cur = db.execute(
            """INSERT INTO customer_quotations (customer_id,quotation_no,quotation_date,sub_total,vat_percent,vat_amount,total_amount,status,terms,notes,location,contact_details)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?)""",
            (cid, q_no, q_date, sub_total, vat_pct, vat_amt, total, "pending", terms, notes, location, contact_details))
        qid = cur.lastrowid
        for idx, it in enumerate(items):
            db.execute("INSERT INTO customer_quotation_items (quotation_id,description,quantity,rate,amount,unit,sort_order) VALUES (?,?,?,?,?,?,?)",
                (qid, it["desc"], it["qty"], it["rate"], it["amt"], it["unit"], idx))
            if it["desc"]:
                try:
                    db.execute("INSERT OR IGNORE INTO service_items (description, default_rate) VALUES (?,?)", (it["desc"], it["rate"]))
                except Exception:
                    pass
        db.commit()
        db.close()
        flash(f"Quotation {q_no} created.", "success")
        return redirect(url_for("customer.customer_profile", cid=cid, tab="quotations"))
    db.close()
    return render_template("customer/quotation_form.html", c=c, q={}, company=company, today=date.today().isoformat(), next_no=next_no)

@customer_bp.route("/<int:cid>/quotation/<int:qid>/edit", methods=["GET", "POST"])
def customer_quotation_edit(cid, qid):
    _ensure_tables()
    c = _get_customer_or_404(cid)
    if not c: return redirect(url_for("customer.customer_dashboard"))
    db = _get_db()
    q = db.execute("SELECT * FROM customer_quotations WHERE id=? AND customer_id=?", (qid, cid)).fetchone()
    if not q:
        db.close()
        flash("Quotation not found.", "error")
        return redirect(url_for("customer.customer_profile", cid=cid, tab="quotations"))
    items = db.execute("SELECT * FROM customer_quotation_items WHERE quotation_id=? ORDER BY sort_order", (qid,)).fetchall()
    company = db.execute("SELECT * FROM company_profile LIMIT 1").fetchone()
    if request.method == "POST":
        q_date = request.form.get("quotation_date", q["quotation_date"])
        q_no = request.form.get("quotation_no", "").strip() or q["quotation_no"]
        location = request.form.get("location", "").strip()
        contact_details = request.form.get("contact_details", "").strip()
        vat_pct = float(request.form.get("vat_percent", q["vat_percent"] or 5))
        terms = request.form.get("terms", "").strip()
        notes = request.form.get("notes", "").strip()
        descs = request.form.getlist("item_desc[]")
        qtys = request.form.getlist("item_qty[]")
        rates = request.form.getlist("item_rate[]")
        units = request.form.getlist("item_unit[]")
        new_items = []
        sub_total = 0
        for i in range(len(descs)):
            desc = descs[i].strip()
            qty = float(qtys[i]) if i < len(qtys) and qtys[i].strip() else 1
            rate = float(rates[i]) if i < len(rates) and rates[i].strip() else 0
            unit = units[i].strip() if i < len(units) and units[i].strip() else "hr"
            if desc or rate > 0:
                amt = round(qty * rate, 2)
                sub_total += amt
                new_items.append({"desc": desc, "qty": qty, "rate": rate, "amt": amt, "unit": unit})
        if not new_items:
            flash("At least one line item is required.", "error")
            db.close()
            return render_template("customer/quotation_form.html", c=c, q=q, items=items, edit=True, company=company, today=date.today().isoformat())
        vat_amt = round(sub_total * vat_pct / 100, 2)
        total = round(sub_total + vat_amt, 2)
        db.execute(
            """UPDATE customer_quotations SET quotation_no=?,quotation_date=?,sub_total=?,vat_percent=?,vat_amount=?,total_amount=?,terms=?,notes=?,location=?,contact_details=? WHERE id=?""",
            (q_no, q_date, sub_total, vat_pct, vat_amt, total, terms, notes, location, contact_details, qid))
        db.execute("DELETE FROM customer_quotation_items WHERE quotation_id=?", (qid,))
        for idx, it in enumerate(new_items):
            db.execute("INSERT INTO customer_quotation_items (quotation_id,description,quantity,rate,amount,unit,sort_order) VALUES (?,?,?,?,?,?,?)",
                (qid, it["desc"], it["qty"], it["rate"], it["amt"], it["unit"], idx))
        db.commit()
        db.close()
        flash(f"Quotation {q_no} updated.", "success")
        return redirect(url_for("customer.customer_profile", cid=cid, tab="quotations"))
    db.close()
    return render_template("customer/quotation_form.html", c=c, q=q, items=items, edit=True, company=company, today=date.today().isoformat())

@customer_bp.route("/<int:cid>/quotation/<int:qid>/delete", methods=["POST"])
def customer_quotation_delete(cid, qid):
    db = _get_db()
    db.execute("DELETE FROM customer_quotations WHERE id=? AND customer_id=?", (qid, cid))
    db.commit()
    db.close()
    flash("Quotation deleted.", "success")
    return redirect(url_for("customer.customer_profile", cid=cid, tab="quotations"))


@customer_bp.route("/<int:cid>/quotation/<int:qid>/approve", methods=["POST"])
def customer_quotation_approve(cid, qid):
    _ensure_tables()
    db = _get_db()
    q = db.execute("SELECT * FROM customer_quotations WHERE id=? AND customer_id=?", (qid, cid)).fetchone()
    if not q:
        db.close()
        flash("Quotation not found.", "error")
        return redirect(url_for("customer.customer_profile", cid=cid, tab="quotations"))
    if q["status"] != "pending":
        db.close()
        flash("Only pending quotations can be approved.", "error")
        return redirect(url_for("customer.customer_quotation_view", cid=cid, qid=qid))
    db.execute("UPDATE customer_quotations SET status='approved' WHERE id=?", (qid,))
    db.commit()
    db.close()
    flash(f"Quotation {q['quotation_no']} approved!", "success")
    return redirect(url_for("customer.customer_quotation_view", cid=cid, qid=qid))

@customer_bp.route("/<int:cid>/quotation/<int:qid>/view")
def customer_quotation_view(cid, qid):
    _ensure_tables()
    c = _get_customer_or_404(cid)
    if not c: return redirect(url_for("customer.customer_dashboard"))
    db = _get_db()
    q = db.execute("SELECT * FROM customer_quotations WHERE id=? AND customer_id=?", (qid, cid)).fetchone()
    if not q:
        db.close()
        flash("Quotation not found.", "error")
        return redirect(url_for("customer.customer_profile", cid=cid, tab="quotations"))
    items = db.execute("SELECT * FROM customer_quotation_items WHERE quotation_id=? ORDER BY sort_order", (qid,)).fetchall()
    company = db.execute("SELECT * FROM company_profile LIMIT 1").fetchone()
    db.close()
    return render_template("customer/quotation_view.html", c=c, q=q, items=items, company=company)

@customer_bp.route("/<int:cid>/quotation/<int:qid>/pdf")
def customer_quotation_pdf(cid, qid):
    import tempfile
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
    from io import BytesIO

    _logo_tmp_files = []
    _ensure_tables()
    db = _get_db()
    c = db.execute("SELECT * FROM customers WHERE id=?", (cid,)).fetchone()
    q = db.execute("SELECT * FROM customer_quotations WHERE id=? AND customer_id=?", (qid, cid)).fetchone()
    items = db.execute("SELECT * FROM customer_quotation_items WHERE quotation_id=? ORDER BY sort_order", (qid,)).fetchall()
    company = db.execute("SELECT * FROM company_profile LIMIT 1").fetchone()
    db.close()
    if not c or not q:
        flash("Quotation not found.", "error")
        return redirect(url_for("customer.customer_dashboard"))

    buf = BytesIO()
    LM, RM, TM, BM = 18*mm, 18*mm, 15*mm, 12*mm
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=LM, rightMargin=RM, topMargin=TM, bottomMargin=BM)
    W = A4[0] - LM - RM

    tc = company["theme_color"] or "#1a3a5c" if company else "#1a3a5c"
    try: TH = colors.HexColor(tc)
    except: TH = colors.HexColor("#1a3a5c")
    WH = colors.white; BG = colors.HexColor("#f8fafc")
    C3 = colors.HexColor("#e2e8f0"); C4 = colors.HexColor("#0f172a")
    C5 = colors.HexColor("#64748b"); C6 = colors.HexColor("#dc2626")

    cn = (company["company_name"] if company else "CURRENT LINK TRANSPORT AND GENERAL CONTRACTING") or "CURRENT LINK TRANSPORT AND GENERAL CONTRACTING"
    c_addr = (company["address"] or "") if company else ""
    c_ph = (company["phone_number"] or "") if company else ""
    c_em = (company["email"] or "") if company else ""
    c_trn = company["trn_no"] or "—" if company else "—"

    def S(name, **kw):
        kw.setdefault("fontSize", 8)
        kw.setdefault("leading", 12)
        return ParagraphStyle(name, **kw)

    def L(t, **kw):
        kw.setdefault("textColor", C5)
        return Paragraph(str(t), S("_L", **kw))

    def V(t, **kw):
        kw.setdefault("fontName", "Helvetica-Bold")
        kw.setdefault("textColor", C4)
        kw.setdefault("fontSize", 8.5)
        return Paragraph(str(t), S("_V", **kw))

    def C(t, **kw):
        kw.setdefault("alignment", TA_CENTER)
        return Paragraph(str(t), S("_C", **kw))

    def R(t, **kw):
        kw.setdefault("alignment", TA_RIGHT)
        return Paragraph(str(t), S("_R", **kw))

    def RB(t, **kw):
        kw.setdefault("fontName", "Helvetica-Bold")
        kw.setdefault("alignment", TA_RIGHT)
        return Paragraph(f"<b>{t}</b>", S("_RB", **kw))

    safe = lambda v, d="—": str(v) if v else d
    els = []
    q_no = q["quotation_no"] or "—"
    q_dt = q["quotation_date"] or "—"

    # HEADER
    logo = None; LW = 0
    if company and company["logo_data"]:
        try:
            lb = base64.b64decode(company["logo_data"])
            f = tempfile.NamedTemporaryFile(delete=False, suffix=".png")
            f.write(lb); f.close()
            logo = Image(f.name, width=90, height=90)
            LW = 90
            _logo_tmp_files.append(f.name)
        except: pass

    ci_lines = []
    if c_addr: ci_lines.append(f"<font size=7 color='#64748b'>{c_addr}</font>")
    c_contact = []
    if c_ph: c_contact.append(f"Phone: {c_ph}")
    if c_em: c_contact.append(f"Email: {c_em}")
    if c_contact: ci_lines.append('<font size=7 color="#64748b">' + ' &middot; '.join(c_contact) + '</font>')
    ci_lines.append(f"<font size=7 color='#64748b'><b>TRN: {c_trn}</b></font>")
    ci_html = f"<font size=12><b>{cn}</b></font><br/>" + "<br/>".join(ci_lines)
    co_p = Paragraph(ci_html, S("CO", fontSize=12, fontName="Helvetica-Bold", textColor=TH, leading=16))

    if logo:
        lh = Table([[logo, Spacer(1, 4*mm), co_p]], colWidths=[LW, 4*mm, W*0.65 - LW - 4*mm])
        lh.setStyle(TableStyle([("VALIGN",(0,0),(-1,-1),"TOP"),("LEFTPADDING",(0,0),(-1,-1),0),("RIGHTPADDING",(0,0),(-1,-1),0)]))
    else:
        lh = co_p

    rh = Paragraph(
        f"<b>QUOTATION</b><br/>"
        f"<font size=7 color='#64748b'># {q_no}<br/>{q_dt}</font>",
        S("TI", fontSize=16, fontName="Helvetica-Bold", textColor=TH, leading=20, alignment=TA_RIGHT))

    ht = Table([[lh, rh]], colWidths=[W*0.65, W*0.35])
    ht.setStyle(TableStyle([("VALIGN",(0,0),(-1,-1),"TOP"),("LEFTPADDING",(0,0),(-1,-1),0),("RIGHTPADDING",(0,0),(-1,-1),0)]))
    els.append(ht)

    bl = Table([[""]], colWidths=[W], rowHeights=[3])
    bl.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,-1),TH),("LEFTPADDING",(0,0),(-1,-1),0),("RIGHTPADDING",(0,0),(-1,-1),0)]))
    els.append(bl)
    els.append(Spacer(1, 5*mm))

    # BILL TO / QUOTATION INFO
    def card(title, pairs):
        cw = W*0.50
        r = [[Paragraph(f"<b>{title}</b>", S("_ch", fontSize=6.5, fontName="Helvetica-Bold", textColor=C5, leading=9)), Paragraph("", S("_cs", fontSize=2, leading=2))]]
        for a, b in pairs:
            r.append([Paragraph(a, S("_cl", fontSize=7.5, textColor=C5, leading=11)), Paragraph(f"{b}", S("_cv", fontSize=8, fontName="Helvetica-Bold", textColor=C4, leading=11.5))])
        t = Table(r, colWidths=[cw*0.28, cw*0.72])
        t.setStyle(TableStyle([("VALIGN",(0,0),(-1,-1),"TOP"),("TOPPADDING",(0,0),(-1,-1),3),("BOTTOMPADDING",(0,0),(-1,-1),3),("LEFTPADDING",(0,0),(-1,-1),10),("RIGHTPADDING",(0,0),(-1,-1),10),("BOX",(0,0),(-1,-1),0.5,C3)]))
        return t

    bd = [("Customer", safe(c["customer_name"])), ("TRN", safe(c["trn"]))]
    if c["phone"]: bd.append(("Phone", c["phone"]))
    if c["email"]: bd.append(("Email", c["email"]))
    if c["address"]: bd.append(("Address", c["address"]))
    id_ = [("Quotation #", q_no), ("Date", q_dt), ("Status", q["status"].upper() if q["status"] else "PENDING")]
    try:
        if q["location"]: id_.append(("Location", q["location"]))
    except (KeyError, IndexError):
        pass

    iw = Table([[card("CUSTOMER", bd), Spacer(1, 4*mm), card("QUOTATION INFO", id_)]], colWidths=[W*0.50, 4*mm, W*0.50])
    iw.setStyle(TableStyle([("VALIGN",(0,0),(-1,-1),"TOP"),("LEFTPADDING",(0,0),(-1,-1),0),("RIGHTPADDING",(0,0),(-1,-1),0)]))
    els.append(iw)
    els.append(Spacer(1, 5*mm))

    # ITEMS TABLE
    fixed_pt = 35*mm + 4*mm + 35*mm + 4*mm + 2*mm + 20*mm + 3*mm + 7*mm + 10*mm + 8*mm + 8*mm
    avail_pt = A4[1] - TM - BM - fixed_pt
    num_rows = len(items)
    fs = 7.0
    if num_rows > 0:
        target = avail_pt / (num_rows + 1)
        fs = max(4.0, min(7.0, target / 2.8))
    ldr = fs * 1.35
    pad_t = max(1.5, fs * 0.5)
    pad_b = max(1.5, fs * 0.5)

    DH = colors.HexColor("#1e293b")
    cw = [10*mm, 52*mm, 16*mm, 16*mm, 22*mm, 26*mm, 26*mm]
    def _pc(t, **kw):
        kw.setdefault("fontSize", fs)
        kw.setdefault("leading", ldr)
        return Paragraph(str(t), S("_pc", **kw))
    hdr = [
        Paragraph("<b>#</b>", S("_h0", fontSize=fs, fontName="Helvetica-Bold", textColor=WH, alignment=TA_CENTER, leading=ldr)),
        Paragraph("<b>Description</b>", S("_h1", fontSize=fs, fontName="Helvetica-Bold", textColor=WH, leading=ldr)),
        Paragraph("<b>Qty</b>", S("_h2", fontSize=fs, fontName="Helvetica-Bold", textColor=WH, alignment=TA_CENTER, leading=ldr)),
        Paragraph("<b>Unit</b>", S("_hu", fontSize=fs, fontName="Helvetica-Bold", textColor=WH, alignment=TA_CENTER, leading=ldr)),
        Paragraph("<b>Rate<br/>(AED)</b>", S("_h3", fontSize=fs, fontName="Helvetica-Bold", textColor=WH, alignment=TA_RIGHT, leading=ldr)),
        Paragraph("<b>Amount<br/>(AED)</b>", S("_h4", fontSize=fs, fontName="Helvetica-Bold", textColor=WH, alignment=TA_RIGHT, leading=ldr)),
        Paragraph("", S("_h5", fontSize=fs, fontName="Helvetica-Bold", textColor=WH, leading=ldr)),
    ]
    rws = [hdr]
    for idx, it in enumerate(items):
        rws.append([
            _pc(str(idx+1), alignment=TA_CENTER, fontName="Helvetica-Bold"),
            _pc(it["description"] or "—"),
            _pc(f"{it['quantity'] or 0:,.2f}", alignment=TA_CENTER),
            _pc((it['unit'] or 'hr').upper(), alignment=TA_CENTER),
            _pc(f"{it['rate'] or 0:,.3f}", alignment=TA_RIGHT),
            _pc(f"{it['amount'] or 0:,.2f}", alignment=TA_RIGHT),
            _pc("", alignment=TA_RIGHT),
        ])

    sub = q["sub_total"] or q["amount"] or 0
    vat = q["vat_amount"] or 0
    tot = q["total_amount"] or sub
    vp = q["vat_percent"] or 0

    itt = Table(rws, colWidths=cw, repeatRows=1)
    itt.setStyle(TableStyle([
        ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
        ("BACKGROUND",(0,0),(-1,0),DH), ("TEXTCOLOR",(0,0),(-1,0),WH),
        ("BOX",(0,0),(-1,-1),0.5,C3),
        ("INNERGRID",(0,0),(-1,-1),0.3,C3),
        ("TOPPADDING",(0,0),(-1,-1),pad_t), ("BOTTOMPADDING",(0,0),(-1,-1),pad_b),
        ("LEFTPADDING",(0,0),(-1,-1),6), ("RIGHTPADDING",(0,0),(-1,-1),6),
        ("ROWBACKGROUNDS",(0,1),(-1,-1),[WH, BG]),
    ]))
    els.append(itt)

    # TOTALS
    tw = 90*mm
    trows = [
        [Paragraph("Sub Total", S("_st", fontSize=9, textColor=C5, leading=14)),
         Paragraph(f"<b>AED {sub:,.2f}</b>", S("_stv", fontSize=9, fontName="Helvetica-Bold", textColor=C4, leading=14, alignment=TA_RIGHT))],
        [Paragraph(f"VAT @ {vp:.0f}%", S("_vt", fontSize=9, textColor=C5, leading=14)),
         Paragraph(f"<b>AED {vat:,.2f}</b>", S("_vtv", fontSize=9, fontName="Helvetica-Bold", textColor=C6, leading=14, alignment=TA_RIGHT))],
        [Paragraph("<b>Total</b>", S("_td", fontSize=11, fontName="Helvetica-Bold", textColor=C4, leading=16)),
         Paragraph(f"<b>AED {tot:,.2f}</b>", S("_tdv", fontSize=13, fontName="Helvetica-Bold", textColor=TH, leading=18, alignment=TA_RIGHT))],
    ]
    tt = Table(trows, colWidths=[tw*0.45, tw*0.55])
    tt.setStyle(TableStyle([
        ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
        ("TOPPADDING",(0,0),(-1,-1),4), ("BOTTOMPADDING",(0,0),(-1,-1),4),
        ("LEFTPADDING",(0,0),(-1,-1),12), ("RIGHTPADDING",(0,0),(-1,-1),12),
        ("BOX",(0,0),(-1,-1),0.5,C3),
        ("LINEABOVE",(0,2),(-1,2),2,TH),
    ]))

    ft = Table([["", tt]], colWidths=[W - tw, tw])
    ft.setStyle(TableStyle([("VALIGN",(0,0),(-1,-1),"TOP"),("LEFTPADDING",(0,0),(-1,-1),0),("RIGHTPADDING",(0,0),(-1,-1),0)]))
    els.append(Spacer(1, 2*mm))
    els.append(ft)

    # AMOUNT IN WORDS
    def n2w(n):
        if n == 0: return "Zero"
        o = ["","One","Two","Three","Four","Five","Six","Seven","Eight","Nine","Ten","Eleven","Twelve",
             "Thirteen","Fourteen","Fifteen","Sixteen","Seventeen","Eighteen","Nineteen"]
        t = ["","","Twenty","Thirty","Forty","Fifty","Sixty","Seventy","Eighty","Ninety"]
        sc = ["","Thousand","Million","Billion"]
        def h(num):
            r = ""
            if num >= 100: r += o[num//100] + " Hundred"; num %= 100
            if num and r: r += " "
            if num >= 20: r += t[num//10]; num %= 10
            if num and r: r += " "
            if num > 0: r += o[num]
            return r.strip()
        ip = int(n)
        dp = min(int(round((n - ip) * 100)), 99)
        if ip == 0: w = "Zero"
        else:
            w = ""; i = 0
            while ip > 0:
                ck = ip % 1000
                if ck:
                    cw = h(ck)
                    if sc[i]: cw += " " + sc[i]
                    w = cw + (" " + w if w else "")
                ip //= 1000; i += 1
        if dp: w += f" and {dp:02d}/100"
        return "AED " + w + " Only"

    els.append(Spacer(1, 3*mm))
    ab = Table([[Paragraph(f"<b>Amount in Words:</b> {n2w(tot)}", S("AW", fontSize=9, textColor=C4, leading=13))]], colWidths=[W])
    ab.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,-1),BG),("LEFTPADDING",(0,0),(-1,-1),8),("RIGHTPADDING",(0,0),(-1,-1),8),("TOPPADDING",(0,0),(-1,-1),5),("BOTTOMPADDING",(0,0),(-1,-1),5)]))
    els.append(ab)

    if q["notes"]:
        els.append(Spacer(1, 3*mm))
        nb = Table([[Paragraph(f"<b>Notes:</b> {q['notes']}", S("NW", fontSize=9, textColor=C4, leading=13))]], colWidths=[W])
        nb.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,-1),BG),("BOX",(0,0),(-1,-1),0.5,C3),("LEFTPADDING",(0,0),(-1,-1),10),("RIGHTPADDING",(0,0),(-1,-1),10),("TOPPADDING",(0,0),(-1,-1),6),("BOTTOMPADDING",(0,0),(-1,-1),6)]))
        els.append(nb)

    # TERMS & CONDITIONS (always shown)
    els.append(Spacer(1, 3*mm))
    tc_lines = []
    if q["terms"]:
        tc_lines.append(f"<li>{q['terms']}</li>")
    tc_lines += [
        "<li>This quotation is valid for 15 days from the date of issue.</li>",
        "<li>Payment is due within 30 days from the date of invoice.</li>",
        "<li>Any alteration or cancellation of order must be notified in writing.</li>",
        "<li>All disputes are subject to UAE jurisdiction.</li>",
        "<li>Delivery / service execution as per agreed schedule.</li>",
        "<li>Rates are exclusive of any applicable taxes unless stated otherwise.</li>",
    ]
    tc_html = "<b>Terms &amp; Conditions:</b><br/><ol>" + "".join(tc_lines) + "</ol>"
    tb = Table([[Paragraph(tc_html, S("TW", fontSize=9, textColor=C4, leading=14))]], colWidths=[W])
    tb.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,-1),colors.HexColor("#fffbeb")),("BOX",(0,0),(-1,-1),0.5,colors.HexColor("#fde68a")),("LEFTPADDING",(0,0),(-1,-1),10),("RIGHTPADDING",(0,0),(-1,-1),10),("TOPPADDING",(0,0),(-1,-1),6),("BOTTOMPADDING",(0,0),(-1,-1),6)]))
    els.append(tb)

    # CONTACT DETAILS
    if company:
        els.append(Spacer(1, 3*mm))
        cd_items = [("Contact Person", "Mr. Nasrullah")]
        if company["phone_number"]: cd_items.append(("Phone", company["phone_number"]))
        if company["email"]: cd_items.append(("Email", company["email"]))
        if company["trn_no"]: cd_items.append(("TRN", company["trn_no"]))
        els.append(Paragraph("<b>CONTACT DETAILS</b>", S("BD", fontSize=8, fontName="Helvetica-Bold", textColor=C5, leading=10, spaceAfter=2)))
        cd_rows = [[Paragraph(f"<font color='#64748b'>{lbl}:</font>", S("_bkl", fontSize=8, textColor=C5, leading=12)), Paragraph(f"<b>{val}</b>", S("_bkv", fontSize=8, fontName="Helvetica-Bold", textColor=C4, leading=12))] for lbl, val in cd_items]
        cdt = Table(cd_rows, colWidths=[22*mm, W - 22*mm])
        cdt.setStyle(TableStyle([("VALIGN",(0,0),(-1,-1),"TOP"),("TOPPADDING",(0,0),(-1,-1),1.5),("BOTTOMPADDING",(0,0),(-1,-1),1.5),("LEFTPADDING",(0,0),(-1,-1),0),("RIGHTPADDING",(0,0),(-1,-1),0)]))
        els.append(cdt)

    # SIGNATURES
    els.append(Spacer(1, 3*mm))
    sg = ParagraphStyle("SG", fontSize=9, alignment=TA_CENTER, leading=14)
    stamp_path = os.path.join(current_app.root_path, 'static', 'Stamp.png')
    sign_path = os.path.join(current_app.root_path, 'static', 'Sign (1).png')
    auth_img = []
    if os.path.exists(stamp_path):
        auth_img.append(Image(stamp_path, width=35, height=35))
    if os.path.exists(sign_path):
        auth_img.append(Image(sign_path, width=35, height=35))
    if auth_img:
        auth_imgs = Table([auth_img], colWidths=[35]*len(auth_img))
        auth_imgs.setStyle(TableStyle([("VALIGN",(0,0),(-1,-1),"TOP"),("TOPPADDING",(0,0),(-1,-1),6),("BOTTOMPADDING",(0,0),(-1,-1),6),("ALIGN",(0,0),(-1,-1),"CENTER")]))
        els.append(auth_imgs)
    st = Table([
        [Paragraph("<br/><br/>___________________________<br/><b>Authorized Signatory</b><br/><font size=7 color='#64748b'>" + cn + "</font>", sg),
         Paragraph("<br/>", sg)],
    ], colWidths=[W*0.5, W*0.5])
    st.setStyle(TableStyle([("VALIGN",(0,0),(-1,-1),"TOP"),("LEFTPADDING",(0,0),(-1,-1),0),("RIGHTPADDING",(0,0),(-1,-1),0)]))
    els.append(st)

    doc.build(els)
    pdf_data = buf.getvalue()
    buf.close()
    for f in _logo_tmp_files:
        try: os.unlink(f)
        except: pass

    from flask import Response
    return Response(pdf_data, mimetype="application/pdf",
                    headers={"Content-Disposition": f'attachment; filename="{q_no}.pdf"'})

@customer_bp.route("/quotation/walkin", methods=["GET", "POST"])
def quotation_walkin():
    _ensure_tables()
    db = _get_db()
    next_no = _next_quotation_no(db)
    company = db.execute("SELECT * FROM company_profile LIMIT 1").fetchone()
    if request.method == "POST":
        q_date = request.form.get("quotation_date", date.today().isoformat())
        q_no = request.form.get("quotation_no", "").strip() or next_no
        location = request.form.get("location", "").strip()
        contact_details = request.form.get("contact_details", "").strip()
        customer_name = request.form.get("customer_name", "").strip() or "Walk-in Customer"
        customer_phone = request.form.get("customer_phone", "").strip() or None
        customer_email = request.form.get("customer_email", "").strip() or None
        vat_pct = float(request.form.get("vat_percent", 5))
        terms = request.form.get("terms", "").strip()
        notes = request.form.get("notes", "").strip()
        descs = request.form.getlist("item_desc[]")
        qtys = request.form.getlist("item_qty[]")
        rates = request.form.getlist("item_rate[]")
        units = request.form.getlist("item_unit[]")
        items = []
        sub_total = 0
        for i in range(len(descs)):
            desc = descs[i].strip()
            qty = float(qtys[i]) if i < len(qtys) and qtys[i].strip() else 1
            rate = float(rates[i]) if i < len(rates) and rates[i].strip() else 0
            unit = units[i].strip() if i < len(units) and units[i].strip() else "hr"
            if desc or rate > 0:
                amt = round(qty * rate, 2)
                sub_total += amt
                items.append({"desc": desc, "qty": qty, "rate": rate, "amt": amt, "unit": unit})
        if not items:
            flash("At least one line item is required.", "error")
            db.close()
            return render_template("customer/quotation_form.html", c=None, q={}, walkin=True, company=company, today=date.today().isoformat(), next_no=next_no)
        vat_amt = round(sub_total * vat_pct / 100, 2)
        total = round(sub_total + vat_amt, 2)
        # Create a temporary customer record for walk-in
        existing = db.execute("SELECT id FROM customers WHERE customer_name=? AND phone=?", (customer_name, customer_phone or "")).fetchone()
        if existing:
            cid = existing["id"]
        else:
            code = "WALKIN"
            cur = db.execute("INSERT INTO customers (customer_name,customer_code,phone,email,status) VALUES (?,?,?,?,?)",
                (customer_name, code, customer_phone, customer_email, "active"))
            cid = cur.lastrowid
            db.execute("UPDATE customers SET customer_code=? WHERE id=?", (f"WCI{cid:04d}", cid))
        cur = db.execute(
            """INSERT INTO customer_quotations (customer_id,quotation_no,quotation_date,sub_total,vat_percent,vat_amount,total_amount,status,terms,notes,location,contact_details)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?)""",
            (cid, q_no, q_date, sub_total, vat_pct, vat_amt, total, "pending", terms, notes, location, contact_details))
        qid = cur.lastrowid
        for idx, it in enumerate(items):
            db.execute("INSERT INTO customer_quotation_items (quotation_id,description,quantity,rate,amount,unit,sort_order) VALUES (?,?,?,?,?,?,?)",
                (qid, it["desc"], it["qty"], it["rate"], it["amt"], it["unit"], idx))
        db.commit()
        db.close()
        flash(f"Quotation {q_no} created for {customer_name}.", "success")
        return redirect(url_for("customer.customer_profile", cid=cid, tab="quotations"))
    db.close()
    return render_template("customer/quotation_form.html", c=None, q={}, walkin=True, company=company, today=date.today().isoformat(), next_no=next_no)

# ─── LPOs ───

@customer_bp.route("/<int:cid>/lpo/add", methods=["GET", "POST"])
def customer_lpo_add(cid):
    c = _get_customer_or_404(cid)
    if not c: return redirect(url_for("customer.customer_dashboard"))
    db = _get_db()
    _ensure_tables()
    if request.method == "POST":
        lpo_no = request.form.get("lpo_no", "").strip()
        lpo_date = request.form.get("lpo_date", date.today().isoformat())
        status = request.form.get("status", "pending")
        so_no = request.form.get("service_order_no", "").strip() or None
        notes = request.form.get("notes", "").strip()
        file_data = None; file_type = None
        file = request.files.get("lpo_file")
        if file and file.filename:
            file_data = base64.b64encode(file.read()).decode("utf-8")
            file_type = file.content_type
        descs = request.form.getlist("item_desc[]")
        qtys = request.form.getlist("item_qty[]")
        units = request.form.getlist("item_unit[]")
        rates = request.form.getlist("item_rate[]")
        vehicles = request.form.getlist("item_vehicle[]")
        total = 0; items = []
        for i in range(len(descs)):
            desc = descs[i].strip()
            qty = float(qtys[i]) if i < len(qtys) and qtys[i].strip() else 1
            unit = units[i] if i < len(units) else "hour"
            rate = float(rates[i]) if i < len(rates) and rates[i].strip() else 0
            vehicle = vehicles[i].strip().upper() if i < len(vehicles) else ""
            if desc or rate > 0:
                amt = round(qty * rate, 2)
                total += amt
                items.append({"desc": desc, "qty": qty, "rate": rate, "amt": amt, "unit": unit, "vehicle": vehicle})
        total = round(total, 2)
        try:
            cur = db.execute("INSERT INTO customer_lpos (customer_id,lpo_no,lpo_date,amount,status,service_order_no,notes,file_data,file_type) VALUES (?,?,?,?,?,?,?,?,?)",
                (cid, lpo_no, lpo_date, total, status, so_no, notes, file_data, file_type))
            lpo_id = cur.lastrowid
            for idx, it in enumerate(items):
                db.execute("INSERT INTO lpo_items (lpo_id,description,quantity,rate,amount,unit_type,vehicle_no,sort_order) VALUES (?,?,?,?,?,?,?,?)",
                    (lpo_id, it["desc"], it["qty"], it["rate"], it["amt"], it["unit"], it["vehicle"], idx))
            db.commit()
            db.close()
            flash("LPO added.", "success")
            return redirect(url_for("customer.customer_profile", cid=cid, tab="lpos"))
        except Exception as e:
            db.rollback()
            db.close()
            import traceback
            current_app.logger.error("LPO add failed: %s\n%s", e, traceback.format_exc())
            flash(f"Error adding LPO: {e}", "error")
            return render_template("customer/lpo_form.html", c=c, lpo={}, items=[], today=date.today().isoformat())
    db.close()
    return render_template("customer/lpo_form.html", c=c, lpo={}, items=[], today=date.today().isoformat())


@customer_bp.route("/<int:cid>/lpo/<int:lid>/edit", methods=["GET", "POST"])
def customer_lpo_edit(cid, lid):
    c = _get_customer_or_404(cid)
    if not c: return redirect(url_for("customer.customer_dashboard"))
    db = _get_db()
    _ensure_tables()
    lpo = db.execute("SELECT * FROM customer_lpos WHERE id=? AND customer_id=?", (lid, cid)).fetchone()
    if not lpo:
        db.close()
        flash("LPO not found.", "error")
        return redirect(url_for("customer.customer_profile", cid=cid, tab="lpos"))
    if request.method == "POST":
        lpo_no = request.form.get("lpo_no", "").strip()
        lpo_date = request.form.get("lpo_date", date.today().isoformat())
        status = request.form.get("status", "pending")
        so_no = request.form.get("service_order_no", "").strip() or None
        notes = request.form.get("notes", "").strip()
        file_data = lpo["file_data"]
        file_type = lpo["file_type"]
        file = request.files.get("lpo_file")
        if file and file.filename:
            file_data = base64.b64encode(file.read()).decode("utf-8")
            file_type = file.content_type
        descs = request.form.getlist("item_desc[]")
        qtys = request.form.getlist("item_qty[]")
        units = request.form.getlist("item_unit[]")
        rates = request.form.getlist("item_rate[]")
        vehicles = request.form.getlist("item_vehicle[]")
        total = 0
        items = []
        for i in range(len(descs)):
            desc = descs[i].strip()
            qty = float(qtys[i]) if i < len(qtys) and qtys[i].strip() else 1
            unit = units[i] if i < len(units) else "hour"
            rate = float(rates[i]) if i < len(rates) and rates[i].strip() else 0
            vehicle = vehicles[i].strip().upper() if i < len(vehicles) else ""
            if desc or rate > 0:
                amt = round(qty * rate, 2)
                total += amt
                items.append({"desc": desc, "qty": qty, "rate": rate, "amt": amt, "unit": unit, "vehicle": vehicle})
        total = round(total, 2)
        db.execute("UPDATE customer_lpos SET lpo_no=?,lpo_date=?,amount=?,status=?,service_order_no=?,notes=?,file_data=?,file_type=? WHERE id=?",
            (lpo_no, lpo_date, total, status, so_no, notes, file_data, file_type, lid))
        db.execute("DELETE FROM lpo_items WHERE lpo_id=?", (lid,))
        for idx, it in enumerate(items):
            db.execute("INSERT INTO lpo_items (lpo_id,description,quantity,rate,amount,unit_type,vehicle_no,sort_order) VALUES (?,?,?,?,?,?,?,?)",
                (lid, it["desc"], it["qty"], it["rate"], it["amt"], it["unit"], it["vehicle"], idx))
        db.commit()
        db.close()
        flash("LPO updated.", "success")
        return redirect(url_for("customer.customer_profile", cid=cid, tab="lpos"))
    items = db.execute("SELECT * FROM lpo_items WHERE lpo_id=? ORDER BY sort_order", (lid,)).fetchall()
    db.close()
    return render_template("customer/lpo_form.html", c=c, lpo=lpo, items=items, edit=True, today=date.today().isoformat())


@customer_bp.route("/<int:cid>/lpo/<int:lid>/delete", methods=["POST"])
def customer_lpo_delete(cid, lid):
    c = _get_customer_or_404(cid)
    if not c: return redirect(url_for("customer.customer_dashboard"))
    db = _get_db()
    lpo = db.execute("SELECT id FROM customer_lpos WHERE id=? AND customer_id=?", (lid, cid)).fetchone()
    if not lpo:
        db.close()
        flash("LPO not found.", "error")
        return redirect(url_for("customer.customer_profile", cid=cid, tab="lpos"))
    db.execute("DELETE FROM lpo_items WHERE lpo_id=?", (lid,))
    db.execute("DELETE FROM customer_lpos WHERE id=?", (lid,))
    db.commit()
    db.close()
    flash("LPO deleted.", "success")
    return redirect(url_for("customer.customer_profile", cid=cid, tab="lpos"))

@customer_bp.route("/<int:cid>/lpo/<int:lid>/close", methods=["POST"])
def customer_lpo_close(cid, lid):
    db = _get_db()
    db.execute("UPDATE customer_lpos SET status='closed' WHERE id=? AND customer_id=?", (lid, cid))
    db.commit()
    db.close()
    flash("LPO closed.", "success")
    return redirect(url_for("customer.customer_profile", cid=cid, tab="lpos"))

@customer_bp.route("/<int:cid>/lpo/<int:lid>/items")
def customer_lpo_items(cid, lid):
    _ensure_tables()
    db = _get_db()
    items = db.execute("SELECT id,description,quantity,rate,amount,unit_type,vehicle_no FROM lpo_items WHERE lpo_id=? ORDER BY sort_order", (lid,)).fetchall()
    lpo = db.execute("SELECT lpo_no,lpo_date FROM customer_lpos WHERE id=? AND customer_id=?", (lid, cid)).fetchone()
    db.close()
    if not lpo:
        return jsonify([])
    return jsonify({
        "lpo_no": lpo["lpo_no"],
        "lpo_date": lpo["lpo_date"],
        "items": [{"id": r["id"], "description": r["description"], "quantity": r["quantity"], "rate": r["rate"], "amount": r["amount"], "unit_type": r["unit_type"], "vehicle_no": r["vehicle_no"]} for r in items]
    })

@customer_bp.route("/<int:cid>/so/<int:sid>/items")
def customer_so_items(cid, sid):
    _ensure_tables()
    db = _get_db()
    items = db.execute("SELECT id,description,quantity,rate,amount,unit_type,vehicle_no FROM customer_so_items WHERE so_id=? ORDER BY sort_order", (sid,)).fetchall()
    so = db.execute("SELECT so_no,so_date FROM customer_service_orders WHERE id=? AND customer_id=?", (sid, cid)).fetchone()
    db.close()
    if not so:
        return jsonify([])
    return jsonify({
        "so_no": so["so_no"],
        "so_date": so["so_date"],
        "items": [{"id": r["id"], "description": r["description"], "quantity": r["quantity"], "rate": r["rate"], "amount": r["amount"], "unit_type": r["unit_type"], "vehicle_no": r["vehicle_no"]} for r in items]
    })

@customer_bp.route("/<int:cid>/lpo/<int:lid>/file")
def customer_lpo_file(cid, lid):
    db = _get_db()
    lpo = db.execute("SELECT * FROM customer_lpos WHERE id=? AND customer_id=?", (lid, cid)).fetchone()
    db.close()
    if not lpo or not lpo["file_data"]:
        flash("LPO file not found.", "error")
        return redirect(url_for("customer.customer_profile", cid=cid, tab="lpos"))
    import io
    data = base64.b64decode(lpo["file_data"])
    return send_file(io.BytesIO(data), mimetype=lpo["file_type"] or "application/pdf",
        as_attachment=True, download_name=f"LPO_{lpo['lpo_no'] or lid}.pdf")

@customer_bp.route("/<int:cid>/lpo/<int:lid>/view")
def customer_lpo_view(cid, lid):
    _ensure_tables()
    c = _get_customer_or_404(cid)
    if not c: return redirect(url_for("customer.customer_dashboard"))
    db = _get_db()
    lpo = db.execute("SELECT * FROM customer_lpos WHERE id=? AND customer_id=?", (lid, cid)).fetchone()
    if not lpo:
        db.close()
        flash("LPO not found.", "error")
        return redirect(url_for("customer.customer_profile", cid=cid, tab="lpos"))
    items = db.execute("SELECT * FROM lpo_items WHERE lpo_id=? ORDER BY sort_order", (lid,)).fetchall()
    db.close()
    return render_template("customer/lpo_view.html", c=c, lpo=lpo, items=items)

# ─── SERVICE ORDERS ───

@customer_bp.route("/<int:cid>/so/add", methods=["GET", "POST"])
def customer_so_add(cid):
    _ensure_tables()
    c = _get_customer_or_404(cid)
    if not c: return redirect(url_for("customer.customer_dashboard"))
    db = _get_db()
    if request.method == "POST":
        so_no = request.form.get("so_no", "").strip()
        so_date = request.form.get("so_date", date.today().isoformat())
        status = request.form.get("status", "pending")
        notes = request.form.get("notes", "").strip()
        descs = request.form.getlist("item_desc[]")
        qtys = request.form.getlist("item_qty[]")
        units = request.form.getlist("item_unit[]")
        rates = request.form.getlist("item_rate[]")
        total = 0; items = []
        for i in range(len(descs)):
            desc = descs[i].strip()
            qty = float(qtys[i]) if i < len(qtys) and qtys[i].strip() else 1
            unit = units[i] if i < len(units) else "hour"
            rate = float(rates[i]) if i < len(rates) and rates[i].strip() else 0
            if desc or rate > 0:
                amt = round(qty * rate, 2)
                total += amt
                items.append({"desc": desc, "qty": qty, "rate": rate, "amt": amt, "unit": unit})
        total = round(total, 2)
        cur = db.execute("INSERT INTO customer_service_orders (customer_id,so_no,so_date,amount,status,notes) VALUES (?,?,?,?,?,?)",
            (cid, so_no, so_date, total, status, notes))
        so_id = cur.lastrowid
        for idx, it in enumerate(items):
            db.execute("INSERT INTO customer_so_items (so_id,description,quantity,rate,amount,unit_type,vehicle_no,sort_order) VALUES (?,?,?,?,?,?,?,?)",
                (so_id, it["desc"], it["qty"], it["rate"], it["amt"], it["unit"], "", idx))
        db.commit()
        db.close()
        flash("Service Order added.", "success")
        return redirect(url_for("customer.customer_profile", cid=cid, tab="service_orders"))
    db.close()
    return render_template("customer/so_form.html", c=c, so={}, items=[], today=date.today().isoformat())

@customer_bp.route("/<int:cid>/so/<int:sid>/edit", methods=["GET", "POST"])
def customer_so_edit(cid, sid):
    _ensure_tables()
    c = _get_customer_or_404(cid)
    if not c: return redirect(url_for("customer.customer_dashboard"))
    db = _get_db()
    so = db.execute("SELECT * FROM customer_service_orders WHERE id=? AND customer_id=?", (sid, cid)).fetchone()
    if not so:
        db.close()
        flash("Service Order not found.", "error")
        return redirect(url_for("customer.customer_profile", cid=cid, tab="service_orders"))
    if request.method == "POST":
        so_no = request.form.get("so_no", "").strip()
        so_date = request.form.get("so_date", date.today().isoformat())
        status = request.form.get("status", "pending")
        notes = request.form.get("notes", "").strip()
        descs = request.form.getlist("item_desc[]")
        qtys = request.form.getlist("item_qty[]")
        units = request.form.getlist("item_unit[]")
        rates = request.form.getlist("item_rate[]")
        total = 0; items = []
        for i in range(len(descs)):
            desc = descs[i].strip()
            qty = float(qtys[i]) if i < len(qtys) and qtys[i].strip() else 1
            unit = units[i] if i < len(units) else "hour"
            rate = float(rates[i]) if i < len(rates) and rates[i].strip() else 0
            if desc or rate > 0:
                amt = round(qty * rate, 2)
                total += amt
                items.append({"desc": desc, "qty": qty, "rate": rate, "amt": amt, "unit": unit})
        total = round(total, 2)
        db.execute("UPDATE customer_service_orders SET so_no=?,so_date=?,amount=?,status=?,notes=? WHERE id=?",
            (so_no, so_date, total, status, notes, sid))
        db.execute("DELETE FROM customer_so_items WHERE so_id=?", (sid,))
        for idx, it in enumerate(items):
            db.execute("INSERT INTO customer_so_items (so_id,description,quantity,rate,amount,unit_type,vehicle_no,sort_order) VALUES (?,?,?,?,?,?,?,?)",
                (sid, it["desc"], it["qty"], it["rate"], it["amt"], it["unit"], it["vehicle"], idx))
        db.commit()
        db.close()
        flash("Service Order updated.", "success")
        return redirect(url_for("customer.customer_profile", cid=cid, tab="service_orders"))
    items = db.execute("SELECT * FROM customer_so_items WHERE so_id=? ORDER BY sort_order", (sid,)).fetchall()
    db.close()
    return render_template("customer/so_form.html", c=c, so=so, items=items, edit=True, today=date.today().isoformat())

@customer_bp.route("/<int:cid>/so/<int:sid>/delete", methods=["POST"])
def customer_so_delete(cid, sid):
    c = _get_customer_or_404(cid)
    if not c: return redirect(url_for("customer.customer_dashboard"))
    db = _get_db()
    db.execute("DELETE FROM customer_so_items WHERE so_id=?", (sid,))
    db.execute("DELETE FROM customer_service_orders WHERE id=? AND customer_id=?", (sid, cid))
    db.commit(); db.close()
    flash("Service Order deleted.", "success")
    return redirect(url_for("customer.customer_profile", cid=cid, tab="service_orders"))

@customer_bp.route("/<int:cid>/so/<int:sid>/view")
def customer_so_view(cid, sid):
    _ensure_tables()
    c = _get_customer_or_404(cid)
    if not c: return redirect(url_for("customer.customer_dashboard"))
    db = _get_db()
    so = db.execute("SELECT * FROM customer_service_orders WHERE id=? AND customer_id=?", (sid, cid)).fetchone()
    if not so:
        db.close()
        flash("Service Order not found.", "error")
        return redirect(url_for("customer.customer_profile", cid=cid, tab="service_orders"))
    items = db.execute("SELECT * FROM customer_so_items WHERE so_id=? ORDER BY sort_order", (sid,)).fetchall()
    db.close()
    return render_template("customer/so_view.html", c=c, so=so, items=items)

# ─── DOCUMENTS ───

@customer_bp.route("/<int:cid>/doc/add", methods=["GET", "POST"])
def customer_doc_add(cid):
    c = _get_customer_or_404(cid)
    if not c: return redirect(url_for("customer.customer_dashboard"))
    if request.method == "POST":
        doc_type = request.form.get("doc_type", "Other")
        doc_name = request.form.get("doc_name", "").strip()
        expiry = request.form.get("expiry_date", "").strip() or None
        file = request.files.get("file")
        file_data = None
        file_type = None
        if file and file.filename:
            file_data = base64.b64encode(file.read()).decode("utf-8")
            file_type = file.content_type
        db = _get_db()
        db.execute("INSERT INTO customer_documents (customer_id,doc_type,doc_name,file_data,file_type,expiry_date) VALUES (?,?,?,?,?,?)",
            (cid, doc_type, doc_name, file_data, file_type, expiry))
        db.commit()
        db.close()
        flash("Document uploaded.", "success")
        return redirect(url_for("customer.customer_profile", cid=cid, tab="documents"))
    return render_template("customer/doc_form.html", c=c)

@customer_bp.route("/<int:cid>/doc/<int:did>/download")
def customer_doc_download(cid, did):
    db = _get_db()
    doc = db.execute("SELECT * FROM customer_documents WHERE id=? AND customer_id=?", (did, cid)).fetchone()
    db.close()
    if not doc or not doc["file_data"]:
        flash("Document not found.", "error")
        return redirect(url_for("customer.customer_profile", cid=cid, tab="documents"))
    import io
    data = base64.b64decode(doc["file_data"])
    return send_file(io.BytesIO(data), mimetype=doc["file_type"] or "application/octet-stream",
        as_attachment=True, download_name=doc["doc_name"] or f"doc_{did}")

@customer_bp.route("/<int:cid>/doc/<int:did>/delete", methods=["POST"])
def customer_doc_delete(cid, did):
    db = _get_db()
    db.execute("DELETE FROM customer_documents WHERE id=? AND customer_id=?", (did, cid))
    db.commit()
    db.close()
    flash("Document deleted.", "success")
    return redirect(url_for("customer.customer_profile", cid=cid, tab="documents"))

# ─── SOA / STATEMENT ───

@customer_bp.route("/<int:cid>/soa")
def customer_soa(cid):
    _ensure_tables()
    c = _get_customer_or_404(cid)
    if not c: return redirect(url_for("customer.customer_dashboard"))
    from_date = request.args.get("from", "")
    to_date = request.args.get("to", "")
    db = _get_db()
    entries = []
    inv_q = """SELECT i.id, i.invoice_date as d, i.invoice_no as ref, i.total_amount as dr,
                      COALESCE((SELECT SUM(p2.amount) FROM customer_payments p2 WHERE p2.invoice_id = i.id),0)
                      + COALESCE((SELECT SUM(cn2.total_amount) FROM customer_credit_notes cn2 WHERE cn2.invoice_id = i.id),0) as cr
               FROM customer_invoices i WHERE i.customer_id=?"""
    inv_p = [cid]
    if from_date: inv_q += " AND i.invoice_date>=?"; inv_p.append(from_date)
    if to_date: inv_q += " AND i.invoice_date<=?"; inv_p.append(to_date)
    inv_q += " ORDER BY i.invoice_date, i.id"
    for inv in db.execute(inv_q, inv_p).fetchall():
        d = dict(inv)
        d["type"] = "Invoice"
        if (d.get("dr",0) or 0) > 0 and (d.get("dr",0) or 0) - (d.get("cr",0) or 0) <= 0.005:
            continue
        entries.append(d)
    unalloc_pmt_q = "SELECT payment_date as d, reference_no as ref, 'Unallocated Payment' as type, 0 as dr, amount as cr FROM customer_payments WHERE customer_id=? AND invoice_id IS NULL"
    unalloc_pmt_p = [cid]
    if from_date: unalloc_pmt_q += " AND payment_date>=?"; unalloc_pmt_p.append(from_date)
    if to_date: unalloc_pmt_q += " AND payment_date<=?"; unalloc_pmt_p.append(to_date)
    unalloc_pmt_q += " ORDER BY payment_date"
    for pmt in db.execute(unalloc_pmt_q, unalloc_pmt_p).fetchall():
        entries.append(dict(pmt))
    entries.sort(key=lambda x: (x.get("d",""), x.get("type","")))
    balance = 0
    for e in entries:
        balance += (e.get("dr",0) or 0) - (e.get("cr",0) or 0)
        e["bal"] = round(balance, 2)
    db.close()
    return render_template("customer/kata.html", c=c, entries=entries, from_date=from_date, to_date=to_date)

@customer_bp.route("/<int:cid>/soa/pdf")
def customer_soa_pdf(cid):
    _ensure_tables()
    import tempfile
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
    from io import BytesIO

    c = _get_customer_or_404(cid)
    if not c: return redirect(url_for("customer.customer_dashboard"))
    from_date = request.args.get("from", "")
    to_date = request.args.get("to", "")
    db = _get_db()
    company = db.execute("SELECT * FROM company_profile LIMIT 1").fetchone()
    entries = []
    inv_q = """SELECT i.id, i.invoice_date as d, i.invoice_no as ref, i.total_amount as dr,
                      COALESCE((SELECT SUM(p2.amount) FROM customer_payments p2 WHERE p2.invoice_id = i.id),0)
                      + COALESCE((SELECT SUM(cn2.total_amount) FROM customer_credit_notes cn2 WHERE cn2.invoice_id = i.id),0) as cr
               FROM customer_invoices i WHERE i.customer_id=?"""
    inv_p = [cid]
    if from_date: inv_q += " AND i.invoice_date>=?"; inv_p.append(from_date)
    if to_date: inv_q += " AND i.invoice_date<=?"; inv_p.append(to_date)
    inv_q += " ORDER BY i.invoice_date, i.id"
    for inv in db.execute(inv_q, inv_p).fetchall():
        d = dict(inv)
        d["type"] = "Invoice"
        if (d.get("dr",0) or 0) > 0 and (d.get("dr",0) or 0) - (d.get("cr",0) or 0) <= 0.005:
            continue
        entries.append(d)
    unalloc_pmt_q = "SELECT payment_date as d, reference_no as ref, 'Unallocated Payment' as type, 0 as dr, amount as cr FROM customer_payments WHERE customer_id=? AND invoice_id IS NULL"
    unalloc_pmt_p = [cid]
    if from_date: unalloc_pmt_q += " AND payment_date>=?"; unalloc_pmt_p.append(from_date)
    if to_date: unalloc_pmt_q += " AND payment_date<=?"; unalloc_pmt_p.append(to_date)
    unalloc_pmt_q += " ORDER BY payment_date"
    for pmt in db.execute(unalloc_pmt_q, unalloc_pmt_p).fetchall():
        entries.append(dict(pmt))
    entries.sort(key=lambda x: (x.get("d",""), x.get("type","")))
    bal = 0
    for e in entries:
        bal += (e.get("dr",0) or 0) - (e.get("cr",0) or 0)
        e["bal"] = round(bal, 2)
    db.close()
    total_dr = sum(e.get("dr",0) or 0 for e in entries)
    total_cr = sum(e.get("cr",0) or 0 for e in entries)
    closing = round(total_dr - total_cr, 2)

    _logo_tmp_files = []
    buf = BytesIO()
    LM, RM, TM, BM = 18*mm, 18*mm, 15*mm, 15*mm
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=LM, rightMargin=RM, topMargin=TM, bottomMargin=BM)
    W = A4[0] - LM - RM

    tc = company["theme_color"] or "#1a3a5c" if company else "#1a3a5c"
    try: TH = colors.HexColor(tc)
    except: TH = colors.HexColor("#1a3a5c")
    BG = colors.HexColor("#f4f6f9"); WH = colors.white; C3 = colors.HexColor("#d1d5db")
    C4 = colors.HexColor("#111827"); C5 = colors.HexColor("#6b7280"); CG = colors.HexColor("#1a7d1a")
    CR = colors.HexColor("#c62828")

    def F(name, **kw):
        kw.setdefault("fontSize", 8); kw.setdefault("leading", 12)
        return ParagraphStyle(name, **kw)

    def C(t, **kw):
        kw.setdefault("alignment", TA_CENTER)
        return Paragraph(str(t), F("_C", **kw))
    def R(t, **kw):
        kw.setdefault("alignment", TA_RIGHT)
        return Paragraph(str(t), F("_R", **kw))
    def L(t, **kw):
        kw.setdefault("textColor", C5)
        return Paragraph(str(t), F("_L", **kw))

    els = []
    cn = company["company_name"] if company else "COMPANY"
    trn = company["trn_no"] or "—" if company else "—"

    # ══════════════════════════════
    # HEADER (matches invoice style)
    # ══════════════════════════════
    logo = None; LW = 0
    if company and company["logo_data"]:
        try:
            lb = base64.b64decode(company["logo_data"])
            f = tempfile.NamedTemporaryFile(delete=False, suffix=".png")
            f.write(lb); f.close()
            from PIL import Image as PILImage
            with PILImage.open(f.name) as img:
                ow, oh = img.size
            max_h = 38
            ratio = max_h / oh
            lw = int(ow * ratio)
            lh = int(max_h)
            logo = Image(f.name, width=lw, height=lh)
            LW = lw
            _logo_tmp_files.append(f.name)
        except: pass

    cl = [f"<font size=11><b>{cn}</b></font>"]
    addr = company["address"] or ""; ph = company["phone_number"] or ""; em = company["email"] or ""
    parts = [x for x in [addr] if x]
    cparts = [x for x in [ph, em, f"TRN: {trn}"] if x and x != f"TRN: —"]
    if parts or cparts:
        info = " &middot; ".join(parts + cparts)
        cl.append(f"<font size=6.5 color='#6b7280'>{info}</font>")
    co_p = Paragraph("<br/>".join(cl), F("CO", fontSize=11, fontName="Helvetica-Bold", textColor=TH, leading=13))
    if logo:
        lh = Table([[logo, Spacer(1, 3*mm), co_p]], colWidths=[LW, 3*mm, W*0.65 - LW - 3*mm])
        lh.setStyle(TableStyle([("VALIGN",(0,0),(-1,-1),"MIDDLE"),("LEFTPADDING",(0,0),(-1,-1),0),("RIGHTPADDING",(0,0),(-1,-1),0)]))
    else:
        lh = co_p
    rh = Paragraph(
        f"<b>STATEMENT<br/>OF ACCOUNT</b>",
        F("TI", fontSize=14, fontName="Helvetica-Bold", textColor=TH, leading=18, alignment=TA_RIGHT))
    ht = Table([[lh, rh]], colWidths=[W*0.65, W*0.35])
    ht.setStyle(TableStyle([("VALIGN",(0,0),(-1,-1),"TOP"),("LEFTPADDING",(0,0),(-1,-1),0),("RIGHTPADDING",(0,0),(-1,-1),0)]))
    els.append(ht)
    els.append(Spacer(1, 2*mm))
    hr = Table([[""]], colWidths=[W], rowHeights=[2])
    hr.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,-1),TH),("LEFTPADDING",(0,0),(-1,-1),0),("RIGHTPADDING",(0,0),(-1,-1),0)]))
    els.append(hr)
    els.append(Spacer(1, 4*mm))

    # ══════════════════════════════
    # CUSTOMER INFO
    # ══════════════════════════════
    cinfo = [
        [Paragraph("<b>Customer</b>", F("_cl", fontSize=8, fontName="Helvetica-Bold", textColor=C4, leading=11)),
         Paragraph(f"<b>{c['customer_name']}</b>", F("_cv", fontSize=9, fontName="Helvetica-Bold", textColor=C4, leading=12))],
    ]
    if c["trn"]: cinfo.append([Paragraph("TRN", F("_l", fontSize=7.5, textColor=C5, leading=10)), Paragraph(c["trn"], F("_v", fontSize=8.5, textColor=C4, leading=11))])
    if c["address"]: cinfo.append([Paragraph("Address", F("_l", fontSize=7.5, textColor=C5, leading=10)), Paragraph(c["address"], F("_v", fontSize=8.5, textColor=C4, leading=11))])
    if c["phone"]: cinfo.append([Paragraph("Phone", F("_l", fontSize=7.5, textColor=C5, leading=10)), Paragraph(c["phone"], F("_v", fontSize=8.5, textColor=C4, leading=11))])
    ct = Table(cinfo, colWidths=[50, W - 50])
    ct.setStyle(TableStyle([("VALIGN",(0,0),(-1,-1),"TOP"),("TOPPADDING",(0,0),(-1,-1),1),("BOTTOMPADDING",(0,0),(-1,-1),1),("LEFTPADDING",(0,0),(-1,-1),0),("RIGHTPADDING",(0,0),(-1,-1),0)]))
    els.append(ct)

    # ══════════════════════════════
    # SUMMARY CARDS
    # ══════════════════════════════
    els.append(Spacer(1, 3*mm))
    sdata = [[
        Paragraph(f"<b>Total Invoiced</b><br/><font size=10 color='#1a3a5c'>AED {total_dr:,.2f}</font>", F("_s1", fontSize=7, textColor=C5, alignment=TA_CENTER, leading=10)),
        Paragraph(f"<b>Total Paid</b><br/><font size=10 color='#1a7d1a'>AED {total_cr:,.2f}</font>", F("_s2", fontSize=7, textColor=C5, alignment=TA_CENTER, leading=10)),
        Paragraph(f"<b>Outstanding</b><br/><font size=10 color='#c62828'>AED {closing if closing > 0 else 0:,.2f}</font>", F("_s3", fontSize=7, textColor=C5, alignment=TA_CENTER, leading=10)),
        Paragraph(f"<b>Transactions</b><br/><font size=10>{len(entries)}</font>", F("_s4", fontSize=7, textColor=C5, alignment=TA_CENTER, leading=10)),
    ]]
    st = Table(sdata, colWidths=[W/4, W/4, W/4, W/4])
    st.setStyle(TableStyle([
        ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
        ("BOX",(0,0),(-1,-1),0.5,C3), ("INNERGRID",(0,0),(-1,-1),0.3,C3),
        ("TOPPADDING",(0,0),(-1,-1),8), ("BOTTOMPADDING",(0,0),(-1,-1),8),
        ("LEFTPADDING",(0,0),(-1,-1),5), ("RIGHTPADDING",(0,0),(-1,-1),5),
        ("BACKGROUND",(0,0),(-1,-1),BG),
    ]))
    els.append(st)
    els.append(Spacer(1, 3*mm))

    if from_date or to_date:
        rng = f"Period: {from_date or '…'} to {to_date or '…'}"
        els.append(Paragraph(
            f"<font size=7 color='#6b7280'>{rng}</font>",
            F("_pr", fontSize=7, textColor=C5, alignment=TA_CENTER, leading=9)))
        els.append(Spacer(1, 2*mm))

    # ══════════════════════════════
    # STATEMENT TABLE
    # ══════════════════════════════
    colw = [45, 38, 65, 38, W - 45 - 38 - 65 - 38 - 65 - 75, 65, 75]
    hdr = [
        Paragraph("<b>Date</b>", F("_h", fontSize=7, fontName="Helvetica-Bold", textColor=WH, alignment=TA_CENTER, leading=10)),
        Paragraph("<b>Month</b>", F("_h", fontSize=7, fontName="Helvetica-Bold", textColor=WH, alignment=TA_CENTER, leading=10)),
        Paragraph("<b>Invoice #</b>", F("_h", fontSize=7, fontName="Helvetica-Bold", textColor=WH, leading=10)),
        Paragraph("<b>Type</b>", F("_h", fontSize=7, fontName="Helvetica-Bold", textColor=WH, alignment=TA_CENTER, leading=10)),
        Paragraph("<b>Dr (AED)</b>", F("_h", fontSize=7, fontName="Helvetica-Bold", textColor=WH, alignment=TA_RIGHT, leading=10)),
        Paragraph("<b>Cr (AED)</b>", F("_h", fontSize=7, fontName="Helvetica-Bold", textColor=WH, alignment=TA_RIGHT, leading=10)),
        Paragraph("<b>Balance (AED)</b>", F("_h", fontSize=7, fontName="Helvetica-Bold", textColor=WH, alignment=TA_RIGHT, leading=10)),
    ]
    rws = [hdr]
    rws.append([
        Paragraph("", F("_o", fontSize=7, leading=10)), Paragraph("", F("_o")),
        Paragraph("", F("_o")), Paragraph("Opening Balance", F("_ol", fontSize=7, textColor=C5, leading=10)),
        Paragraph("", F("_o")), Paragraph("", F("_o")),
        Paragraph("<b>0.00</b>", F("_ob", fontSize=7, fontName="Helvetica-Bold", textColor=C4, alignment=TA_RIGHT, leading=10)),
    ])
    for e in entries:
        d = str(e.get("d",""))
        month = d[:7] if d and len(d) >= 7 else ""
        bal_val = e.get("bal",0) or 0
        bal_display = "0.00" if bal_val <= 0 else f"{bal_val:,.2f}"
        bal_color = "#c62828" if bal_val > 0 else "#1a7d1a"
        rws.append([
            Paragraph(d, F("_d", fontSize=7, leading=10)),
            Paragraph(f"<font color='{C5}'>{month}</font>" if month else "", F("_m", fontSize=6.5, textColor=C5, leading=10)),
            Paragraph(str(e.get("ref","—")), F("_r", fontSize=7, fontName="Helvetica-Bold", textColor=C4, leading=10)),
            Paragraph(f"<font color=\"{'#1a56db' if e['type']=='Invoice' else '#e65100' if e['type']=='Credit Note' else '#c62828' if e['type']=='Unallocated Payment' else '#1a7d1a'}\">{e['type']}</font>", F("_t", fontSize=7, alignment=TA_CENTER, leading=10)),
            Paragraph(f"<b>{e.get('dr',0) or 0:,.2f}</b>" if e.get("dr") else '<font color="#cccccc">—</font>', F("_dr", fontSize=7, textColor="#c62828" if e.get("dr") else C5, alignment=TA_RIGHT, leading=10)),
            Paragraph(f"<b>{e.get('cr',0) or 0:,.2f}</b>" if e.get("cr") else '<font color="#cccccc">—</font>', F("_cr", fontSize=7, textColor="#1a7d1a" if e.get("cr") else C5, alignment=TA_RIGHT, leading=10)),
            Paragraph(f"<b>{bal_display}</b>", F("_bl", fontSize=7, fontName="Helvetica-Bold", textColor=bal_color, alignment=TA_RIGHT, leading=10)),
        ])

    # ── CLOSING ROW (inside main table for perfect alignment) ──
    rws.append([
        Paragraph("<b>Closing Balance</b>", F("_cb", fontSize=8, fontName="Helvetica-Bold", textColor=WH, leading=11)),
        Paragraph("", F("_x", fontSize=7, leading=10)),
        Paragraph("", F("_x", fontSize=7, leading=10)),
        Paragraph("", F("_x", fontSize=7, leading=10)),
        Paragraph(f"<b>{total_dr:,.2f}</b>", F("_cd", fontSize=8, fontName="Helvetica-Bold", textColor=WH, alignment=TA_RIGHT, leading=11)),
        Paragraph(f"<b>{total_cr:,.2f}</b>", F("_cc", fontSize=8, fontName="Helvetica-Bold", textColor=WH, alignment=TA_RIGHT, leading=11)),
        Paragraph(f"<b>{(closing if closing > 0 else 0):,.2f}</b>", F("_ccl", fontSize=8, fontName="Helvetica-Bold", textColor=WH, alignment=TA_RIGHT, leading=11)),
    ])

    it = Table(rws, colWidths=colw, repeatRows=1)
    it.setStyle(TableStyle([
        ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
        ("BACKGROUND",(0,0),(-1,0),TH), ("TEXTCOLOR",(0,0),(-1,0),WH),
        ("BOX",(0,0),(-1,-1),0.5,C3), ("INNERGRID",(0,0),(-1,-1),0.3,C3),
        ("TOPPADDING",(0,0),(-1,-1),2), ("BOTTOMPADDING",(0,0),(-1,-1),2),
        ("LEFTPADDING",(0,0),(-1,-1),3), ("RIGHTPADDING",(0,0),(-1,-1),3),
        ("BACKGROUND",(0,-1),(-1,-1),TH), ("TEXTCOLOR",(0,-1),(-1,-1),WH),
        ("FONTNAME",(0,-1),(-1,-1),"Helvetica-Bold"),
        ("ROWBACKGROUNDS",(0,1),(-2,-2),[WH, BG]),
    ]))
    els.append(it)

    # ══════════════════════════════
    # SIGNATURES (SOA)
    # ══════════════════════════════
    els.append(Spacer(1, 8*mm))
    s_sg = ParagraphStyle("SSG", fontSize=9, alignment=TA_CENTER, leading=14)
    s_stamp_path = os.path.join(current_app.root_path, 'static', 'Stamp.png')
    s_sign_path = os.path.join(current_app.root_path, 'static', 'Sign (1).png')
    s_auth_cells = []
    s_auth_cells.append(Paragraph("_________________________", s_sg))
    if os.path.exists(s_stamp_path):
        s_auth_cells.append(Image(s_stamp_path, width=40, height=40))
    if os.path.exists(s_sign_path):
        s_auth_cells.append(Image(s_sign_path, width=40, height=40))
    s_auth_cells.append(Paragraph("<b>Authorized Signatory</b>", s_sg))
    s_auth_cell = Table([[c] for c in s_auth_cells], colWidths=[W*0.35])
    s_auth_cell.setStyle(TableStyle([
        ("ALIGN",(0,0),(-1,-1),"CENTER"),
        ("VALIGN",(0,0),(-1,-1),"MIDDLE"),
        ("TOPPADDING",(0,0),(-1,-1),0),
        ("BOTTOMPADDING",(0,0),(-1,-1),2),
    ]))
    soa_sig = Table([[
        s_auth_cell,
        C("", fontSize=4),
        Paragraph("", s_sg),
    ]], colWidths=[W*0.35, W*0.30, W*0.35])
    soa_sig.setStyle(TableStyle([
        ("VALIGN",(0,0),(-1,-1),"TOP"),
        ("LINEABOVE",(0,0),(0,0),0.5,C5), ("LINEABOVE",(2,0),(2,0),0.5,C5),
        ("LEFTPADDING",(0,0),(-1,-1),0), ("RIGHTPADDING",(0,0),(-1,-1),0),
    ]))
    els.append(soa_sig)

    # ══════════════════════════════
    # FOOTER
    # ══════════════════════════════
    els.append(Spacer(1, 8*mm))
    fh = Table([[""]], colWidths=[W], rowHeights=[0.5])
    fh.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,-1),TH),("LEFTPADDING",(0,0),(-1,-1),0),("RIGHTPADDING",(0,0),(-1,-1),0)]))
    els.append(fh)
    els.append(Spacer(1, 2*mm))
    ft_txt = "This is a computer-generated Statement of Account."
    if from_date or to_date:
        rng = f"Period: {from_date or '…'} to {to_date or '…'}"
        ft_txt += f" | {rng}"
    els.append(Paragraph(ft_txt, F("_ft", fontSize=7, textColor=C5, alignment=TA_CENTER, leading=9)))

    doc.build(els)
    for f in _logo_tmp_files:
        try: os.remove(f)
        except: pass
    pdf_data = buf.getvalue(); buf.close()
    return send_file(BytesIO(pdf_data), mimetype="application/pdf", as_attachment=True, download_name=f"SOA_{c['customer_name']}.pdf")

# ─── LIST ───

@customer_bp.route("/list")
def customer_list():
    _ensure_tables()
    db = _get_db()
    search = request.args.get("search", "").strip()
    if search:
        customers = db.execute("SELECT * FROM customers WHERE customer_name LIKE ? OR phone LIKE ? OR email LIKE ? ORDER BY customer_name",
            (f"%{search}%", f"%{search}%", f"%{search}%")).fetchall()
    else:
        customers = db.execute("SELECT * FROM customers ORDER BY customer_name").fetchall()
    db.close()
    return render_template("customer/list.html", customers=customers, search=search)


# ═══════════════════════════════════════════════════════════
# SETTINGS
# ═══════════════════════════════════════════════════════════

@customer_bp.route("/tax-report")
def customer_tax_report():
    _ensure_tables()
    db = _get_db()
    from_filter = request.args.get("from", "")
    to_filter = request.args.get("to", "")
    where = ""
    params = []
    if from_filter:
        where += " AND i.invoice_date >= ?"
        params.append(from_filter)
    if to_filter:
        where += " AND i.invoice_date <= ?"
        params.append(to_filter)
    customers = db.execute(f"""
        SELECT c.id, c.customer_name, c.trn, c.customer_code, c.status,
               COUNT(i.id) AS invoice_count,
               COALESCE(SUM(i.amount),0) AS total_taxable,
               COALESCE(SUM(i.vat_amount),0) AS total_vat,
               COALESCE(SUM(i.total_amount),0) AS total_invoiced
        FROM customers c
        LEFT JOIN customer_invoices i ON i.customer_id = c.id {where}
        GROUP BY c.id, c.customer_name, c.trn, c.customer_code, c.status
        ORDER BY c.customer_name
    """, params).fetchall()
    invoices = db.execute(f"""
        SELECT i.invoice_date, i.invoice_no, c.customer_name,
               i.amount AS net_sale, i.vat_amount, i.total_amount
        FROM customer_invoices i
        JOIN customers c ON c.id = i.customer_id
        WHERE 1=1 {where.replace('i.','')}
        ORDER BY i.invoice_date DESC, i.invoice_no DESC
    """, params).fetchall()
    invoices_where = ""
    invoices_params = []
    if from_filter:
        invoices_where += " AND i.invoice_date >= ?"
        invoices_params.append(from_filter)
    if to_filter:
        invoices_where += " AND i.invoice_date <= ?"
        invoices_params.append(to_filter)
    invoices = db.execute(f"""
        SELECT i.invoice_date, i.invoice_no, c.customer_name,
               i.amount AS net_sale, i.vat_amount, i.total_amount
        FROM customer_invoices i
        JOIN customers c ON c.id = i.customer_id
        WHERE 1=1 {invoices_where}
        ORDER BY i.invoice_date DESC, i.invoice_no DESC
    """, invoices_params).fetchall()
    total_taxable = sum(r["total_taxable"] for r in customers)
    total_vat = sum(r["total_vat"] for r in customers)
    total_invoiced = sum(r["total_invoiced"] for r in customers)
    db.close()
    return render_template(
        "customer/tax_report.html",
        customers=customers,
        invoices=invoices,
        total_taxable=total_taxable,
        total_vat=total_vat,
        total_invoiced=total_invoiced,
        from_filter=from_filter,
        to_filter=to_filter,
    )

@customer_bp.route("/tax-report/export/excel")
def customer_tax_report_excel():
    _ensure_tables()
    db = _get_db()
    from_filter = request.args.get("from", "")
    to_filter = request.args.get("to", "")
    where = ""; params = []
    if from_filter: where += " AND i.invoice_date >= ?"; params.append(from_filter)
    if to_filter: where += " AND i.invoice_date <= ?"; params.append(to_filter)
    invoices = db.execute(f"""
        SELECT i.invoice_date, i.invoice_no, c.customer_name,
               i.amount AS net_sale, i.vat_amount, i.total_amount
        FROM customer_invoices i
        JOIN customers c ON c.id = i.customer_id
        WHERE 1=1 {where.replace('i.','')}
        ORDER BY i.invoice_date DESC, i.invoice_no DESC
    """, params).fetchall()
    total_taxable = sum(r["net_sale"] or 0 for r in invoices)
    total_vat = sum(r["vat_amount"] or 0 for r in invoices)
    total_invoiced = sum(r["total_amount"] or 0 for r in invoices)
    db.close()

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from io import BytesIO
    wb = Workbook()
    ws = wb.active
    ws.title = "Tax Report"

    hf = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    hfill = PatternFill("solid", fgColor="1a3a5c")
    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right", vertical="center")
    thin = Side(style="thin", color="d8e4f5")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)

    heads = ["Date", "Invoice No", "Customer", "Net Sale (AED)", "VAT (AED)", "Gross Sale (AED)"]
    for ci, h in enumerate(heads, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = hf; c.fill = hfill; c.alignment = center; c.border = border

    period = f"{from_filter or 'All'} to {to_filter or 'All'}"
    ws.cell(row=2, column=1, value=f"Period: {period}").font = Font(italic=True, color="6b7280", size=10)

    for ri, inv in enumerate(invoices, 3):
        vals = [inv["invoice_date"], inv["invoice_no"], inv["customer_name"],
                inv["net_sale"] or 0, inv["vat_amount"] or 0, inv["total_amount"] or 0]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.border = border
            if ci >= 4: c.alignment = right; c.number_format = '#,##0.00'
            if ci == 5: c.font = Font(color="f7931e")

    tr = 3 + len(invoices)
    totals = ["", "", "TOTALS", total_taxable, total_vat, total_invoiced]
    tf = Font(bold=True, size=11)
    tfill = PatternFill("solid", fgColor="f5f8fe")
    for ci, v in enumerate(totals, 1):
        c = ws.cell(row=tr, column=ci, value=v)
        c.font = tf; c.fill = tfill; c.border = border
        if ci >= 4: c.alignment = right; c.number_format = '#,##0.00'

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 30
    for col in ["D","E","F"]: ws.column_dimensions[col].width = 18

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    fn = f"Tax_Report_{from_filter or 'start'}_to_{to_filter or 'end'}.xlsx"
    return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=fn)

@customer_bp.route("/tax-report/export/pdf")
def customer_tax_report_pdf():
    _ensure_tables()
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
    from io import BytesIO

    db = _get_db()
    from_filter = request.args.get("from", "")
    to_filter = request.args.get("to", "")
    where = ""; params = []
    if from_filter: where += " AND i.invoice_date >= ?"; params.append(from_filter)
    if to_filter: where += " AND i.invoice_date <= ?"; params.append(to_filter)
    company = db.execute("SELECT * FROM company_profile LIMIT 1").fetchone()
    invoices = db.execute(f"""
        SELECT i.invoice_date, i.invoice_no, c.customer_name,
               i.amount AS net_sale, i.vat_amount, i.total_amount
        FROM customer_invoices i
        JOIN customers c ON c.id = i.customer_id
        WHERE 1=1 {where.replace('i.','')}
        ORDER BY i.invoice_date DESC, i.invoice_no DESC
    """, params).fetchall()
    total_taxable = sum(r["net_sale"] or 0 for r in invoices)
    total_vat = sum(r["vat_amount"] or 0 for r in invoices)
    total_invoiced = sum(r["total_amount"] or 0 for r in invoices)
    db.close()

    buf = BytesIO()
    LM, RM, TM, BM = 15*mm, 15*mm, 15*mm, 15*mm
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=LM, rightMargin=RM, topMargin=TM, bottomMargin=BM)
    W = landscape(A4)[0] - LM - RM

    tc = company["theme_color"] or "#1a3a5c" if company else "#1a3a5c"
    try: TH = colors.HexColor(tc)
    except: TH = colors.HexColor("#1a3a5c")
    WH = colors.white; C5 = colors.HexColor("#6b7280")
    CR = colors.HexColor("#c62828")

    def F(name, **kw):
        kw.setdefault("fontSize", 7); kw.setdefault("leading", 10)
        return ParagraphStyle(name, **kw)

    els = []
    cn = company["company_name"] if company else "Tax Report"
    els.append(Paragraph(f"<b>{cn}</b>", F("T", fontSize=12, textColor=TH, spaceAfter=2)))
    period = f"Period: {from_filter or 'Start'} to {to_filter or 'End'}"
    els.append(Paragraph(period, F("P", fontSize=7, textColor=C5, spaceAfter=10)))
    els.append(Spacer(1, 3*mm))

    hdr = ["Date", "Invoice No", "Customer", "Net Sale", "VAT", "Gross Sale"]
    data = [hdr]
    for inv in invoices:
        data.append([
            inv["invoice_date"] or "—",
            inv["invoice_no"] or "—",
            inv["customer_name"],
            f"{inv['net_sale'] or 0:,.2f}",
            f"{inv['vat_amount'] or 0:,.2f}",
            f"{inv['total_amount'] or 0:,.2f}",
        ])

    data.append(["", "", "TOTALS", f"{total_taxable:,.2f}", f"{total_vat:,.2f}", f"{total_invoiced:,.2f}"])

    col_w = [W*0.12, W*0.14, W*0.30, W*0.14, W*0.14, W*0.16]
    tbl = Table(data, colWidths=col_w, repeatRows=1)
    tbl.setStyle(TableStyle([
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE", (0,0), (-1,-1), 7),
        ("BACKGROUND", (0,0), (-1,0), TH),
        ("TEXTCOLOR", (0,0), (-1,0), WH),
        ("ALIGN", (0,0), (-1,-1), "LEFT"),
        ("ALIGN", (3,0), (-1,-1), "RIGHT"),
        ("TEXTCOLOR", (3,1), (3,-2), CR),
        ("FONTNAME", (0,-1), (-1,-1), "Helvetica-Bold"),
        ("BACKGROUND", (0,-1), (-1,-1), colors.HexColor("#f5f8fe")),
        ("TOPPADDING", (0,0), (-1,-1), 4),
        ("BOTTOMPADDING", (0,0), (-1,-1), 4),
        ("GRID", (0,0), (-1,-1), 0.3, colors.HexColor("#d8e4f5")),
    ]))
    els.append(tbl)

    doc.build(els)
    buf.seek(0)
    fn = f"Tax_Report_{from_filter or 'start'}_to_{to_filter or 'end'}.pdf"
    return send_file(buf, mimetype="application/pdf", as_attachment=True, download_name=fn)

# ═══════════════════════════════════════════════════════════
# TABREED TRIPSHEET
# ═══════════════════════════════════════════════════════════

TANKER_GLN_OPTIONS = ["10000 GLN", "5000 GLN", "3000 GLN", "2000 GLN", "1500 GLN", "1000 GLN", "500 GLN"]

@customer_bp.route("/<int:cid>/tripsheet/add", methods=["GET", "POST"])
def customer_tripsheet_add(cid):
    _ensure_tables()
    c = _get_customer_or_404(cid)
    if not c: return redirect(url_for("customer.customer_dashboard"))
    if request.method == "POST":
        entry_date = request.form.get("entry_date", "").strip()
        time_in = request.form.get("time_in", "").strip()
        time_out = request.form.get("time_out", "").strip()
        total_reading = round(float(request.form.get("total_reading", 0) or 0), 2)
        tanker_gln = request.form.get("tanker_gln", "").strip()
        trips = float(request.form.get("trips", 1) or 1)
        tanker_reg = request.form.get("tanker_reg", "").strip().upper()
        notes = request.form.get("notes", "").strip()
        if not entry_date:
            flash("Date is required.", "error")
            return render_template("customer/tripsheet_form.html", c=c, today=date.today().isoformat())
        db = _get_db()
        db.execute("""INSERT INTO tabreed_tripsheets
            (customer_id, entry_date, time_in, time_out, total_reading, tanker_gln, trips, tanker_reg, notes)
            VALUES (?,?,?,?,?,?,?,?,?)""",
            (cid, entry_date, time_in or None, time_out or None, total_reading, tanker_gln or None, trips, tanker_reg or None, notes or None))
        db.commit()
        db.close()
        flash("Tripsheet entry added.", "success")
        if request.args.get("save_and_new"):
            return redirect(url_for("customer.customer_tripsheet_add", cid=cid, date=entry_date))
        return redirect(url_for("customer.customer_profile", cid=cid, tab="tripsheet"))
    default_date = request.args.get("date", date.today().isoformat())
    return render_template("customer/tripsheet_form.html", c=c, today=default_date, tanker_options=TANKER_GLN_OPTIONS)

@customer_bp.route("/<int:cid>/tripsheet/<int:tid>/edit", methods=["GET", "POST"])
def customer_tripsheet_edit(cid, tid):
    _ensure_tables()
    c = _get_customer_or_404(cid)
    if not c: return redirect(url_for("customer.customer_dashboard"))
    db = _get_db()
    row = db.execute("SELECT * FROM tabreed_tripsheets WHERE id=? AND customer_id=?", (tid, cid)).fetchone()
    if not row:
        db.close()
        flash("Tripsheet entry not found.", "error")
        return redirect(url_for("customer.customer_profile", cid=cid, tab="tripsheet"))
    if request.method == "POST":
        entry_date = request.form.get("entry_date", "").strip()
        time_in = request.form.get("time_in", "").strip()
        time_out = request.form.get("time_out", "").strip()
        total_reading = round(float(request.form.get("total_reading", 0) or 0), 2)
        tanker_gln = request.form.get("tanker_gln", "").strip()
        trips = float(request.form.get("trips", 1) or 1)
        tanker_reg = request.form.get("tanker_reg", "").strip().upper()
        notes = request.form.get("notes", "").strip()
        if not entry_date:
            flash("Date is required.", "error")
            return render_template("customer/tripsheet_form.html", c=c, row=row, today=entry_date, tanker_options=TANKER_GLN_OPTIONS, edit=True)
        db.execute("""UPDATE tabreed_tripsheets SET entry_date=?, time_in=?, time_out=?, total_reading=?, tanker_gln=?, trips=?, tanker_reg=?, notes=? WHERE id=? AND customer_id=?""",
            (entry_date, time_in or None, time_out or None, total_reading, tanker_gln or None, trips, tanker_reg or None, notes or None, tid, cid))
        db.commit()
        db.close()
        flash("Tripsheet entry updated.", "success")
        return redirect(url_for("customer.customer_profile", cid=cid, tab="tripsheet"))
    db.close()
    return render_template("customer/tripsheet_form.html", c=c, row=row, today=row["entry_date"], tanker_options=TANKER_GLN_OPTIONS, edit=True)

@customer_bp.route("/<int:cid>/tripsheet/<int:tid>/delete", methods=["POST"])
def customer_tripsheet_delete(cid, tid):
    db = _get_db()
    db.execute("DELETE FROM tabreed_tripsheets WHERE id=? AND customer_id=?", (tid, cid))
    db.commit()
    db.close()
    flash("Tripsheet entry deleted.", "success")
    return redirect(url_for("customer.customer_profile", cid=cid, tab="tripsheet"))

@customer_bp.route("/<int:cid>/tripsheet/report")
def customer_tripsheet_report(cid):
    _ensure_tables()
    c = _get_customer_or_404(cid)
    if not c: return redirect(url_for("customer.customer_dashboard"))
    month = request.args.get("month", date.today().strftime("%Y-%m"))
    db = _get_db()
    rows = db.execute("""
        SELECT * FROM tabreed_tripsheets
        WHERE customer_id=? AND substr(entry_date,1,7)=?
        ORDER BY entry_date, id
    """, (cid, month)).fetchall()
    total_trips = sum(r["trips"] or 0 for r in rows)
    total_reading = sum(r["total_reading"] or 0 for r in rows)
    db.close()
    return render_template("customer/tripsheet_report.html", c=c, rows=rows, month=month,
        total_trips=total_trips, total_reading=total_reading)

@customer_bp.route("/<int:cid>/tripsheet/report/pdf")
def customer_tripsheet_report_pdf(cid):
    _ensure_tables()
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
    from io import BytesIO
    import calendar

    c = _get_customer_or_404(cid)
    if not c: return redirect(url_for("customer.customer_dashboard"))
    month = request.args.get("month", date.today().strftime("%Y-%m"))
    db = _get_db()
    company = db.execute("SELECT * FROM company_profile LIMIT 1").fetchone()
    rows = db.execute("""
        SELECT * FROM tabreed_tripsheets
        WHERE customer_id=? AND substr(entry_date,1,7)=?
        ORDER BY entry_date, id
    """, (cid, month)).fetchall()
    total_trips = sum(r["trips"] or 0 for r in rows)
    total_reading_sum = sum(r["total_reading"] or 0 for r in rows)
    db.close()

    buf = BytesIO()
    LM, RM, TM, BM = 12*mm, 12*mm, 10*mm, 10*mm
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=LM, rightMargin=RM, topMargin=TM, bottomMargin=BM)
    W = A4[0] - LM - RM
    FS = 5

    tc = company["theme_color"] or "#1a3a5c" if company else "#1a3a5c"
    try: TH = colors.HexColor(tc)
    except: TH = colors.HexColor("#1a3a5c")
    WH = colors.white
    C4 = colors.HexColor("#111827")
    C5_g = colors.HexColor("#6b7280")

    _logo_tmp_files = []
    els = []
    cn = company["company_name"] if company else "Current Link"
    trn = company["trn_no"] or "—" if company else "—"

    # ── COMPACT HEADER ──
    logo = None; LW = 0
    if company and company["logo_data"]:
        try:
            import tempfile
            lb = base64.b64decode(company["logo_data"])
            f = tempfile.NamedTemporaryFile(delete=False, suffix=".png")
            f.write(lb); f.close()
            logo = Image(f.name, width=30, height=30)
            LW = 30
            _logo_tmp_files.append(f.name)
        except: pass

    month_name = f"{calendar.month_name[int(month.split('-')[1])]} {month.split('-')[0]}" if '-' in month else month
    co_text = f"<font size=9><b>{cn}</b></font><br/><font size=5.5 color='#6b7280'>{month_name} &middot; {c['customer_name']}</font>"
    co_p = Paragraph(co_text, ParagraphStyle("CO", fontSize=9, fontName="Helvetica-Bold", textColor=TH, leading=11))
    if logo:
        lh = Table([[logo, Spacer(1, 2*mm), co_p]], colWidths=[LW, 2*mm, W*0.65 - LW - 2*mm])
        lh.setStyle(TableStyle([("VALIGN",(0,0),(-1,-1),"MIDDLE"),("LEFTPADDING",(0,0),(-1,-1),0),("RIGHTPADDING",(0,0),(-1,-1),0)]))
    else:
        lh = co_p
    rh = Paragraph(f"<b>TRIPSHEET REPORT</b>", ParagraphStyle("TI", fontSize=11, fontName="Helvetica-Bold", textColor=TH, leading=13, alignment=TA_RIGHT))
    ht = Table([[lh, rh]], colWidths=[W*0.65, W*0.35])
    ht.setStyle(TableStyle([("VALIGN",(0,0),(-1,-1),"TOP"),("LEFTPADDING",(0,0),(-1,-1),0),("RIGHTPADDING",(0,0),(-1,-1),0)]))
    els.append(ht)
    hr = Table([[""]], colWidths=[W], rowHeights=[1])
    hr.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,-1),TH),("LEFTPADDING",(0,0),(-1,-1),0),("RIGHTPADDING",(0,0),(-1,-1),0)]))
    els.append(hr)
    els.append(Spacer(1, 2*mm))

    PPS = lambda name, **kw: ParagraphStyle(name, **kw)
    HDR = ["#", "Date", "Time In", "Time Out", "Total Reading", "Tanker GLN", "Trips", "Tanker Reg"]
    hdr_p = [Paragraph(f"<b>{h}</b>", PPS("h", fontSize=6, fontName="Helvetica-Bold", textColor=WH, leading=7, alignment=TA_CENTER)) for h in HDR]
    data = [hdr_p]
    for idx, r in enumerate(rows, 1):
        data.append([
            Paragraph(str(idx), PPS("c", fontSize=FS, leading=FS + 0.5, alignment=TA_CENTER)),
            Paragraph(r["entry_date"] or "—", PPS("c", fontSize=FS, leading=FS + 0.5, alignment=TA_CENTER)),
            Paragraph(r["time_in"] or "—", PPS("c", fontSize=FS, leading=FS + 0.5, alignment=TA_CENTER)),
            Paragraph(r["time_out"] or "—", PPS("c", fontSize=FS, leading=FS + 0.5, alignment=TA_CENTER)),
            Paragraph(f"{r['total_reading'] or 0:,.2f}", PPS("c", fontSize=FS, fontName="Helvetica-Bold", leading=FS + 0.5, alignment=TA_RIGHT)),
            Paragraph(r["tanker_gln"] or "—", PPS("c", fontSize=FS, leading=FS + 0.5, alignment=TA_CENTER)),
            Paragraph(f"{r['trips'] or 0:,.0f}", PPS("c", fontSize=FS, leading=FS + 0.5, alignment=TA_CENTER)),
            Paragraph(r["tanker_reg"] or "—", PPS("c", fontSize=FS, leading=FS + 0.5, alignment=TA_CENTER)),
        ])
    data.append([
        Paragraph("<b>Total</b>", PPS("t", fontSize=6, fontName="Helvetica-Bold", leading=7, alignment=TA_CENTER)),
        Paragraph("", PPS("c", fontSize=FS, leading=FS + 0.5)),
        Paragraph("", PPS("c", fontSize=FS, leading=FS + 0.5)),
        Paragraph("", PPS("c", fontSize=FS, leading=FS + 0.5)),
        Paragraph(f"<b>{total_reading_sum:,.2f}</b>", PPS("t", fontSize=6, fontName="Helvetica-Bold", leading=7, alignment=TA_RIGHT)),
        Paragraph("", PPS("c", fontSize=FS, leading=FS + 0.5)),
        Paragraph(f"<b>{total_trips:,.0f}</b>", PPS("t", fontSize=6, fontName="Helvetica-Bold", leading=7, alignment=TA_CENTER)),
        Paragraph("", PPS("c", fontSize=FS, leading=FS + 0.5)),
    ])

    col_w = [7*mm, 24*mm, 20*mm, 20*mm, 28*mm, 32*mm, 12*mm, W - 7 - 24 - 20 - 20 - 28 - 32 - 12]
    tbl = Table(data, colWidths=col_w, repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), TH), ("TEXTCOLOR", (0,0), (-1,0), WH),
        ("ALIGN", (0,0), (-1,-1), "CENTER"),
        ("ALIGN", (4,1), (4,-1), "RIGHT"),
        ("FONTNAME", (0,-1), (-1,-1), "Helvetica-Bold"),
        ("BACKGROUND", (0,-1), (-1,-1), colors.HexColor("#f5f8fe")),
        ("TOPPADDING", (0,0), (-1,-1), 0.5),
        ("BOTTOMPADDING", (0,0), (-1,-1), 0.5),
        ("GRID", (0,0), (-1,-1), 0.3, colors.HexColor("#b0bed0")),
        ("LINEBELOW", (0,0), (-1,0), 0.6, TH),
    ]))
    els.append(tbl)

    doc.build(els)
    for f in _logo_tmp_files:
        try: os.remove(f)
        except: pass
    buf.seek(0)
    fn = f"Tabreed_Tripsheet_{month}_{c['customer_name']}.pdf"
    return send_file(buf, mimetype="application/pdf", as_attachment=True, download_name=fn)
