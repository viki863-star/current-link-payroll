import os
from datetime import date, datetime
from pathlib import Path
from io import BytesIO
import base64

from flask import (
    current_app, flash, redirect, render_template, request,
    send_file, session, url_for, jsonify
)
from werkzeug.security import generate_password_hash
from werkzeug.utils import secure_filename

from ..database import open_db
from ..routes import (
    _login_required, _audit_log, _touch_admin_workspace, _current_month_value,
    ValidationError, _parse_decimal, _normalize_month, _next_month_value,
    _advance_summary, _outstanding_advance, _timesheet_total_for_month,
    _calculate_salary_preview, _default_salary_form, _salary_form_from_row,
    _salary_preview_from_row, format_month_label, _driver_month_calendar,
    _timesheet_month_summary, _driver_kata_month_data, _regenerate_kata_for_driver,
    SALARY_MODE_OPTIONS, PAYMENT_SOURCES
)
from . import hr_bp
from .forms import (
    employee_form_data, validate_employee_form,
    EMPLOYEE_TYPES, DEPARTMENTS, DESIGNATIONS, STATUS_OPTIONS,
    GENDER_OPTIONS, SHIFT_OPTIONS, CONTRACT_TYPE_OPTIONS
)
from .services import (
    sync_drivers_to_employees, save_employee_photo,
    employee_search_filter, next_employee_id,
    employee_departments, employee_types,
    sync_field_staff_to_employees,
    EMPLOYEE_SCHEMA, EMPLOYEE_SCHEMA_POSTGRES
)


def ensure_employees_table():
    db = open_db()
    backend = current_app.config.get("DATABASE_BACKEND", "sqlite")

    try:
        schema = EMPLOYEE_SCHEMA_POSTGRES if backend == "postgres" else EMPLOYEE_SCHEMA
        db.executescript(schema)
        db.commit()
    except Exception as e:
        import traceback
        current_app.logger.error("ensure_employees_table schema error: %s\n%s", e, traceback.format_exc())

    # Migration: add missing columns for tables created before schema was expanded
    employee_cols = [
        ("employee_id", "TEXT"),
        ("email", "TEXT"),
        ("gender", "TEXT"),
        ("shift", "TEXT DEFAULT 'Morning'"),
        ("contract_type", "TEXT DEFAULT 'Permanent'"),
        ("nationality", "TEXT"),
        ("iqama_no", "TEXT"),
        ("passport_no", "TEXT"),
        ("bank_name", "TEXT"),
        ("bank_account", "TEXT"),
        ("iban", "TEXT"),
        ("emergency_contact", "TEXT"),
        ("emergency_name", "TEXT"),
        ("address", "TEXT"),
        ("photo_name", "TEXT"),
        ("photo_data", "TEXT"),
        ("photo_content_type", "TEXT"),
        ("termination_date", "TEXT"),
        ("remarks", "TEXT"),
        ("updated_at", "TEXT DEFAULT CURRENT_TIMESTAMP"),
    ]
    for col_name, col_type in employee_cols:
        try:
            if backend == "postgres":
                db.execute(f"ALTER TABLE employees ADD COLUMN IF NOT EXISTS {col_name} {col_type}")
            else:
                db.execute(f"ALTER TABLE employees ADD COLUMN {col_name} {col_type}")
            db.commit()
        except Exception:
            db.rollback()

    try:
        sync_drivers_to_employees(db)
    except Exception:
        db.rollback()

    try:
        sync_field_staff_to_employees(db)
    except Exception:
        db.rollback()


def _fetch_employee(db, employee_id):
    return db.execute(
        "SELECT * FROM employees WHERE UPPER(employee_id) = ?",
        (employee_id.strip().upper(),),
    ).fetchone()


def _sync_employee_to_field_staff(db, staff_id, full_name, phone, status):
    from werkzeug.security import generate_password_hash
    username = staff_id.lower()
    pw_hash = generate_password_hash("changeme123")
    is_active = 1 if status == "Active" else 0
    try:
        existing = db.execute("SELECT staff_id FROM field_staff WHERE staff_id = ?", (staff_id,)).fetchone()
        if existing:
            db.execute("UPDATE field_staff SET full_name=?, phone=?, is_active=? WHERE staff_id=?", (full_name, phone or "", is_active, staff_id))
        else:
            db.execute("INSERT INTO field_staff (staff_id, full_name, phone, username, password_hash, is_active) VALUES (?,?,?,?,?,?)", (staff_id, full_name, phone or "", username, pw_hash, is_active))
    except Exception:
        pass


def _employee_photo_url(app, employee):
    if not employee:
        return None
    if employee.get("photo_data") and employee.get("photo_content_type"):
        return f"data:{employee['photo_content_type']};base64,{employee['photo_data']}"
    if employee.get("photo_url"):
        return employee["photo_url"]
    if employee.get("photo_name"):
        try:
            import base64
            from pathlib import Path
            photo_dir = Path(app.config.get("DRIVER_FILES_DIR", "")) / "employee_photos"
            photo_path = photo_dir / employee["photo_name"]
            if photo_path.exists():
                ct = "image/jpeg"
                ext = photo_path.suffix.lower()
                if ext == ".png": ct = "image/png"
                elif ext == ".gif": ct = "image/gif"
                elif ext == ".webp": ct = "image/webp"
                with open(photo_path, "rb") as f:
                    data = base64.b64encode(f.read()).decode("utf-8")
                return f"data:{ct};base64,{data}"
        except Exception:
            pass
    return None


# ── HR Dashboard ────────────────────────────────────────────────

@hr_bp.route("/hr")
@_login_required("admin")
def hr_dashboard():
    try:
        _touch_admin_workspace("hr")
        ensure_employees_table()
        db = open_db()

        employees = db.execute(
            "SELECT employee_id, full_name, employee_type, department, status, join_date FROM employees ORDER BY full_name"
        ).fetchall()

        total = len(employees)
        active = sum(1 for e in employees if (e["status"] or "").lower() == "active")
        inactive = total - active

        stored_this_month = db.execute(
            "SELECT COUNT(*) AS c FROM salary_store WHERE salary_month = ?",
            (_current_month_value(),),
        ).fetchone()["c"] or 0

        advances_pending = db.execute(
            "SELECT COUNT(*) AS c FROM driver_transactions WHERE txn_type IN ('advance','loan')"
        ).fetchone()["c"] or 0

        # ── Department breakdown ──
        dept_rows = db.execute(
            "SELECT department, COUNT(*) AS c FROM employees GROUP BY department ORDER BY c DESC"
        ).fetchall()
        departments = {r["department"]: r["c"] for r in dept_rows}

        # ── Type breakdown ──
        type_rows = db.execute(
            "SELECT employee_type, COUNT(*) AS c FROM employees GROUP BY employee_type ORDER BY c DESC"
        ).fetchall()
        employee_types_dict = {r["employee_type"]: r["c"] for r in type_rows}

        # ── Monthly salary store trend (last 6 months) ──
        from datetime import date
        trend_months = []
        trend_counts = []
        today = date.today()
        for i in range(5, -1, -1):
            m = today.month - i
            y = today.year
            if m < 1:
                m += 12
                y -= 1
            ym = f"{y:04d}-{m:02d}"
            cnt = db.execute(
                "SELECT COUNT(*) AS c FROM salary_store WHERE salary_month = ?", (ym,)
            ).fetchone()["c"] or 0
            trend_months.append(ym)
            trend_counts.append(cnt)

        # ── Active drivers payroll amount this month ──
        payroll_row = db.execute(
            "SELECT COALESCE(SUM(ss.net_salary), 0) AS total FROM salary_store ss JOIN employees e ON e.employee_id = ss.driver_id WHERE ss.salary_month = ? AND LOWER(e.status) = 'active'",
            (_current_month_value(),),
        ).fetchone()
        payroll_amount = payroll_row["total"] if payroll_row else 0

        # ── Recent employees ──
        recent = db.execute(
            "SELECT employee_id, full_name, employee_type, department, join_date, status FROM employees ORDER BY id DESC LIMIT 5"
        ).fetchall()

        # ── Count by status ──
        on_leave = sum(1 for e in employees if (e["status"] or "").lower() == "on leave")
        terminated = sum(1 for e in employees if (e["status"] or "").lower() == "terminated")

        return render_template(
            "hr/dashboard.html",
            total=total,
            active_count=active,
            inactive_count=inactive,
            on_leave=on_leave,
            terminated=terminated,
            stored_this_month=stored_this_month,
            payroll_amount=payroll_amount,
            advances_pending=advances_pending,
            departments=departments,
            employee_types=employee_types_dict,
            trend_months=trend_months,
            trend_counts=trend_counts,
            recent_employees=recent,
        )
    except Exception as e:
        import traceback
        tb = traceback.format_exc()
        current_app.logger.error("HR dashboard error: %s | type=%s\n%s", e, type(e).__name__, tb)
        flash(f"HR Dashboard error: {type(e).__name__}: {e}\n\nTraceback:\n{tb[:2000]}", "error")
        return redirect(url_for("dashboard"))


# ── Employee List ────────────────────────────────────────────────

@hr_bp.route("/hr/employees")
@_login_required("admin")
def employee_list():
    _touch_admin_workspace("hr")
    ensure_employees_table()
    db = open_db()

    query = request.args.get("q", "").strip()
    status_filter = request.args.get("status", "").strip()
    department_filter = request.args.get("department", "").strip()
    employee_type_filter = request.args.get("type", "").strip()

    status_counts = db.execute(
        "SELECT LOWER(status) AS st, COUNT(*) AS c FROM employees GROUP BY LOWER(status)"
    ).fetchall()
    total_active = sum(r["c"] for r in status_counts if r["st"] == "active")
    total_inactive = sum(r["c"] for r in status_counts if r["st"] in ("inactive", "on leave"))
    total_terminated = sum(r["c"] for r in status_counts if r["st"] == "terminated")

    if not status_filter:
        status_filter = "Active"

    where_sql, params = employee_search_filter(query, status_filter, department_filter, employee_type_filter)

    employees = db.execute(
        f"""
        SELECT e.employee_id, e.full_name, e.phone_number, e.email, e.employee_type,
               e.department, e.designation, e.join_date, e.basic_salary, e.status, e.photo_name, e.termination_date,
               COALESCE(
                   (SELECT v.plate_no FROM vehicle_assignments va
                    JOIN vehicles v ON v.plate_no = va.vehicle_id
                    WHERE va.driver_id = e.employee_id
                    ORDER BY va.id DESC LIMIT 1),
                   (SELECT d.vehicle_no FROM drivers d WHERE d.driver_id = e.employee_id LIMIT 1)
               ) AS plate_no
        FROM employees e
        {where_sql}
        ORDER BY e.full_name ASC
        """,
        params,
    ).fetchall()

    all_departments = employee_departments(db)
    all_types = employee_types(db)

    return render_template(
        "hr/employee_list.html",
        employees=employees,
        query=query,
        status_filter=status_filter,
        department_filter=department_filter,
        employee_type_filter=employee_type_filter,
        departments=all_departments,
        employee_types=all_types,
        status_options=STATUS_OPTIONS,
        total_active=total_active,
        total_inactive=total_inactive,
        total_terminated=total_terminated,
    )


# ── Add Employee ─────────────────────────────────────────────────

@hr_bp.route("/hr/employees/new", methods=["GET", "POST"])
@_login_required("admin")
def employee_new():
    _touch_admin_workspace("hr")
    ensure_employees_table()
    db = open_db()

    values = employee_form_data()
    if not values["employee_id"]:
        try:
            values["employee_id"] = next_employee_id(db)
        except Exception as exc:
            import traceback
            current_app.logger.error("next_employee_id failed: %s\n%s", exc, traceback.format_exc())
            values["employee_id"] = ""

    # Get vehicles if table exists, otherwise empty list
    try:
        vehicles = db.execute("SELECT plate_no, vehicle_type, model FROM vehicles WHERE status = 'Active' ORDER BY plate_no").fetchall()
    except Exception:
        vehicles = []

    if request.method == "POST":
        errors = validate_employee_form(values)
        if errors:
            for err in errors:
                flash(err, "error")
            return render_template(
                "hr/employee_form.html",
                values=values,
                page_title="Add Employee",
                submit_label="Save Employee",
                edit_mode=False,
                employee_types=EMPLOYEE_TYPES,
                departments=DEPARTMENTS,
                designations=DESIGNATIONS,
                status_options=STATUS_OPTIONS,
                gender_options=GENDER_OPTIONS,
                shift_options=SHIFT_OPTIONS,
                contract_options=CONTRACT_TYPE_OPTIONS,
                vehicles=vehicles,
                assigned_vehicle=values.get("vehicle_id", ""),
            )

        try:
            salary = float(values["basic_salary"])
            ot_rate = float(values.get("ot_rate", 0) or 0)

            uploaded_photo = save_employee_photo(
                current_app._get_current_object(), values["employee_id"],
                values["full_name"], request.files.get("photo_file")
            )

            db.execute(
                """
                INSERT INTO employees (
                    employee_id, full_name, phone_number, email,
                    employee_type, department, designation, gender,
                    shift, contract_type, join_date, basic_salary, ot_rate,
                    nationality, iqama_no, passport_no,
                    bank_name, bank_account, iban,
                    emergency_contact, emergency_name, address,
                    photo_name, photo_data, photo_content_type,
                    status, termination_date, remarks
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    values["employee_id"], values["full_name"], values["phone_number"], values["email"] or None,
                    values["employee_type"], values["department"], values["designation"], values["gender"] or None,
                    values["shift"] or "Morning", values["contract_type"] or "Permanent",
                    values["join_date"], salary, ot_rate,
                    values["nationality"] or None, values["iqama_no"] or None, values["passport_no"] or None,
                    values["bank_name"] or None, values["bank_account"] or None, values["iban"] or None,
                    values["emergency_contact"] or None, values["emergency_name"] or None, values["address"] or None,
                    uploaded_photo["photo_name"] if uploaded_photo else None,
                    uploaded_photo["photo_data"] if uploaded_photo else None,
                    uploaded_photo["photo_content_type"] if uploaded_photo else None,
                    values["status"], values["termination_date"] or None, values["remarks"] or None,
                ),
            )

            db.execute(
                """
                INSERT INTO drivers (
                    driver_id, full_name, phone_number, vehicle_no, shift, vehicle_type,
                    basic_salary, ot_rate, duty_start, photo_name, photo_data, photo_content_type,
                    status, termination_date
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(driver_id) DO UPDATE SET
                    full_name=excluded.full_name,
                    phone_number=excluded.phone_number,
                    shift=excluded.shift,
                    basic_salary=excluded.basic_salary,
                    ot_rate=excluded.ot_rate,
                    photo_name=excluded.photo_name,
                    photo_data=excluded.photo_data,
                    photo_content_type=excluded.photo_content_type,
                    status=excluded.status,
                    termination_date=excluded.termination_date
                """,
                (
                    values["employee_id"], values["full_name"], values["phone_number"] or None,
                    values.get("vehicle_id", "") or "",
                    values["shift"] or "Morning",
                    values.get("vehicle_type", "Car") or "Car",
                    salary, ot_rate,
                    values["join_date"] or None,
                    uploaded_photo["photo_name"] if uploaded_photo else None,
                    uploaded_photo["photo_data"] if uploaded_photo else None,
                    uploaded_photo["photo_content_type"] if uploaded_photo else None,
                    values["status"], values["termination_date"] or None,
                ),
            )

            if values["vehicle_id"]:
                db.execute(
                    "UPDATE vehicle_assignments SET is_current = 0, assigned_until = ? WHERE driver_id = ? AND is_current = 1",
                    (date.today().isoformat(), values["employee_id"]),
                )
                db.execute(
                    "INSERT INTO vehicle_assignments (vehicle_id, driver_id, assigned_from, is_current) VALUES (?, ?, ?, 1)",
                    (values["vehicle_id"], values["employee_id"], date.today().isoformat()),
                )
                db.execute("UPDATE drivers SET vehicle_no = ? WHERE driver_id = ?", (values["vehicle_id"], values["employee_id"]))

            if values["employee_type"] == "Field Staff":
                _sync_employee_to_field_staff(db, values["employee_id"], values["full_name"], values["phone_number"], values["status"])

            _audit_log(
                db, "employee_created",
                entity_type="employee",
                entity_id=values["employee_id"],
                details=f"{values['full_name']} / {values['employee_type']} / {values['department']}",
            )
            db.commit()
            flash(f"Employee {values['employee_id']} - {values['full_name']} created successfully.", "success")
            return redirect(url_for("hr.employee_detail", employee_id=values["employee_id"]))
        except Exception as e:
            db.rollback()
            import traceback
            current_app.logger.error(f"Employee creation failed: {e}\n{traceback.format_exc()}")
            flash(f"Error creating employee: {str(e)}", "error")
            return render_template(
                "hr/employee_form.html",
                values=values,
                page_title="Add Employee",
                submit_label="Save Employee",
                edit_mode=False,
                employee_types=EMPLOYEE_TYPES,
                departments=DEPARTMENTS,
                designations=DESIGNATIONS,
                status_options=STATUS_OPTIONS,
                gender_options=GENDER_OPTIONS,
                shift_options=SHIFT_OPTIONS,
                contract_options=CONTRACT_TYPE_OPTIONS,
                vehicles=vehicles,
                assigned_vehicle=values.get("vehicle_id", ""),
            )


    # For new employees, no assigned vehicle query needed
    assigned_vehicle = None
    return render_template(
        "hr/employee_form.html",
        values=values,
        page_title="Add Employee",
        submit_label="Save Employee",
        edit_mode=False,
        employee_types=EMPLOYEE_TYPES,
        departments=DEPARTMENTS,
        designations=DESIGNATIONS,
        status_options=STATUS_OPTIONS,
        gender_options=GENDER_OPTIONS,
        shift_options=SHIFT_OPTIONS,
        contract_options=CONTRACT_TYPE_OPTIONS,
        vehicles=vehicles,
        assigned_vehicle="",
    )


# ── Employee Detail (Profile + 3 Tabs) ───────────────────────────

@hr_bp.route("/hr/employees/<employee_id>")
@_login_required("admin")
def employee_detail(employee_id):
    _touch_admin_workspace("hr")
    ensure_employees_table()
    db = open_db()

    employee = _fetch_employee(db, employee_id)
    if employee is None:
        flash("Employee not found.", "error")
        return redirect(url_for("hr.employee_list"))

    return redirect(url_for("hr.employee_transactions", employee_id=employee_id))


# ── Transactions Tab ─────────────────────────────────────────────

@hr_bp.route("/hr/employees/<employee_id>/transactions", methods=["GET", "POST"])
@_login_required("admin")
def employee_transactions(employee_id):
    _touch_admin_workspace("hr")
    ensure_employees_table()
    db = open_db()

    employee = _fetch_employee(db, employee_id)
    if employee is None:
        flash("Employee not found.", "error")
        return redirect(url_for("hr.employee_list"))

    eid = employee["employee_id"]
    today = date.today().isoformat()
    form_values = {
        "entry_date": today,
        "amount": "",
        "source": "Cash",
        "given_by": "",
        "details": "",
    }

    if request.method == "POST":
        edit_id = request.form.get("edit_id", "").strip()
        form_values = {
            "entry_date": request.form.get("entry_date", today).strip() or today,
            "amount": request.form.get("amount", "0").strip(),
            "source": request.form.get("source", "Cash").strip(),
            "given_by": request.form.get("given_by", "").strip(),
            "details": request.form.get("details", "").strip(),
        }
        try:
            amount = _parse_decimal(form_values["amount"], "Amount", minimum=0.01)
            if not form_values["given_by"]:
                raise ValidationError("Given by (person name) is required.")
            if not form_values["details"]:
                raise ValidationError("Details / reason is required.")
            txn_type = request.form.get("txn_type", "Advance").strip()
            salary_month = request.form.get("salary_month", "").strip() or _current_month_value()

            if edit_id and edit_id.isdigit():
                db.execute(
                    """UPDATE driver_transactions SET entry_date=?, salary_month=?, txn_type=?, source=?, given_by=?, amount=?, details=? WHERE id=? AND driver_id=?""",
                    (form_values["entry_date"], salary_month, txn_type, form_values["source"],
                     form_values["given_by"], amount, form_values["details"], int(edit_id), eid),
                )
                _audit_log(db, "employee_transaction_updated", entity_type="employee_transaction", entity_id=eid, details=f"AED {amount:.2f} / txn#{edit_id}")
                db.commit()
                flash(f"Transaction #{edit_id} updated.", "success")
                return redirect(url_for("hr.employee_transactions", employee_id=eid))
            else:
                db.execute(
                    """INSERT INTO driver_transactions (driver_id, entry_date, salary_month, txn_type, source, given_by, amount, details) VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
                    (eid, form_values["entry_date"], salary_month, txn_type, form_values["source"],
                     form_values["given_by"], amount, form_values["details"]),
                )
                _audit_log(db, "employee_transaction_created", entity_type="employee_transaction", entity_id=eid, details=f"AED {amount:.2f} / {form_values['source']} / {form_values['details']}")
                db.commit()
                flash(f"Transaction of AED {amount:.2f} recorded for {employee['full_name']}.", "success")
                return redirect(url_for("hr.employee_transactions", employee_id=eid))
        except ValidationError as exc:
            flash(str(exc), "error")

    transactions = db.execute(
        """SELECT id, entry_date, salary_month, txn_type, source, given_by, amount, details, created_at FROM driver_transactions WHERE driver_id = ? ORDER BY entry_date DESC, id DESC LIMIT 50""",
        (eid,),
    ).fetchall()

    total_advance = db.execute(
        "SELECT COALESCE(SUM(amount), 0) FROM driver_transactions WHERE driver_id = ?",
        (eid,),
    ).fetchone()[0]

    edit_txn = None
    edit_id = request.args.get("edit", "").strip()
    if edit_id and edit_id.isdigit():
        edit_txn = db.execute("SELECT * FROM driver_transactions WHERE id = ? AND driver_id = ?", (int(edit_id), eid)).fetchone()

    photo_url = _employee_photo_url(current_app._get_current_object(), employee)

    # Fetch current vehicle directly (same logic as edit page)
    current_vehicle = None
    if employee.get("employee_type") == "Driver":
        veh_row = db.execute(
            "SELECT va.vehicle_id, va.assigned_from, v.plate_no, v.vehicle_type, v.model "
            "FROM vehicle_assignments va JOIN vehicles v ON v.plate_no = va.vehicle_id "
            "WHERE va.driver_id = ? AND va.is_current = 1 LIMIT 1",
            (eid,),
        ).fetchone()
        if not veh_row:
            veh_row = db.execute(
                "SELECT va.vehicle_id, va.assigned_from, v.plate_no, v.vehicle_type, v.model "
                "FROM vehicle_assignments va JOIN vehicles v ON v.plate_no = va.vehicle_id "
                "WHERE va.driver_id = ? ORDER BY va.id DESC LIMIT 1",
                (eid,),
            ).fetchone()
        if not veh_row:
            legacy = db.execute("SELECT vehicle_no FROM drivers WHERE driver_id = ? LIMIT 1", (eid,)).fetchone()
            if legacy and legacy["vehicle_no"]:
                veh_row = db.execute("SELECT plate_no AS vehicle_id, NULL AS assigned_from, plate_no, vehicle_type, model FROM vehicles WHERE plate_no = ? LIMIT 1", (legacy["vehicle_no"],)).fetchone()
                if not veh_row:
                    current_vehicle = {"vehicle_id": legacy["vehicle_no"], "assigned_from": "", "plate_no": legacy["vehicle_no"], "vehicle_type": "", "model": ""}
        if veh_row:
            current_vehicle = dict(veh_row)

    return render_template(
        "hr/employee_detail.html",
        employee=employee,
        photo_url=photo_url,
        active_tab="transactions",
        txn_form=form_values,
        transactions=transactions,
        total_advance=total_advance,
        edit_txn=edit_txn,
        current_vehicle=current_vehicle,
    )


@hr_bp.route("/hr/employees/<employee_id>/transactions/<int:txn_id>/delete", methods=["POST"])
@_login_required("admin")
def employee_transaction_delete(employee_id, txn_id):
    _touch_admin_workspace("hr")
    ensure_employees_table()
    db = open_db()
    try:
        db.execute("DELETE FROM driver_transactions WHERE id=? AND driver_id=?", (txn_id, employee_id))
        _audit_log(db, "employee_transaction_deleted", entity_type="employee_transaction", entity_id=employee_id, details=f"txn#{txn_id}")
        db.commit()
        flash("Transaction deleted.", "info")
    except Exception as e:
        db.rollback()
        flash(f"Error deleting transaction: {e}", "error")
    return redirect(url_for("hr.employee_transactions", employee_id=employee_id))


# ── Store Salary Tab ─────────────────────────────────────────────

@hr_bp.route("/hr/employees/<employee_id>/salary-store", methods=["GET", "POST"])
@_login_required("admin")
def employee_salary_store(employee_id):
    _touch_admin_workspace("hr")
    ensure_employees_table()
    db = open_db()

    employee = _fetch_employee(db, employee_id)
    if employee is None:
        flash("Employee not found.", "error")
        return redirect(url_for("hr.employee_list"))

    eid = employee["employee_id"]

    # Build a driver-like dict for salary helpers
    driver_like = {
        "driver_id": eid,
        "basic_salary": employee["basic_salary"] or 0,
        "ot_rate": employee["ot_rate"] or 0,
        "duty_start": employee["join_date"],
    }

    selected_month = request.args.get("month", "").strip() or _current_month_value()
    existing_row = db.execute(
        "SELECT * FROM salary_store WHERE driver_id = ? AND salary_month = ?",
        (eid, selected_month),
    ).fetchone()

    form = _default_salary_form(selected_month, employee.get("join_date"))
    preview = _calculate_salary_preview(driver_like, form)
    if existing_row is not None:
        form = _salary_form_from_row(existing_row)
        preview = _salary_preview_from_row(existing_row)

    if request.method == "POST":
        form = {
            "entry_date": request.form.get("entry_date", date.today().isoformat()).strip() or date.today().isoformat(),
            "salary_month": _normalize_month(request.form.get("salary_month", selected_month).strip() or selected_month),
            "ot_month": "",
            "salary_mode": (request.form.get("salary_mode", "full").strip() or "full").lower(),
            "basic_salary": request.form.get("basic_salary", "").strip(),
            "prorata_start_date": request.form.get("prorata_start_date", "").strip(),
            "prorata_end_date": request.form.get("prorata_end_date", "").strip(),
            "ot_hours": request.form.get("ot_hours", "0").strip() or "0",
            "ot_type": (request.form.get("ot_type", "hours").strip() or "hours").lower(),
            "ot_trips": request.form.get("ot_trips", "0").strip() or "0",
            "personal_vehicle": request.form.get("personal_vehicle", "0").strip() or "0",
            "personal_vehicle_note": request.form.get("personal_vehicle_note", "").strip(),
            "remarks": request.form.get("remarks", "").strip(),
        }
        form["ot_month"] = _previous_month_value(form["salary_month"])
        if form["basic_salary"]:
            driver_like["basic_salary"] = float(form["basic_salary"])
        try:
            preview = _calculate_salary_preview(driver_like, form)
        except ValidationError as exc:
            flash(str(exc), "error")
        else:
            if employee.get("termination_date") and (employee.get("status") or "").lower() == "terminated":
                term_date = employee["termination_date"]
                if term_date < form["salary_month"] + "-01":
                    flash(f"Cannot store salary: employee was terminated on {term_date}. Select a month on or before {term_date[:7]}.", "error")
                    return redirect(url_for("hr.employee_salary_store", employee_id=eid, month=form["salary_month"]))
            action = request.form.get("action", "calculate")
            if action == "save":
                existing_month_row = db.execute(
                    "SELECT id FROM salary_store WHERE driver_id = ? AND salary_month = ?",
                    (eid, form["salary_month"]),
                ).fetchone()

                db.execute(
                    """
                    INSERT INTO salary_store (
                        driver_id, entry_date, salary_month, ot_month, salary_mode, prorata_start_date,
                        salary_days, daily_rate, monthly_basic_salary, basic_salary, ot_hours, ot_rate,
                        ot_amount, ot_type, ot_trips, personal_vehicle, personal_vehicle_note, net_salary, remarks
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ON CONFLICT(driver_id, salary_month) DO UPDATE SET
                        entry_date = excluded.entry_date, ot_month = excluded.ot_month,
                        salary_mode = excluded.salary_mode, prorata_start_date = excluded.prorata_start_date,
                        salary_days = excluded.salary_days, daily_rate = excluded.daily_rate,
                        monthly_basic_salary = excluded.monthly_basic_salary,
                        basic_salary = excluded.basic_salary, ot_hours = excluded.ot_hours,
                        ot_rate = excluded.ot_rate, ot_amount = excluded.ot_amount,
                        ot_type = excluded.ot_type, ot_trips = excluded.ot_trips,
                        personal_vehicle = excluded.personal_vehicle,
                        personal_vehicle_note = excluded.personal_vehicle_note,
                        net_salary = excluded.net_salary, remarks = excluded.remarks
                    """,
                    (
                        eid, form["entry_date"], form["salary_month"],
                        preview["ot_month"], preview["salary_mode"],
                        preview["prorata_start_date"] or None,
                        preview["salary_days"], preview["daily_rate"],
                        preview["monthly_basic_salary"], preview["basic_salary"],
                        preview["ot_hours"], preview["ot_rate"],
                        preview["ot_amount"], preview["ot_type"],
                        preview["ot_trips"], preview["personal_vehicle"],
                        preview["personal_vehicle_note"] or None,
                        preview["net_salary"], form["remarks"],
                    ),
                )
                _audit_log(
                    db, "employee_salary_store_saved",
                    entity_type="salary_store",
                    entity_id=f"{eid}:{form['salary_month']}",
                    details=f"{preview['salary_mode_label']} / net AED {preview['net_salary']:.2f}",
                )
                db.commit()
                if existing_month_row:
                    flash("Salary updated for this month.", "success")
                else:
                    flash("Salary stored successfully.", "success")
                return redirect(url_for("hr.employee_salary_store", employee_id=eid, month=form["salary_month"]))

    salary_rows = db.execute(
        """
        SELECT id, entry_date, salary_month, ot_month, salary_mode, prorata_start_date,
               salary_days, daily_rate, monthly_basic_salary, basic_salary, ot_hours,
               ot_amount, personal_vehicle, personal_vehicle_note, net_salary, remarks
        FROM salary_store
        WHERE driver_id = ?
        ORDER BY salary_month DESC
        LIMIT 12
        """,
        (eid,),
    ).fetchall()

    timesheet_hours = _timesheet_total_for_month(db, eid, form["salary_month"])
    photo_url = _employee_photo_url(current_app._get_current_object(), employee)

    return render_template(
        "hr/employee_detail.html",
        employee=employee,
        photo_url=photo_url,
        active_tab="salary_store",
        salary_form=form,
        salary_preview=preview,
        salary_rows=salary_rows,
        selected_month_label=format_month_label(form["salary_month"]),
        timesheet_hours=timesheet_hours,
        existing_month=existing_row is not None,
        salary_mode_options=SALARY_MODE_OPTIONS,
    )


@hr_bp.route("/hr/employees/<employee_id>/salary-store/<int:store_id>/delete", methods=["GET", "POST"])
@_login_required("admin")
def employee_salary_store_delete(employee_id, store_id):
    _touch_admin_workspace("hr")
    ensure_employees_table()
    try:
        db = open_db()
        employee = _fetch_employee(db, employee_id)
        if employee is None:
            flash("Employee not found.", "error")
            return redirect(url_for("hr.employee_list"))

        eid = employee["employee_id"]
        row = db.execute(
            "SELECT * FROM salary_store WHERE id = ? AND driver_id = ?",
            (store_id, eid),
        ).fetchone()
        if row is None:
            flash("Salary store not found.", "error")
        else:
            db.execute("DELETE FROM salary_payments WHERE salary_store_id = ? AND driver_id = ?", (store_id, eid))
            db.execute("DELETE FROM salary_slips WHERE salary_store_id = ? AND driver_id = ?", (store_id, eid))
            db.execute("DELETE FROM salary_store WHERE id = ? AND driver_id = ?", (store_id, eid))
            _audit_log(db, "employee_salary_store_deleted", entity_type="salary_store", entity_id=f"{eid}:{row['salary_month']}")
            db.commit()
            from flask import current_app
            driver_row = db.execute("SELECT driver_id, full_name FROM drivers WHERE driver_id = ?", (eid,)).fetchone()
            _regenerate_kata_for_driver(current_app._get_current_object(), db, dict(driver_row) if driver_row else {"driver_id": eid, "full_name": eid})
            flash(f"Salary store for {row['salary_month']} deleted.", "success")
    except Exception as ex:
        flash(f"Error deleting salary store: {ex}", "error")
    return redirect(url_for("hr.employee_salary_store", employee_id=eid))


def _previous_month_value(month_value: str) -> str:
    y, m = int(month_value[:4]), int(month_value[5:7])
    m -= 1
    if m == 0:
        y -= 1
        m = 12
    return f"{y:04d}-{m:02d}"


# ── Run Salary / Salary Slip Tab ─────────────────────────────────

@hr_bp.route("/hr/employees/<employee_id>/salary-slip", methods=["GET", "POST"])
@_login_required("admin")
def employee_salary_slip(employee_id):
    _touch_admin_workspace("hr")
    ensure_employees_table()
    db = open_db()

    employee = _fetch_employee(db, employee_id)
    if employee is None:
        flash("Employee not found.", "error")
        return redirect(url_for("hr.employee_list"))

    eid = employee["employee_id"]

    salary_rows = db.execute(
        "SELECT * FROM salary_store WHERE driver_id = ? ORDER BY salary_month DESC",
        (eid,),
    ).fetchall()

    slip_store_ids = {
        r["salary_store_id"]
        for r in db.execute(
            "SELECT DISTINCT salary_store_id FROM salary_slips WHERE driver_id = ?", (eid,)
        ).fetchall()
    }

    selected_salary_id = request.args.get("salary_store_id", "").strip()
    if not selected_salary_id and salary_rows:
        selected_salary_id = str(salary_rows[0]["id"])

    selected_salary = None
    existing_slip = None
    existing_payment = None
    advance_summary = _advance_summary(db, eid)
    available_advance = advance_summary["remaining_advance"]

    values = {
        "deduction_amount": "0.00",
        "payment_date": date.today().isoformat(),
        "actual_paid_amount": "",
        "payment_source": PAYMENT_SOURCES[0],
        "paid_by": "",
        "payment_notes": "",
        "deduction_mode": "manual",
        "selected_txn_ids": [],
    }

    if selected_salary_id:
        selected_salary = db.execute(
            "SELECT * FROM salary_store WHERE id = ? AND driver_id = ?",
            (selected_salary_id, eid),
        ).fetchone()
        if selected_salary is not None:
            existing_slip = db.execute(
                """
                SELECT * FROM salary_slips
                WHERE salary_store_id = ? AND driver_id = ?
                ORDER BY id DESC LIMIT 1
                """,
                (selected_salary_id, eid),
            ).fetchone()
            available_advance = _advance_summary(
                db, eid, exclude_salary_store_id=int(selected_salary_id),
            )["remaining_advance"]
            if existing_slip:
                values["deduction_amount"] = f"{float(existing_slip['total_deductions']):.2f}"

    # Load driver transactions for selection
    driver_txns = db.execute(
        "SELECT * FROM driver_transactions WHERE driver_id = ? ORDER BY entry_date ASC",
        (eid,),
    ).fetchall()

    # For each transaction, compute how much has been deducted already
    deducted_txn_ids = set()
    for txn in driver_txns:
        already = float(db.execute(
            "SELECT COALESCE(SUM(amount_deducted), 0) FROM salary_slip_deductions WHERE driver_transaction_id = ?",
            (txn["id"],),
        ).fetchone()[0])
        remaining = float(txn["amount"]) - already
        txn["_already_deducted"] = already
        txn["_remaining"] = max(remaining, 0.0)
        txn["_is_fully_deducted"] = remaining <= 0.001
        if txn["_is_fully_deducted"]:
            deducted_txn_ids.add(txn["id"])

    # If existing slip, pre-select its linked transactions
    selected_txn_ids = set()
    if existing_slip:
        linked = db.execute(
            "SELECT driver_transaction_id FROM salary_slip_deductions WHERE salary_slip_id = ?",
            (existing_slip["id"],),
        ).fetchall()
        selected_txn_ids = {r["driver_transaction_id"] for r in linked}

    if request.method == "POST":
        selected_salary_id = request.form.get("salary_store_id", "").strip()
        deduction_amount = float(request.form.get("deduction_amount", "0").strip() or "0")
        selected_ids_int = []
        values = {
            "deduction_amount": f"{deduction_amount:.2f}",
            "payment_date": request.form.get("payment_date", date.today().isoformat()).strip() or date.today().isoformat(),
            "actual_paid_amount": request.form.get("actual_paid_amount", "").strip(),
            "payment_source": request.form.get("payment_source", PAYMENT_SOURCES[0]).strip() or PAYMENT_SOURCES[0],
            "paid_by": request.form.get("paid_by", "").strip(),
            "payment_notes": request.form.get("payment_notes", "").strip(),
            "deduction_mode": "manual",
            "selected_txn_ids": [],
        }

        if not selected_salary_id:
            flash("Select a stored salary month first.", "error")
        else:
            selected_salary = db.execute(
                "SELECT * FROM salary_store WHERE id = ? AND driver_id = ?",
                (selected_salary_id, eid),
            ).fetchone()
            existing_slip = db.execute(
                """
                SELECT * FROM salary_slips
                WHERE salary_store_id = ? AND driver_id = ?
                ORDER BY id DESC LIMIT 1
                """,
                (selected_salary_id, eid),
            ).fetchone()
            available_advance = _advance_summary(
                db, eid, exclude_salary_store_id=int(selected_salary_id),
            )["remaining_advance"]

            try:
                deduction_amount = _parse_decimal(values["deduction_amount"], "Deduction", required=False, default=0.0, minimum=0.0)
            except ValidationError as exc:
                flash(str(exc), "error")
                deduction_amount = None

            if deduction_amount is not None and selected_salary is not None:
                if deduction_amount < 0 or deduction_amount > available_advance + 0.001:
                    flash(f"Deduction must be between 0 and {available_advance:,.2f}.", "error")
                else:
                    salary_after_deduction = float(selected_salary["net_salary"]) - deduction_amount
                    if salary_after_deduction < 0:
                        flash("Deduction cannot exceed salary amount.", "error")
                    else:
                        remaining_advance = max(available_advance - deduction_amount, 0.0)
                        from ..pdf_service import generate_salary_slip_pdf
                        company = db.execute("SELECT * FROM company_profile LIMIT 1").fetchone()
                        company_profile = dict(company) if company else None

                        if existing_slip is not None:
                            db.execute(
                                """
                                UPDATE salary_slips
                                SET total_deductions=?, available_advance=?, remaining_advance=?,
                                    salary_after_deduction=?, net_payable=?,
                                    generated_at=CURRENT_TIMESTAMP
                                WHERE id=? AND driver_id=?
                                """,
                                (deduction_amount, available_advance, remaining_advance,
                                 salary_after_deduction, salary_after_deduction,
                                 existing_slip["id"], eid),
                            )
                            slip_id = existing_slip["id"]
                        else:
                            result = db.execute(
                                """
                                INSERT INTO salary_slips (
                                    driver_id, salary_store_id, salary_month, source_filter,
                                    total_deductions, available_advance, remaining_advance,
                                    salary_after_deduction, actual_paid_amount,
                                    company_balance_due, payment_source, paid_by, net_payable, pdf_path
                                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                                """,
                                (eid, selected_salary["id"], selected_salary["salary_month"], "",
                                 deduction_amount, available_advance, remaining_advance,
                                 salary_after_deduction, 0.0, salary_after_deduction,
                                 values["payment_source"], values["paid_by"] or None,
                                 salary_after_deduction, ""),
                            )
                            slip_id = result.lastrowid

                        # Save linked transactions to salary_slip_deductions
                        db.execute("DELETE FROM salary_slip_deductions WHERE salary_slip_id = ?", (slip_id,))
                        if selected_ids_int:
                            for txn_id in selected_ids_int:
                                txn = db.execute("SELECT * FROM driver_transactions WHERE id = ? AND driver_id = ?", (txn_id, eid)).fetchone()
                                if txn:
                                    already = float(db.execute(
                                        "SELECT COALESCE(SUM(amount_deducted), 0) FROM salary_slip_deductions WHERE driver_transaction_id = ?",
                                        (txn["id"],),
                                    ).fetchone()[0])
                                    remaining = max(float(txn["amount"]) - already, 0.0)
                                    if remaining > 0.001:
                                        db.execute(
                                            "INSERT INTO salary_slip_deductions (salary_slip_id, driver_transaction_id, amount_deducted) VALUES (?,?,?)",
                                            (slip_id, txn["id"], remaining),
                                        )

                        slip_row = db.execute("SELECT * FROM salary_slips WHERE id = ?", (slip_id,)).fetchone()
                        driver_display = {"driver_id": eid, "full_name": employee["full_name"],
                                          "basic_salary": employee["basic_salary"] or 0,
                                          "vehicle_no": "", "shift": employee.get("shift", ""),
                                          "duty_start": employee.get("join_date", "")}

                        dr = db.execute("SELECT vehicle_no, photo_name, photo_data, photo_content_type FROM drivers WHERE driver_id=?", (eid,)).fetchone()
                        if dr:
                            driver_display["vehicle_no"] = dr["vehicle_no"] or ""
                            driver_display["photo_data"] = dr["photo_data"] or ""
                            driver_display["photo_name"] = dr["photo_name"] or ""

                        generated_dir = current_app.config["GENERATED_DIR"]
                        slip_output_dir = str(Path(generated_dir) / "salary_slips")
                        _ap = float(salary_after_deduction)
                        pdf_path = generate_salary_slip_pdf(
                            driver_display,
                            selected_salary,
                            {
                                "available_advance": float(available_advance),
                                "deduction_amount": float(deduction_amount),
                                "remaining_advance": float(remaining_advance),
                                "salary_after_deduction": float(salary_after_deduction),
                                "actual_paid_amount": _ap,
                                "company_balance_due": max(float(salary_after_deduction) - _ap, 0.0),
                                "payment_source": values["payment_source"],
                                "paid_by": values["paid_by"] or "",
                                "net_payable": float(salary_after_deduction),
                                "_vehicle_no": (dr["vehicle_no"] or "") if dr else "",
                                "_photo_name": (dr["photo_name"] or "") if dr else "",
                                "_photo_data": (dr["photo_data"] or "") if dr else "",
                            },
                            slip_output_dir,
                            current_app.config["STATIC_ASSETS_DIR"],
                            generated_dir,
                            company_profile=company_profile,
                        )
                        relative_pdf = Path(pdf_path).relative_to(current_app.config["GENERATED_DIR"]).as_posix() if pdf_path else ""
                        db.execute("UPDATE salary_slips SET pdf_path=? WHERE id=?", (relative_pdf, slip_id))
                        _audit_log(db, "employee_salary_slip_generated", entity_type="salary_slip",
                                   entity_id=f"{eid}:{selected_salary['salary_month']}",
                                   details=f"Deduction AED {deduction_amount:.2f} / Payable AED {salary_after_deduction:.2f}")
                        db.commit()
                        flash(f"Salary slip generated for {selected_salary['salary_month']}.", "success")
                        return redirect(url_for("hr.employee_salary_slip", employee_id=eid))

    photo_url = _employee_photo_url(current_app._get_current_object(), employee)

    return render_template(
        "hr/employee_detail.html",
        employee=employee,
        photo_url=photo_url,
        active_tab="salary_slip",
        slip_salary_rows=salary_rows,
        slip_selected_salary=selected_salary,
        slip_existing_slip=existing_slip,
        slip_values=values,
        slip_available_advance=available_advance,
        slip_advance_summary=advance_summary,
        slip_store_ids=slip_store_ids,
        slip_driver_txns=driver_txns,
        slip_selected_txn_ids=selected_txn_ids,
    )


@hr_bp.route("/hr/employees/<employee_id>/salary-slip/<int:store_id>/delete", methods=["GET"])
@_login_required("admin")
def employee_salary_slip_delete(employee_id, store_id):
    _touch_admin_workspace("hr")
    ensure_employees_table()
    db = open_db()
    employee = _fetch_employee(db, employee_id)
    if employee is None:
        flash("Employee not found.", "error")
        return redirect(url_for("hr.employee_list"))
    eid = employee["employee_id"]
    slip = db.execute(
        "SELECT * FROM salary_slips WHERE salary_store_id = ? AND driver_id = ?",
        (store_id, eid),
    ).fetchone()
    if slip is None:
        flash("Salary slip not found.", "error")
    else:
        db.execute("DELETE FROM salary_slip_deductions WHERE salary_slip_id = ?", (slip["id"],))
        db.execute("DELETE FROM salary_slips WHERE id = ?", (slip["id"],))
        _audit_log(db, "employee_salary_slip_deleted", entity_type="salary_slip",
                    entity_id=f"{eid}:{slip['salary_month']}")
        db.commit()
        flash(f"Salary slip for {slip['salary_month']} deleted.", "success")
    return redirect(url_for("hr.employee_salary_slip", employee_id=eid, salary_store_id=store_id))


@hr_bp.route("/hr/employees/<employee_id>/salary-slip/<int:store_id>/deduction-statement")
@_login_required("admin")
def employee_deduction_statement(employee_id, store_id):
    _touch_admin_workspace("hr")
    ensure_employees_table()
    db = open_db()
    employee = _fetch_employee(db, employee_id)
    if employee is None:
        flash("Employee not found.", "error")
        return redirect(url_for("hr.employee_list"))
    eid = employee["employee_id"]
    salary_store_row = db.execute("SELECT * FROM salary_store WHERE id = ? AND driver_id = ?", (store_id, eid)).fetchone()
    if not salary_store_row:
        flash("Salary store not found.", "error")
        return redirect(url_for("hr.employee_salary_slip", employee_id=eid))
    slip = db.execute("SELECT * FROM salary_slips WHERE salary_store_id = ? AND driver_id = ? ORDER BY id DESC LIMIT 1", (store_id, eid)).fetchone()
    if not slip:
        flash("Salary slip not found. Generate slip first.", "error")
        return redirect(url_for("hr.employee_salary_slip", employee_id=eid, salary_store_id=store_id))
    driver_display = {"driver_id": eid, "full_name": employee["full_name"], "basic_salary": employee["basic_salary"] or 0}
    company = db.execute("SELECT * FROM company_profile LIMIT 1").fetchone()
    company_profile = dict(company) if company else None
    # Get deducted transactions
    deducted = db.execute(
        """SELECT dt.*, sd.amount_deducted
           FROM salary_slip_deductions sd
           JOIN driver_transactions dt ON dt.id = sd.driver_transaction_id
           WHERE sd.salary_slip_id = ?""",
        (slip["id"],),
    ).fetchall()
    from ..pdf_service import generate_deduction_statement_pdf
    pdf_path = generate_deduction_statement_pdf(
        driver_display, salary_store_row, slip, deducted,
        str(Path(current_app.config["GENERATED_DIR"]) / "salary_slips"),
        current_app.config["STATIC_ASSETS_DIR"],
        company_profile=company_profile,
    )
    if pdf_path:
        rel = Path(pdf_path).relative_to(current_app.config["GENERATED_DIR"]).as_posix()
        return redirect(url_for("generated_file", filename=rel))
    flash("Could not generate deduction statement.", "error")
    return redirect(url_for("hr.employee_salary_slip", employee_id=eid, salary_store_id=store_id))


# ── Employee Kata ────────────────────────────────────────────────

@hr_bp.route("/hr/employees/<employee_id>/kata")
@_login_required("admin")
def employee_kata(employee_id):
    _touch_admin_workspace("hr")
    ensure_employees_table()
    db = open_db()

    employee = _fetch_employee(db, employee_id)
    if employee is None:
        flash("Employee not found.", "error")
        return redirect(url_for("hr.employee_list"))

    eid = employee["employee_id"]

    available_months = db.execute(
        """
        SELECT DISTINCT salary_month FROM (
            SELECT salary_month FROM salary_store WHERE driver_id = ?
            UNION
            SELECT salary_month FROM salary_slips WHERE driver_id = ?
            UNION
            SELECT COALESCE(salary_month, SUBSTR(entry_date, 1, 7)) AS salary_month
            FROM driver_transactions WHERE driver_id = ?
        ) ORDER BY salary_month DESC
        """,
        (eid, eid, eid),
    ).fetchall()

    selected_month = request.args.get("month", "").strip()
    if not selected_month and available_months:
        selected_month = available_months[0]["salary_month"]

    entries = []
    summary = {}
    salary_row = None
    slip_row = None
    kata_advances = []
    kata_prev_remaining = 0.0
    kata_this_deduction = 0.0
    kata_remaining = 0.0
    deduction_history = []

    if selected_month:
        active_entries, closed_entries, summary_ret = _driver_kata_month_data(db, eid, selected_month)
        entries = active_entries + closed_entries
        summary = summary_ret

        salary_row = db.execute(
            "SELECT * FROM salary_store WHERE driver_id = ? AND salary_month = ?",
            (eid, selected_month),
        ).fetchone()

        slip_row = db.execute(
            "SELECT * FROM salary_slips WHERE driver_id = ? AND salary_month = ? ORDER BY id DESC LIMIT 1",
            (eid, selected_month),
        ).fetchone()

        all_advances = db.execute(
            "SELECT * FROM driver_transactions WHERE driver_id = ? ORDER BY entry_date ASC, id ASC",
            (eid,),
        ).fetchall()

        total_advance_amount = sum(float(r["amount"]) for r in all_advances)

        prev_deductions = float(db.execute(
            "SELECT COALESCE(SUM(total_deductions), 0) FROM salary_slips WHERE driver_id = ? AND salary_month < ?",
            (eid, selected_month),
        ).fetchone()[0])

        this_deduction = float(slip_row["total_deductions"]) if slip_row else 0.0

        remaining_before = max(total_advance_amount - prev_deductions, 0.0)
        remaining_after = max(remaining_before - this_deduction, 0.0)

        kata_prev_remaining = remaining_before
        kata_this_deduction = this_deduction
        kata_remaining = remaining_after

        # Build actual deduction data per transaction from salary_slip_deductions
        txn_deducted_map = {}
        deducted_rows = db.execute(
            """
            SELECT sd.driver_transaction_id, sd.amount_deducted, ss.salary_month
            FROM salary_slip_deductions sd
            JOIN salary_slips ss ON ss.id = sd.salary_slip_id
            WHERE ss.driver_id = ?
            """,
            (eid,),
        ).fetchall()
        for dr in deducted_rows:
            txn_id = dr["driver_transaction_id"]
            if txn_id not in txn_deducted_map:
                txn_deducted_map[txn_id] = {"total": 0.0, "slips": []}
            txn_deducted_map[txn_id]["total"] += float(dr["amount_deducted"])
            txn_deducted_map[txn_id]["slips"].append(dr["salary_month"])

        # Compute remaining deduction (from old-style lump sums not tracked per transaction)
        old_style_deductions = max(prev_deductions + this_deduction - sum(d["total"] for d in txn_deducted_map.values()), 0.0)
        remaining_deduction = old_style_deductions

        for a in all_advances:
            amt = float(a["amount"])
            txn_id = a["id"]
            already_deducted = txn_deducted_map.get(txn_id, {}).get("total", 0.0)
            if already_deducted >= amt - 0.001:
                # Fully deducted via actual tracking
                kata_advances.append({
                    "entry_date": a["entry_date"],
                    "amount": amt,
                    "source": a["source"],
                    "given_by": a["given_by"],
                    "details": a["details"],
                    "remaining": 0.0,
                    "deducted": amt,
                    "status": "cleared",
                    "deducted_in": txn_deducted_map.get(txn_id, {}).get("slips", []),
                })
            else:
                not_deducted_yet = max(amt - already_deducted, 0.0)
                # Apply remaining old-style deduction FIFO
                if remaining_deduction <= 0:
                    ded = already_deducted
                    rem = not_deducted_yet
                    status = "cleared" if rem <= 0.001 else "outstanding"
                elif remaining_deduction >= not_deducted_yet:
                    ded = amt
                    rem = 0.0
                    status = "cleared"
                    remaining_deduction -= not_deducted_yet
                else:
                    ded = already_deducted + remaining_deduction
                    rem = not_deducted_yet - remaining_deduction
                    status = "cleared" if rem <= 0.001 else "partial" if remaining_deduction > 0 else "outstanding"
                    remaining_deduction = 0.0
                kata_advances.append({
                    "entry_date": a["entry_date"],
                    "amount": amt,
                    "source": a["source"],
                    "given_by": a["given_by"],
                    "details": a["details"],
                    "remaining": max(amt - ded, 0.0),
                    "deducted": ded,
                    "status": status,
                    "deducted_in": txn_deducted_map.get(txn_id, {}).get("slips", []),
                })

        # Deduction history for display
        deduction_history = db.execute(
            """
            SELECT sd.*, dt.amount as txn_amount, dt.details as txn_details, dt.entry_date as txn_date,
                   ss.salary_month, ss.salary_store_id
            FROM salary_slip_deductions sd
            JOIN driver_transactions dt ON dt.id = sd.driver_transaction_id
            JOIN salary_slips ss ON ss.id = sd.salary_slip_id
            WHERE ss.driver_id = ?
            ORDER BY ss.salary_month DESC, sd.created_at DESC
            """,
            (eid,),
        ).fetchall()

    pdf_url = None
    transactions_pdf_url = None
    if selected_month:
        from ..pdf_service import generate_simple_kata_pdf, generate_transactions_kata_pdf

        driver_display = {
            "driver_id": eid,
            "full_name": employee["full_name"],
            "basic_salary": employee["basic_salary"] or 0,
        }
        company = db.execute("SELECT * FROM company_profile LIMIT 1").fetchone()
        company_profile = dict(company) if company else None

        # All salary store rows that don't have a generated slip yet
        unpaid_salary_rows = db.execute(
            "SELECT ss.* FROM salary_store ss WHERE ss.driver_id = ? AND ss.id NOT IN (SELECT DISTINCT salary_store_id FROM salary_slips WHERE driver_id = ?) ORDER BY ss.salary_month ASC",
            (eid, eid),
        ).fetchall()

        pdf_path = generate_simple_kata_pdf(
            driver_display, salary_row, unpaid_salary_rows, kata_advances,
            kata_prev_remaining, kata_this_deduction, kata_remaining,
            selected_month,
            str(Path(current_app.config["GENERATED_DIR"]) / "kata_pdfs"),
            current_app.config["STATIC_ASSETS_DIR"],
            company_profile=company_profile,
        )
        if pdf_path:
            rel = Path(pdf_path).relative_to(current_app.config["GENERATED_DIR"]).as_posix()
            pdf_url = url_for("generated_file", filename=rel)

        outstanding = [a for a in kata_advances if a["status"] != "cleared"]
        tx_path = generate_transactions_kata_pdf(
            driver_display, outstanding, selected_month,
            str(Path(current_app.config["GENERATED_DIR"]) / "kata_pdfs"),
            current_app.config["STATIC_ASSETS_DIR"],
            company_profile=company_profile,
        )
        if tx_path:
            rel = Path(tx_path).relative_to(current_app.config["GENERATED_DIR"]).as_posix()
            transactions_pdf_url = url_for("generated_file", filename=rel)

    photo_url = _employee_photo_url(current_app._get_current_object(), employee)

    return render_template(
        "hr/employee_detail.html",
        employee=employee,
        photo_url=photo_url,
        active_tab="kata",
        kata_available_months=available_months,
        kata_selected_month=selected_month,
        kata_entries=entries,
        kata_summary=summary,
        kata_salary_row=salary_row,
        kata_slip_row=slip_row,
        kata_advances=kata_advances,
        kata_prev_remaining=kata_prev_remaining,
        kata_this_deduction=kata_this_deduction,
        kata_remaining=kata_remaining,
        kata_deduction_history=deduction_history,
        kata_pdf_url=pdf_url,
        kata_transactions_pdf_url=transactions_pdf_url,
    )


# ── Employee Edit ────────────────────────────────────────────────

@hr_bp.route("/hr/employees/<employee_id>/edit", methods=["GET", "POST"])
@_login_required("admin")
def employee_edit(employee_id):
    _touch_admin_workspace("hr")
    ensure_employees_table()
    db = open_db()

    employee_id = employee_id.strip().upper()
    employee = _fetch_employee(db, employee_id)
    if employee is None:
        flash("Employee not found.", "error")
        return redirect(url_for("hr.employee_list"))

    # Get vehicles if table exists, otherwise empty list
    try:
        vehicles = db.execute("SELECT plate_no, vehicle_type, model FROM vehicles WHERE status = 'Active' ORDER BY plate_no").fetchall()
    except Exception:
        vehicles = []
    assigned = db.execute("SELECT vehicle_id FROM vehicle_assignments WHERE driver_id = ? AND is_current = 1 LIMIT 1", (employee_id,)).fetchone()
    assigned_vehicle = assigned["vehicle_id"] if assigned else ""
    if not assigned_vehicle:
        legacy = db.execute("SELECT vehicle_no FROM drivers WHERE driver_id = ? LIMIT 1", (employee_id,)).fetchone()
        if legacy and legacy["vehicle_no"]:
            assigned_vehicle = legacy["vehicle_no"]

    if request.method == "POST":
        values = employee_form_data()
        values["employee_id"] = employee_id

        # Upload photo FIRST so it survives validation errors
        uploaded_photo = None
        try:
            uploaded_photo = save_employee_photo(
                current_app._get_current_object(), employee_id,
                request.form.get("full_name", "").strip().title(),
                request.files.get("photo_file"))
        except Exception as exc:
            flash(f"Photo upload failed: {exc}", "error")

        errors = validate_employee_form(values)
        if errors:
            for err in errors:
                flash(err, "error")
            assigned_vehicle = values.get("vehicle_id", "")
            # Save photo to DB even if other fields fail validation
            if uploaded_photo:
                try:
                    db.execute(
                        "UPDATE employees SET photo_name=?, photo_data=?, photo_content_type=?, updated_at=CURRENT_TIMESTAMP WHERE UPPER(employee_id)=?",
                        (uploaded_photo["photo_name"], uploaded_photo["photo_data"],
                         uploaded_photo["photo_content_type"], employee_id),
                    )
                    db.commit()
                except Exception:
                    db.rollback()
        else:
            salary = float(values["basic_salary"])
            ot_rate = float(values.get("ot_rate", 0) or 0)
            if values.get("status", "").lower() == "terminated":
                if not values.get("termination_date"):
                    values["termination_date"] = date.today().isoformat()
            else:
                values["termination_date"] = ""

            try:
                db.execute(
                    """
                    UPDATE employees SET
                        full_name=?, phone_number=?, email=?,
                        employee_type=?, department=?, designation=?, gender=?,
                        shift=?, contract_type=?, join_date=?,
                        basic_salary=?, ot_rate=?,
                        nationality=?, iqama_no=?, passport_no=?,
                        bank_name=?, bank_account=?, iban=?,
                        emergency_contact=?, emergency_name=?, address=?,
                        photo_name=COALESCE(?, photo_name),
                        photo_data=COALESCE(?, photo_data),
                        photo_content_type=COALESCE(?, photo_content_type),
                        status=?, termination_date=?, remarks=?,
                        updated_at=CURRENT_TIMESTAMP
                    WHERE UPPER(employee_id)=?
                    """,
                    (
                        values["full_name"], values["phone_number"], values["email"] or None,
                        values["employee_type"], values["department"], values["designation"], values["gender"] or None,
                        values["shift"] or "Morning", values["contract_type"] or "Permanent",
                        values["join_date"], salary, ot_rate,
                        values["nationality"] or None, values["iqama_no"] or None, values["passport_no"] or None,
                        values["bank_name"] or None, values["bank_account"] or None, values["iban"] or None,
                        values["emergency_contact"] or None, values["emergency_name"] or None, values["address"] or None,
                        uploaded_photo["photo_name"] if uploaded_photo else None,
                        uploaded_photo["photo_data"] if uploaded_photo else None,
                        uploaded_photo["photo_content_type"] if uploaded_photo else None,
                        values["status"], values["termination_date"] or None, values["remarks"] or None,
                        employee_id,
                    ),
                )
            except Exception as exc:
                flash(f"Database error updating employee: {exc}", "error")
                db.rollback()
                return redirect(url_for("hr.employee_edit", employee_id=employee_id))

            # Always sync status/termination to drivers table so sync_drivers_to_employees doesn't revert it
            try:
                db.execute(
                    "UPDATE drivers SET status=?, termination_date=?, basic_salary=?, ot_rate=?, shift=?, full_name=?, phone_number=? WHERE UPPER(driver_id)=?",
                    (values["status"], values["termination_date"] or None, salary, ot_rate,
                     values["shift"] or "Morning", values["full_name"], values["phone_number"] or None, employee_id),
                )
            except Exception:
                pass  # drivers table may not exist

            if values.get("employee_type") == "Driver":
                if values["vehicle_id"]:
                    db.execute(
                        "UPDATE vehicle_assignments SET is_current = 0, assigned_until = ? WHERE driver_id = ? AND is_current = 1",
                        (date.today().isoformat(), employee_id),
                    )
                    existing = db.execute(
                        "SELECT id FROM vehicle_assignments WHERE vehicle_id = ? AND driver_id = ? AND is_current = 1",
                        (values["vehicle_id"], employee_id),
                    ).fetchone()
                    if not existing:
                        db.execute(
                            "INSERT INTO vehicle_assignments (vehicle_id, driver_id, assigned_from, is_current) VALUES (?, ?, ?, 1)",
                            (values["vehicle_id"], employee_id, date.today().isoformat()),
                        )
                    db.execute("UPDATE drivers SET vehicle_no = ? WHERE driver_id = ?", (values["vehicle_id"], employee_id))
                else:
                    db.execute(
                        "UPDATE vehicle_assignments SET is_current = 0, assigned_until = ? WHERE driver_id = ? AND is_current = 1",
                        (date.today().isoformat(), employee_id),
                    )
                    db.execute("UPDATE drivers SET vehicle_no = NULL WHERE driver_id = ?", (employee_id,))

            if values.get("employee_type") == "Field Staff":
                _sync_employee_to_field_staff(db, employee_id, values["full_name"], values["phone_number"], values["status"])
            else:
                try:
                    db.execute("DELETE FROM field_staff WHERE staff_id = ?", (employee_id,))
                except Exception:
                    pass

            _audit_log(db, "employee_updated", entity_type="employee", entity_id=employee_id, details=f"{values['full_name']} updated")
            db.commit()

            # Verify the update took effect
            check = _fetch_employee(db, employee_id)
            current_app.logger.info(f"Employee {employee_id} updated: status={check['status'] if check else 'NOT FOUND'}")
            flash("Employee updated successfully.", "success")
            return redirect(url_for("hr.employee_detail", employee_id=employee_id))

    values = dict(employee)
    for key in ("basic_salary", "ot_rate"):
        try:
            values[key] = f"{float(values[key] or 0):.2f}" if values.get(key) else "0"
        except (ValueError, TypeError):
            values[key] = "0"

    return render_template(
        "hr/employee_form.html",
        values=values,
        page_title="Edit Employee",
        submit_label="Update Employee",
        edit_mode=True,
        employee_types=EMPLOYEE_TYPES,
        departments=DEPARTMENTS,
        designations=DESIGNATIONS,
        status_options=STATUS_OPTIONS,
        gender_options=GENDER_OPTIONS,
        shift_options=SHIFT_OPTIONS,
        contract_options=CONTRACT_TYPE_OPTIONS,
        vehicles=vehicles,
        assigned_vehicle=assigned_vehicle,
    )


@hr_bp.route("/hr/employees/<employee_id>/delete", methods=["POST"])
@_login_required("admin")
def employee_delete(employee_id):
    _touch_admin_workspace("hr")
    ensure_employees_table()
    db = open_db()
    employee_id = employee_id.strip().upper()

    employee = _fetch_employee(db, employee_id)
    if employee is None:
        flash("Employee not found.", "error")
        return redirect(url_for("hr.employee_list"))

    try:
        # Clean up related data
        db.execute("DELETE FROM salary_store WHERE driver_id = ?", (employee_id,))
        db.execute("DELETE FROM salary_slips WHERE driver_id = ?", (employee_id,))
        db.execute("DELETE FROM salary_payments WHERE driver_id = ?", (employee_id,))
        db.execute("DELETE FROM driver_transactions WHERE driver_id = ?", (employee_id,))
        db.execute("DELETE FROM vehicle_assignments WHERE driver_id = ?", (employee_id,))
        db.execute("DELETE FROM drivers WHERE driver_id = ?", (employee_id,))
        db.execute("DELETE FROM cash_receipts WHERE staff_id = ?", (employee_id,))
        db.execute("DELETE FROM field_staff WHERE staff_id = ?", (employee_id,))
        db.execute("DELETE FROM employees WHERE employee_id = ?", (employee_id,))
        db.execute("DELETE FROM audit_logs WHERE entity_id = ? AND entity_type = 'employee'", (employee_id,))
        db.commit()
        _audit_log(db, "employee_deleted", entity_type="employee", entity_id=employee_id, details=f"{employee['full_name']} deleted")
        flash(f"Employee {employee['full_name']} ({employee_id}) deleted permanently.", "success")
    except Exception as e:
        db.rollback()
        flash(f"Error deleting employee: {e}", "error")

    return redirect(url_for("hr.employee_list"))


@hr_bp.route("/hr/employees/<employee_id>/restore", methods=["GET", "POST"])
@_login_required("admin")
def employee_restore(employee_id):
    import traceback
    try:
        _touch_admin_workspace("hr")
        ensure_employees_table()
        db = open_db()
        employee_id = employee_id.strip().upper()

        existing = _fetch_employee(db, employee_id)
        if existing:
            flash(f"Employee {employee_id} already exists.", "success")
            return redirect(url_for("hr.employee_edit", employee_id=employee_id))

        tech = db.execute("SELECT * FROM technicians WHERE technician_code = ?", (employee_id,)).fetchone()
        if not tech:
            flash(f"No data found for {employee_id} in technicians table.", "error")
            return redirect(url_for("hr.employee_list"))

        name = tech["specialization"] or f"Staff {employee_id}"
        phone = tech["phone_number"] or ""
        username = tech["user_id"] or employee_id.lower()
        pw_hash = tech["password_hash"] or generate_password_hash("changeme123")
        is_active = 1 if tech.get("status") == "Active" else 0
        raw_created = tech.get("created_at")
        if isinstance(raw_created, (datetime, date)):
            join_date = raw_created.strftime("%Y-%m-%d")
        elif raw_created:
            join_date = str(raw_created).split(" ")[0]
        else:
            join_date = date.today().isoformat()

        try:
            db.execute("""
                INSERT INTO field_staff (staff_id, full_name, phone, username, password_hash, is_active)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (employee_id, name, phone, username, pw_hash, is_active))
        except Exception as e:
            current_app.logger.warning("restore field_staff: %s", e)

        try:
            db.execute("""
                INSERT INTO employees (employee_id, full_name, phone_number, employee_type, department, designation, status, join_date, basic_salary, ot_rate)
                VALUES (?, ?, ?, 'Field Staff', 'Field Staff', 'Field Staff', ?, ?, 0, 0)
            """, (employee_id, name, phone, "Active" if is_active else "Inactive", join_date))
        except Exception as e:
            current_app.logger.warning("restore employees: %s", e)
            db.rollback()
            flash(f"Restore failed (employees): {e}", "error")
            return redirect(url_for("hr.employee_list"))

        db.commit()
        flash(f"Employee {name} ({employee_id}) restored successfully.", "success")
        return redirect(url_for("hr.employee_edit", employee_id=employee_id))
    except Exception as e:
        current_app.logger.error("restore error: %s\n%s", e, traceback.format_exc())
        flash(f"Restore error: {e}", "error")
        return redirect(url_for("hr.employee_list"))


@hr_bp.route("/hr/employees/download/excel")
@_login_required("admin")
def employee_list_excel():
    _touch_admin_workspace("hr")
    ensure_employees_table()
    db = open_db()

    query = request.args.get("q", "").strip()
    status_filter = request.args.get("status", "").strip()
    department_filter = request.args.get("department", "").strip()
    employee_type_filter = request.args.get("type", "").strip()

    where_sql, params = employee_search_filter(query, status_filter, department_filter, employee_type_filter)

    employees = db.execute(
        f"""
        SELECT e.employee_id, e.full_name, e.phone_number, e.email, e.employee_type,
               e.department, e.designation, e.join_date, e.basic_salary, e.status, e.termination_date,
               COALESCE(
                   (SELECT v.plate_no FROM vehicle_assignments va
                    JOIN vehicles v ON v.plate_no = va.vehicle_id
                    WHERE va.driver_id = e.employee_id
                    ORDER BY va.id DESC LIMIT 1),
                   (SELECT d.vehicle_no FROM drivers d WHERE d.driver_id = e.employee_id LIMIT 1)
               ) AS plate_no
        FROM employees e
        {where_sql}
        ORDER BY e.full_name ASC
        """,
        params,
    ).fetchall()

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO

    wb = Workbook()
    ws = wb.active
    ws.title = "Employees"

    hf = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    hfill = PatternFill("solid", fgColor="1a3a5c")
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="d8e4f5")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)

    heads = ["Employee ID", "Full Name", "Phone", "Email", "Type", "Department", "Designation", "Join Date", "Salary (AED)", "Vehicle", "Status", "Termination Date"]
    for ci, h in enumerate(heads, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = hf; c.fill = hfill; c.alignment = center; c.border = border

    for ri, emp in enumerate(employees, 2):
        vals = [emp["employee_id"], emp["full_name"], emp["phone_number"] or "", emp["email"] or "",
                emp["employee_type"], emp["department"], emp["designation"],
                emp["join_date"], emp["basic_salary"] or 0, emp["plate_no"] or "", emp["status"] or "",
                emp["termination_date"] or ""]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.border = border
            if ci == 9:
                c.number_format = '#,##0.00'
                c.alignment = Alignment(horizontal="right", vertical="center")

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 26
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 16
    ws.column_dimensions["G"].width = 16
    ws.column_dimensions["H"].width = 14
    ws.column_dimensions["I"].width = 16
    ws.column_dimensions["J"].width = 14
    ws.column_dimensions["K"].width = 12

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name="employees.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@hr_bp.route("/hr/employees/download/pdf")
@_login_required("admin")
def employee_list_pdf():
    _touch_admin_workspace("hr")
    ensure_employees_table()
    db = open_db()

    query = request.args.get("q", "").strip()
    status_filter = request.args.get("status", "").strip()
    department_filter = request.args.get("department", "").strip()
    employee_type_filter = request.args.get("type", "").strip()

    where_sql, params = employee_search_filter(query, status_filter, department_filter, employee_type_filter)

    employees = db.execute(
        f"""
        SELECT e.employee_id, e.full_name, e.phone_number, e.email, e.employee_type,
               e.department, e.designation, e.join_date, e.basic_salary, e.status, e.termination_date
        FROM employees e
        {where_sql}
        ORDER BY e.full_name ASC
        """,
        params,
    ).fetchall()

    try:
        company = db.execute("SELECT * FROM company_profile LIMIT 1").fetchone()
        company_profile = dict(company) if company else None
    except Exception:
        company_profile = None

    from ..pdf_service import _generate_employee_list_pdf
    from pathlib import Path
    from flask import current_app

    output_dir = Path(current_app.config["GENERATED_DIR"]) / "temp"
    pdf_path = _generate_employee_list_pdf(employees, str(output_dir), company_profile)
    return send_file(pdf_path, as_attachment=True, download_name="employees.pdf", mimetype="application/pdf")


@hr_bp.app_template_global()
def employee_current_vehicle(employee_id):
    from ..database import open_db
    try:
        d = open_db()
        row = d.execute(
            "SELECT va.vehicle_id, va.assigned_from, v.plate_no, v.vehicle_type, v.model "
            "FROM vehicle_assignments va JOIN vehicles v ON v.plate_no = va.vehicle_id "
            "WHERE va.driver_id = ? AND va.is_current = 1 LIMIT 1",
            (employee_id,),
        ).fetchone()
        if not row:
            row = d.execute(
                "SELECT va.vehicle_id, va.assigned_from, v.plate_no, v.vehicle_type, v.model "
                "FROM vehicle_assignments va JOIN vehicles v ON v.plate_no = va.vehicle_id "
                "WHERE va.driver_id = ? ORDER BY va.id DESC LIMIT 1",
                (employee_id,),
            ).fetchone()
        if not row:
            legacy = d.execute("SELECT vehicle_no FROM drivers WHERE driver_id = ? LIMIT 1", (employee_id,)).fetchone()
            if legacy and legacy["vehicle_no"]:
                row = d.execute("SELECT plate_no AS vehicle_id, NULL AS assigned_from, plate_no, vehicle_type, model FROM vehicles WHERE plate_no = ? LIMIT 1", (legacy["vehicle_no"],)).fetchone()
                if not row:
                    row = {"vehicle_id": legacy["vehicle_no"], "assigned_from": "", "plate_no": legacy["vehicle_no"], "vehicle_type": "", "model": ""}
        d.close()
        return dict(row) if row else None
    except Exception:
        return None

@hr_bp.app_template_global()
def employee_vehicle_history(employee_id):
    from ..database import open_db
    try:
        d = open_db()
        rows = d.execute(
            "SELECT va.vehicle_id, va.assigned_from, va.assigned_until, v.plate_no, v.vehicle_type, v.model FROM vehicle_assignments va JOIN vehicles v ON v.plate_no = va.vehicle_id WHERE va.driver_id = ? ORDER BY va.id DESC",
            (employee_id,),
        ).fetchall()
        d.close()
        return [dict(r) for r in rows]
    except Exception:
        return []

# ── Salary Dashboard ──────────────────────────────────────────────

def _salary_dashboard_data(status_filter=""):
    db = open_db()

    where = ""
    params = []
    if status_filter == "active":
        where = "WHERE LOWER(status) = 'active'"
    elif status_filter == "inactive":
        where = "WHERE LOWER(status) IN ('inactive','on leave')"
    elif status_filter == "terminated":
        where = "WHERE LOWER(status) = 'terminated'"

    employees = db.execute(
        f"SELECT employee_id, full_name, department, status, termination_date FROM employees {where} ORDER BY full_name"
    ).fetchall()

    store_rows = db.execute(
        "SELECT driver_id, salary_month, net_salary, ot_amount, monthly_basic_salary, basic_salary FROM salary_store"
    ).fetchall()
    store_by_emp = {}
    ot_hist = {}
    basic_by_emp = {}
    for r in store_rows:
        store_by_emp.setdefault(r["driver_id"], {})[r["salary_month"]] = r["net_salary"]
        ot_hist.setdefault(r["driver_id"], []).append((r["salary_month"], float(r["ot_amount"] or 0)))
        bs = float(r["monthly_basic_salary"] or r["basic_salary"] or 0)
        if bs > 0:
            basic_by_emp.setdefault(r["driver_id"], []).append((r["salary_month"], bs))
    ot_sort = {}
    for eid, lst in ot_hist.items():
        lst.sort(key=lambda x: x[0], reverse=True)
        ot_sort[eid] = lst
    basic_sort = {}
    for eid, lst in basic_by_emp.items():
        lst.sort(key=lambda x: x[0], reverse=True)
        basic_sort[eid] = lst

    slip_rows = db.execute(
        "SELECT driver_id, salary_month, salary_after_deduction, actual_paid_amount "
        "FROM salary_slips WHERE salary_after_deduction > 0 OR actual_paid_amount > 0"
    ).fetchall()
    paid_by_emp = {}
    for r in slip_rows:
        paid_by_emp.setdefault(r["driver_id"], set()).add(r["salary_month"])

    pay_rows = db.execute(
        "SELECT driver_id, salary_month FROM salary_payments WHERE amount > 0"
    ).fetchall()
    pay_by_emp = {}
    for r in pay_rows:
        pay_by_emp.setdefault(r["driver_id"], set()).add(r["salary_month"])

    month_rows = db.execute(
        "SELECT DISTINCT salary_month FROM salary_store UNION "
        "SELECT DISTINCT salary_month FROM salary_slips UNION "
        "SELECT DISTINCT salary_month FROM salary_payments "
        "ORDER BY salary_month DESC"
    ).fetchall()
    available_months = [r["salary_month"] for r in month_rows]

    emp_list = []
    for emp in employees:
        eid = emp["employee_id"]
        term_date = emp.get("termination_date") or ""
        term_month = term_date[:7] if term_date and len(term_date) >= 7 else ""

        statuses = {}
        amounts = {}
        for m in available_months:
            emp_is_terminated = (emp["status"] or "").lower() == "terminated"
            if emp_is_terminated and term_month and m > term_month:
                statuses[m] = "No Record"
                amounts[m] = 0.0
            else:
                if eid in paid_by_emp and m in paid_by_emp[eid]:
                    statuses[m] = "Paid"
                elif eid in store_by_emp and m in store_by_emp[eid]:
                    statuses[m] = "Unpaid"
                else:
                    statuses[m] = "Pending"
                amounts[m] = float(store_by_emp[eid][m]) if eid in store_by_emp and m in store_by_emp[eid] else 0.0
        ot_history = ot_sort.get(eid, [])
        est = 0.0
        if ot_history:
            recent = [x[1] for x in ot_history[:3] if x[1] > 0]
            if recent:
                est = sum(recent) / len(recent)
        basic_list = basic_sort.get(eid, [])
        basic_val = basic_list[0][1] if basic_list else 0.0
        emp_list.append({
            "id": eid,
            "name": emp["full_name"],
            "department": emp["department"] or "",
            "emp_status": emp["status"] or "",
            "term_date": term_date,
            "statuses": statuses,
            "amounts": amounts,
            "estimated": round(est, 0),
            "basic_salary": basic_val,
        })

    db.close()
    return emp_list, available_months


def _format_month_options(months):
    names = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
    return [(m, f"{names[int(m.split('-')[1])-1]} {m.split('-')[0]}") for m in months]


def _month_name(ym):
    names = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
    return f"{names[int(ym.split('-')[1])-1]} {ym.split('-')[0]}"


@hr_bp.route("/hr/salary-dashboard")
@_login_required("admin")
def salary_dashboard():
    _touch_admin_workspace("hr")
    ensure_employees_table()
    sf = request.args.get("status", "active")
    sel_month_raw = request.args.get("month", "")
    emp_list, available_months = _salary_dashboard_data(status_filter=sf)
    return render_template(
        "hr/employee_salary_dashboard.html",
        employees=emp_list,
        months=_format_month_options(available_months),
        selected_status=sf,
        sel_month=sel_month_raw or (available_months[0] if available_months else ""),
    )


@hr_bp.route("/hr/salary-dashboard/download/excel")
@_login_required("admin")
def salary_dashboard_excel():
    _touch_admin_workspace("hr")
    ensure_employees_table()

    sf = request.args.get("status", "all")
    selected_month = request.args.get("month", "")
    emp_list, available_months = _salary_dashboard_data(status_filter=sf)
    if not selected_month and available_months:
        selected_month = available_months[0]

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from io import BytesIO

    wb = Workbook()
    ws = wb.active
    ws.title = "Salary Status"

    hf = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    hfill = PatternFill("solid", fgColor="1a3a5c")
    hfill_green = PatternFill("solid", fgColor="059669")
    hfill_red = PatternFill("solid", fgColor="DC2626")
    hfill_amber = PatternFill("solid", fgColor="D97706")
    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right", vertical="center")
    thin = Side(style="thin", color="d8e4f5")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)

    heads = ["#", "Employee Name", "Department", "Status", "Basic Salary", "Est. OT (AI)", "Total (Basic+OT)", "Salary Status"]
    for ci, h in enumerate(heads, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = hf; c.fill = hfill; c.alignment = center; c.border = border

    status_fills = {
        "Paid": PatternFill("solid", fgColor="D5F5E3"),
        "Unpaid": PatternFill("solid", fgColor="FADBD8"),
        "Pending": PatternFill("solid", fgColor="FDEBD0"),
        "No Record": PatternFill("solid", fgColor="F2F3F4"),
    }

    row_idx = 2
    for emp in emp_list:
        st = emp["statuses"].get(selected_month, "No Record")
        if st == "No Record":
            continue
        amt = emp["amounts"].get(selected_month, 0)
        est = emp.get("estimated", 0)
        basic = emp.get("basic_salary", 0)
        total = basic + est
        vals = [row_idx - 1, emp["name"], emp["department"], emp["emp_status"], basic, est, total if total > 0 else "", st]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=row_idx, column=ci, value=v)
            c.border = border
            if ci in (5, 6, 7):
                c.number_format = '#,##0.00'
                c.alignment = right
            sf = status_fills.get(st)
            if sf:
                c.fill = sf
        row_idx += 1

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 16
    ws.column_dimensions["G"].width = 18
    ws.column_dimensions["H"].width = 18

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"salary_status_{selected_month}.xlsx" if selected_month else "salary_status.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@hr_bp.app_template_global()
def employee_photo_url(employee):
    if not employee:
        return None
    if employee.get("photo_data") and employee.get("photo_content_type"):
        return f"data:{employee['photo_content_type']};base64,{employee['photo_data']}"
    return None
